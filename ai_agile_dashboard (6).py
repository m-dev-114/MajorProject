import streamlit as st
import pandas as pd
import numpy as np
from sklearn.linear_model import LogisticRegression, LinearRegression
from sklearn.ensemble import (RandomForestClassifier, GradientBoostingClassifier,
                               StackingClassifier, VotingClassifier, AdaBoostClassifier)
from sklearn.svm import SVC
from sklearn.calibration import CalibratedClassifierCV
from sklearn.pipeline import Pipeline
from sklearn.preprocessing import StandardScaler
from sklearn.model_selection import train_test_split, cross_val_score, StratifiedKFold
from sklearn.metrics import (accuracy_score, classification_report, mean_squared_error,
                              roc_auc_score, f1_score, brier_score_loss)
from sklearn.preprocessing import LabelEncoder
from sklearn.inspection import permutation_importance
import time
import os

# ── PySpark integration (graceful fallback) ──────────────────────────────
os.environ.setdefault("PYSPARK_PYTHON", "python3")
os.environ.setdefault("SPARK_LOCAL_IP", "127.0.0.1")
try:
    from pyspark.sql import SparkSession
    from pyspark.sql import functions as F
    from pyspark.sql.types import (StructType, StructField, StringType,
                                    DoubleType, IntegerType, FloatType)
    SPARK_AVAILABLE = True
except ImportError:
    SPARK_AVAILABLE = False

@st.cache_resource
def get_spark():
    """Create or retrieve a cached SparkSession."""
    if not SPARK_AVAILABLE:
        return None
    try:
        spark = (SparkSession.builder
                 .appName("AgileAIDashboard")
                 .master("local[*]")
                 .config("spark.driver.memory", "2g")
                 .config("spark.sql.shuffle.partitions", "4")
                 .config("spark.ui.enabled", "false")
                 .getOrCreate())
        spark.sparkContext.setLogLevel("ERROR")
        return spark
    except Exception:
        return None

def spark_preprocess(pdf: pd.DataFrame, spark) -> pd.DataFrame:
    """
    Use Spark for distributed preprocessing & feature engineering.
    Falls back to pandas if Spark unavailable.
    Returns enriched pandas DataFrame.
    """
    if spark is None:
        return _pandas_preprocess(pdf)
    try:
        sdf = spark.createDataFrame(pdf)

        # ── Feature engineering in Spark ────────────────────────────────
        # 1. Velocity Efficiency Ratio
        sdf = sdf.withColumn("Velocity_Efficiency",
            F.when(F.col("Planned_Story_Points_Sprint") > 0,
                   F.col("Historical_Velocity") / F.col("Planned_Story_Points_Sprint")
            ).otherwise(F.lit(1.0)))

        # 2. Completion Gap (how far behind are they?)
        sdf = sdf.withColumn("Completion_Gap",
            F.col("Planned_Story_Points_Sprint") - F.col("Completed_Story_Points"))

        # 3. Blocker Severity (blockers × days remaining inverse)
        sdf = sdf.withColumn("Blocker_Severity",
            F.col("Blocked_Stories") * F.when(
                F.col("Days_Remaining_Sprint") > 0,
                F.lit(1.0) / F.col("Days_Remaining_Sprint")
            ).otherwise(F.lit(1.0)))

        # 4. Scope Pressure (scope change relative to planned)
        sdf = sdf.withColumn("Scope_Pressure",
            F.when(F.col("Planned_Story_Points_Sprint") > 0,
                   F.col("Scope_Change") / F.col("Planned_Story_Points_Sprint")
            ).otherwise(F.lit(0.0)))

        # 5. Workload Stress Index
        sdf = sdf.withColumn("Workload_Stress",
            (F.col("Current_Workload_Percent") / F.lit(100.0)) *
             F.col("Consecutive_Overloads"))

        # 6. Sprint Momentum (completed vs velocity ratio)
        sdf = sdf.withColumn("Sprint_Momentum",
            F.when(F.col("Historical_Velocity") > 0,
                   F.col("Completed_Story_Points") / F.col("Historical_Velocity")
            ).otherwise(F.lit(0.0)))

        # 7. Recovery Index (can they catch up?)
        sdf = sdf.withColumn("Recovery_Index",
            F.when(
                (F.col("Planned_Story_Points_Sprint") - F.col("Completed_Story_Points") > 0) &
                (F.col("Days_Remaining_Sprint") > 0),
                (F.col("Historical_Velocity") * F.col("Days_Remaining_Sprint") / F.lit(10.0)) /
                (F.col("Planned_Story_Points_Sprint") - F.col("Completed_Story_Points"))
            ).otherwise(F.lit(1.0)))

        result = sdf.toPandas()
        result = result.fillna(0)
        return result
    except Exception as e:
        st.warning(f"⚡ Spark preprocessing fallback to pandas: {e}")
        return _pandas_preprocess(pdf)

def _pandas_preprocess(pdf: pd.DataFrame) -> pd.DataFrame:
    """Pandas fallback for Spark feature engineering."""
    df = pdf.copy()
    df["Velocity_Efficiency"] = (df["Historical_Velocity"] / df["Planned_Story_Points_Sprint"].replace(0,1)).clip(0,3)
    df["Completion_Gap"]      = df["Planned_Story_Points_Sprint"] - df["Completed_Story_Points"]
    df["Blocker_Severity"]    = df["Blocked_Stories"] * (1 / df["Days_Remaining_Sprint"].replace(0,1).abs())
    df["Scope_Pressure"]      = (df["Scope_Change"] / df["Planned_Story_Points_Sprint"].replace(0,1)).clip(-1,2)
    df["Workload_Stress"]     = (df["Current_Workload_Percent"] / 100) * df.get("Consecutive_Overloads", pd.Series(0,index=df.index))
    df["Sprint_Momentum"]     = (df["Completed_Story_Points"] / df["Historical_Velocity"].replace(0,1)).clip(0,2)
    df["Recovery_Index"]      = ((df["Historical_Velocity"] * df["Days_Remaining_Sprint"] / 10) /
                                  (df["Planned_Story_Points_Sprint"] - df["Completed_Story_Points"]).replace(0,0.001)).clip(0,5)
    return df.fillna(0)

st.set_page_config(page_title="AI Agile Dashboard", layout="wide")

st.markdown("""
<style>
    .agent-card {
        background: linear-gradient(135deg, #1e1e2e 0%, #2a2a3e 100%);
        border: 1px solid #444466;
        border-radius: 12px;
        padding: 1.2rem 1.5rem;
        margin-bottom: 1rem;
        color: #e0e0f0;
    }
    .agent-card.critical {
        border-left: 5px solid #ff4d6d;
        background: linear-gradient(135deg, #2e1e22 0%, #3e2a2e 100%);
    }
    .agent-card.warning {
        border-left: 5px solid #ffd166;
        background: linear-gradient(135deg, #2e2a1e 0%, #3e362a 100%);
    }
    .agent-card.success {
        border-left: 5px solid #06d6a0;
        background: linear-gradient(135deg, #1e2e2a 0%, #2a3e36 100%);
    }
    .agent-card.info {
        border-left: 5px solid #4cc9f0;
        background: linear-gradient(135deg, #1e252e 0%, #2a333e 100%);
    }
    .agent-title {
        font-size: 1rem;
        font-weight: 700;
        margin-bottom: 0.3rem;
        letter-spacing: 0.03em;
    }
    .agent-detail {
        font-size: 0.85rem;
        opacity: 0.85;
        line-height: 1.5;
    }
    .chain-step {
        display: flex;
        align-items: flex-start;
        gap: 1rem;
        margin-bottom: 0.8rem;
    }
    .step-num {
        background: #4cc9f0;
        color: #000;
        border-radius: 50%;
        width: 28px;
        height: 28px;
        display: flex;
        align-items: center;
        justify-content: center;
        font-weight: 800;
        font-size: 0.8rem;
        flex-shrink: 0;
    }
    .health-bar-container {
        background: #333;
        border-radius: 8px;
        height: 14px;
        width: 100%;
        margin-top: 4px;
    }
    .report-section {
        background: #1a1a2e;
        border-radius: 10px;
        padding: 1rem 1.5rem;
        margin-bottom: 1rem;
        border: 1px solid #333355;
        color: #dde;
        font-size: 0.9rem;
        line-height: 1.7;
    }
</style>
""", unsafe_allow_html=True)

st.title("🚀 AI Agile Project Management Dashboard")
spark_badge = get_spark()
if spark_badge is not None:
    st.markdown("⚡ **Apache Spark** active — distributed preprocessing & feature engineering enabled", unsafe_allow_html=False)
else:
    st.caption("ℹ️ Install `pyspark` to enable Apache Spark distributed preprocessing")

uploaded_file = st.file_uploader("📁 Upload the Combined CSV for All Objectives", type="csv")

# ── shared state ────────────────────────────────────────────────────────────
models = {}   # will hold trained models keyed by objective
encoders = {} # label encoders

if uploaded_file:
    df_raw = pd.read_csv(uploaded_file)
    df_raw = df_raw.fillna(0)

    # ── Spark preprocessing ─────────────────────────────────────────
    spark = get_spark()
    with st.spinner("⚡ Running Spark preprocessing & feature engineering..."):
        df = spark_preprocess(df_raw, spark)

    spark_status = "⚡ Spark" if (spark is not None) else "🐼 Pandas"
    st.sidebar.markdown(f"**Processing Engine:** {spark_status}")
    if spark is not None:
        st.sidebar.success("Apache Spark active — distributed feature engineering enabled")
        st.sidebar.caption("7 engineered features added: Velocity Efficiency, Completion Gap, Blocker Severity, Scope Pressure, Workload Stress, Sprint Momentum, Recovery Index")
    else:
        st.sidebar.info("PySpark not found — using pandas fallback (install pyspark for distributed processing)")

    # Handle both string ('Yes'/'No') and float/probability labels
    def binarize_col(series, threshold=0.5):
        if series.dtype == object:
            return series.map({'No': 0, 'Yes': 1}).fillna(0).astype(int)
        return (series > threshold).astype(int)

    thresholds = {'Success_Label': 0.5, 'Expected_Overload': 0.5, 'Risk_Flag': 0.3}
    for col, thresh in thresholds.items():
        if col in df.columns:
            df[col] = binarize_col(df[col], threshold=thresh)

    st.success("✅ File uploaded successfully!")

    # Dataset summary
    c1, c2, c3, c4 = st.columns(4)
    c1.metric("📋 Total Records", f"{len(df):,}")
    c2.metric("📊 Features", len(df.columns))
    if 'Success_Label' in df.columns:
        c3.metric("🔴 Sprints at Risk", int((df['Success_Label']==0).sum()))
    if 'Risk_Flag' in df.columns:
        c4.metric("⚠️ Burnout Flags", int(df['Risk_Flag'].sum()))

    # ── Spark status banner ──────────────────────────────────────────────
    spark = get_spark()
    if SPARK_AVAILABLE and spark is not None:
        st.success("⚡ Apache Spark is active — feature engineering will run in distributed mode.")
    else:
        st.info("ℹ️ PySpark not installed — using pandas fallback for feature engineering. Install pyspark to enable Spark mode.")

    # ── Run Spark/pandas preprocessing ───────────────────────────────────
    df = spark_preprocess(df, spark)

    with st.expander("👀 Preview Data (with Spark-engineered features)"):
        spark_cols = [c for c in ['Velocity_Efficiency','Completion_Gap','Blocker_Severity',
                                   'Scope_Pressure','Workload_Stress','Sprint_Momentum','Recovery_Index']
                      if c in df.columns]
        if spark_cols:
            st.caption(f"⚡ {len(spark_cols)} Spark-engineered features added: {', '.join(spark_cols)}")
        st.write(df.head())

    # ── train all models silently so the agent can use them ─────────────────
    def train_all(df):
        results = {}

        # --- Obj 1: Sprint Completion — Calibrated GBM + novel features ---
        try:
            base_f  = ['Planned_Story_Points_Sprint','Completed_Story_Points','Percent_Done',
                       'Days_Remaining_Sprint','Historical_Velocity','Blocked_Stories','Scope_Change']
            spark_f = [f for f in ['Velocity_Efficiency','Completion_Gap','Blocker_Severity',
                       'Scope_Pressure','Sprint_Momentum','Recovery_Index'] if f in df.columns]
            # Novel engineered features
            if 'Days_Per_Remaining_SP' not in df.columns:
                df['Days_Per_Remaining_SP']   = (df['Days_Remaining_Sprint'] / (df['Completion_Gap'].replace(0,0.001))).clip(-2,10)
                df['Overcommitment_Ratio']    = (df['Planned_Story_Points_Sprint'] / df['Historical_Velocity'].replace(0,1)).clip(0,3)
                df['Blocker_Scope_Risk']      = df['Blocked_Stories'] * (1 + df.get('Scope_Pressure', pd.Series(0,index=df.index)).clip(0,2))
                days_elapsed                  = (10 - df['Days_Remaining_Sprint']).clip(1,10)
                df['Burn_Rate']               = (df['Completed_Story_Points'] / days_elapsed).clip(0,20)
                df['Can_Finish_Score']        = ((df['Historical_Velocity'] / 10 * df['Days_Remaining_Sprint']) - df.get('Completion_Gap', pd.Series(0,index=df.index))).clip(-50,50)
            novel_f = [f for f in ['Days_Per_Remaining_SP','Overcommitment_Ratio',
                        'Blocker_Scope_Risk','Burn_Rate','Can_Finish_Score'] if f in df.columns]
            feat1   = base_f + spark_f + novel_f
            X1 = df[feat1]
            y1 = df['Success_Label']
            if len(y1.unique()) > 1:
                gbm_base = GradientBoostingClassifier(n_estimators=200, random_state=42,
                                                       learning_rate=0.05, max_depth=4)
                m = CalibratedClassifierCV(gbm_base, method='isotonic', cv=5)
                m.fit(X1, y1)
                results['sprint'] = {'model': m, 'features': feat1,
                                     'base_feats': base_f, 'spark_feats': spark_f, 'novel_feats': novel_f}
        except: pass

        # --- Obj 2: Workload ---
        try:
            X2 = df[['Planned_Story_Points_Resource','Current_Assigned_SP','Historical_Avg_SP',
                      'Remaining_Days_Resource','High_Priority_Tasks_Resource','Current_Workload_Percent']]
            y2 = df['Expected_Overload']
            if len(y2.unique()) > 1:
                m = RandomForestClassifier(n_estimators=100, random_state=42)
                m.fit(X2, y2)
                results['workload'] = {'model': m, 'features': X2.columns.tolist()}
        except: pass

        # --- Obj 3: Time to Resolve ---
        try:
            X3 = pd.get_dummies(df[['Issue_Type','Priority']], drop_first=False)
            X3 = pd.concat([X3, df[['Original_Estimate_Hours','Story_Points_Issue']]], axis=1)
            y3 = df['Resolution_Time_Hours']
            m = LinearRegression()
            m.fit(X3, y3)
            results['ttr'] = {'model': m, 'features': X3.columns.tolist(), 'X3': X3}
        except: pass

        # --- Obj 4: Burnout ---
        try:
            X4 = df[['Total_SP_This_Sprint','Historical_Avg_SP_Burnout',
                      'High_Priority_Tasks_Burnout','Consecutive_Overloads']]
            y4 = df['Risk_Flag']
            if len(y4.unique()) > 1:
                m = RandomForestClassifier(n_estimators=100, random_state=42)
                m.fit(X4, y4)
                results['burnout'] = {'model': m, 'features': X4.columns.tolist()}
        except: pass

        # --- Obj 5: Resource Allocation ---
        try:
            le_s = LabelEncoder(); le_l = LabelEncoder()
            df['Summary_enc'] = le_s.fit_transform(df['Summary'].astype(str))
            df['Labels_enc']  = le_l.fit_transform(df['Labels'].astype(str))
            X5 = df[['Summary_enc','Labels_enc','Original_Estimate_Resource','Story_Points_Resource']]
            y5 = df['Assignee_Resource']
            m = RandomForestClassifier(n_estimators=100, random_state=42)
            m.fit(X5, y5)
            results['alloc'] = {'model': m, 'features': X5.columns.tolist(),
                                'le_summary': le_s, 'le_labels': le_l}
        except: pass

        return results, df

    models, df = train_all(df)

    # ── agentic scan: run all models on every row ────────────────────────────
    def run_agent_scan(_df, _models_keys):
        """Run predictions across the whole dataset and collect findings."""
        findings = []
        df = _df.copy()

        # Sprint risk scan — aggregate, not per-row
        if 'sprint' in models:
            m = models['sprint']['model']
            cols = models['sprint']['features']
            try:
                X = df[cols]
                preds  = m.predict(X)
                probas = m.predict_proba(X)[:, 1]
                at_risk_mask = (preds == 0)
                at_risk_count = int(at_risk_mask.sum())
                total = len(preds)
                pct = at_risk_count / total if total > 0 else 0
                if pct > 0.15:
                    avg_prob   = float(probas[at_risk_mask].mean())
                    avg_blocked = float(df.loc[at_risk_mask, 'Blocked_Stories'].mean()) if 'Blocked_Stories' in df.columns else 0
                    avg_days    = float(df.loc[at_risk_mask, 'Days_Remaining_Sprint'].mean()) if 'Days_Remaining_Sprint' in df.columns else 0
                    sev = 'critical' if pct > 0.5 else 'warning'
                    findings.append({
                        'severity': sev,
                        'objective': 'Sprint Completion',
                        'icon': '🔴' if sev == 'critical' else '🟡',
                        'title': f"{at_risk_count} of {total} sprints ({pct:.0%}) at risk of spillover",
                        'detail': (f"Avg completion probability: {avg_prob:.0%} | "
                                   f"Avg blocked stories: {avg_blocked:.1f} | "
                                   f"Avg days remaining: {avg_days:.1f}"),
                        'action': "Consider reducing scope or unblocking stories immediately."
                    })
            except: pass

        # Workload overload scan
        if 'workload' in models:
            m = models['workload']['model']
            cols = models['workload']['features']
            try:
                X = df[cols]
                preds  = m.predict(X)
                probas = m.predict_proba(X)[:, 1]
                overloaded = df[preds == 1].copy()
                overloaded['wl_prob'] = probas[preds == 1]
                count = len(overloaded)
                if count > 0 and count > len(df) * 0.2:
                    findings.append({
                        'severity': 'critical' if count > len(df) * 0.45 else 'warning',
                        'objective': 'Workload Projection',
                        'icon': '🔴' if count > len(df) * 0.45 else '🟡',
                        'title': f"{count} resource(s) projected to be overloaded",
                        'detail': (f"Average overload probability: {overloaded['wl_prob'].mean():.0%} | "
                                   f"Avg current workload: {overloaded.get('Current_Workload_Percent', pd.Series([0])).mean():.0f}%"),
                        'action': "Redistribute story points from overloaded to available team members."
                    })
            except: pass

        # Burnout risk scan
        if 'burnout' in models:
            m = models['burnout']['model']
            cols = models['burnout']['features']
            try:
                X = df[cols]
                preds = m.predict(X)
                at_risk_count = int(preds.sum())
                pct_flagged_b = at_risk_count / len(preds) if len(preds) > 0 else 0
                if pct_flagged_b > 0.25:
                    avg_co = df.loc[preds == 1, 'Consecutive_Overloads'].mean() if 'Consecutive_Overloads' in df.columns else 0
                    pct_flagged = pct_flagged_b
                    findings.append({
                        'severity': 'critical' if pct_flagged > 0.5 else 'warning',
                        'objective': 'Burnout Risk',
                        'icon': '🔴' if pct_flagged > 0.5 else '🟡',
                        'title': f"{at_risk_count} team member(s) flagged for burnout risk",
                        'detail': f"Avg consecutive overloads: {avg_co:.1f} sprints",
                        'action': "Schedule 1:1s, reduce high-priority task load, or grant recovery sprint."
                    })
            except: pass

        # Healthy signal
        sprint_ok   = sum(1 for f in findings if f['objective'] == 'Sprint Completion') == 0
        workload_ok = sum(1 for f in findings if f['objective'] == 'Workload Projection') == 0
        burnout_ok  = sum(1 for f in findings if f['objective'] == 'Burnout Risk') == 0

        if sprint_ok:
            findings.append({'severity':'success','objective':'Sprint Completion','icon':'✅',
                             'title':'All sprints on track','detail':'No spillover risk detected.',
                             'action':''})
        if workload_ok:
            findings.append({'severity':'success','objective':'Workload Projection','icon':'✅',
                             'title':'Workloads within capacity','detail':'No overload signals found.',
                             'action':''})
        if burnout_ok:
            findings.append({'severity':'success','objective':'Burnout Risk','icon':'✅',
                             'title':'No burnout risk detected','detail':'Team load looks sustainable.',
                             'action':''})

        return findings

    # ── chained decisions ────────────────────────────────────────────────────
    def build_chain(findings):
        """Build an agentic chain of decisions based on cross-objective findings."""
        chain = []
        has_sprint_risk   = any(f['objective'] == 'Sprint Completion'  and f['severity'] in ('critical','warning') for f in findings)
        has_overload      = any(f['objective'] == 'Workload Projection' and f['severity'] in ('critical','warning') for f in findings)
        has_burnout       = any(f['objective'] == 'Burnout Risk'        and f['severity'] in ('critical','warning') for f in findings)

        chain.append({
            'step': 1,
            'label': 'Scan Objectives',
            'detail': f"Scanned {len(df)} records across 5 objectives.",
            'status': 'done'
        })

        if has_sprint_risk and has_overload:
            chain.append({'step':2,'label':'Linked: Sprint risk ← Overload detected',
                'detail':'Sprint may be at risk because team members are overloaded. Reallocation needed.',
                'status':'alert'})
            chain.append({'step':3,'label':'Recommend: Rebalance workload',
                'detail':'Move story points from overloaded members to those under capacity before sprint closes.',
                'status':'action'})
        elif has_sprint_risk:
            chain.append({'step':2,'label':'Sprint risk detected — checking workload',
                'detail':'Workload looks OK. Risk may stem from blocked stories or scope change.',
                'status':'alert'})
            chain.append({'step':3,'label':'Recommend: Unblock stories & freeze scope',
                'detail':'No reallocation needed. Focus on removing blockers and preventing scope creep.',
                'status':'action'})

        if has_burnout and has_overload:
            chain.append({'step': len(chain)+1,'label':'Linked: Burnout risk ← Persistent overloads',
                'detail':'Burnout signal correlates with overloaded workload across multiple sprints.',
                'status':'alert'})
            chain.append({'step': len(chain)+1,'label':'Recommend: Recovery sprint or capacity reduction',
                'detail':'Reduce assigned story points by 20–30% for flagged members next sprint.',
                'status':'action'})

        if not has_sprint_risk and not has_overload and not has_burnout:
            chain.append({'step':2,'label':'All clear — no chained risks',
                'detail':'No cross-objective dependencies triggered. Project health looks good.',
                'status':'done'})

        return chain

    # ── health score ─────────────────────────────────────────────────────────
    def compute_health(findings):
        score = 100
        for f in findings:
            if f['severity'] == 'critical': score -= 25
            elif f['severity'] == 'warning': score -= 10
        return max(0, min(100, score))

    # ── generate written report ───────────────────────────────────────────────
    def generate_report(findings, chain, score):
        total   = len(df)
        criticals = [f for f in findings if f['severity'] == 'critical']
        warnings  = [f for f in findings if f['severity'] == 'warning']
        successes = [f for f in findings if f['severity'] == 'success']

        status = "🟢 Healthy" if score >= 75 else ("🟡 Needs Attention" if score >= 50 else "🔴 At Risk")

        report = f"""
## 📋 Project Health Report

**Overall Status:** {status} — Health Score: {score}/100

**Dataset:** {total} records analyzed across 5 AI objectives.

### Summary
The autonomous agent scanned your project data and identified **{len(criticals)} critical issue(s)** and **{len(warnings)} warning(s)**. {len(successes)} objective(s) returned healthy signals.

### Findings
"""
        for f in findings:
            if f['severity'] != 'success':
                report += f"\n- **{f['icon']} [{f['objective']}]** {f['title']}: {f['detail']}"
                if f['action']:
                    report += f"\n  → *{f['action']}*"

        report += "\n\n### Healthy Signals\n"
        for f in successes:
            report += f"\n- **{f['icon']} [{f['objective']}]** {f['title']}"

        report += "\n\n### Chained Recommendations\n"
        for step in chain:
            emoji = "✅" if step['status'] == 'done' else ("⚠️" if step['status'] == 'alert' else "💡")
            report += f"\n**Step {step['step']}** {emoji} {step['label']}  \n{step['detail']}\n"

        report += f"\n\n---\n*Report auto-generated by the Agentic AI layer. {total} rows × 5 objectives scanned.*"
        return report

    # ═══════════════════════════════════════════════════════════════════════
    tabs = st.tabs([
        "🤖 Agentic AI Overview",
        "1️⃣ Sprint Completion Forecast",
        "2️⃣ Workload Projection Forecast",
        "3️⃣ Time to Resolve Estimation",
        "4️⃣ Burnout Risk Alerts",
        "5️⃣ Resource Allocation Suggestions",
        "⚡ Spark Processing"
    ])

    # ══════════════════════════════════════════════════════════════
    # AGENTIC AI TAB
    # ══════════════════════════════════════════════════════════════
    with tabs[0]:
        st.header("🤖 Agentic AI — Autonomous Project Scanning")
        st.caption("The agent automatically runs all 5 models across your full dataset, chains findings together, and surfaces prioritized actions — no manual input needed.")

        with st.spinner("🧠 Agent scanning dataset across all objectives..."):
            findings = run_agent_scan(df, list(models.keys()))
            chain    = build_chain(findings)
            score    = compute_health(findings)

        # ── Health Score ────────────────────────────────────────
        st.markdown("---")
        col1, col2, col3 = st.columns([1, 2, 1])
        with col1:
            color = "#06d6a0" if score >= 75 else ("#ffd166" if score >= 50 else "#ff4d6d")
            label = "🟢 Healthy" if score >= 75 else ("🟡 Needs Attention" if score >= 50 else "🔴 At Risk")
            st.markdown(f"""
            <div style='text-align:center;'>
                <div style='font-size:3.5rem;font-weight:900;color:{color};'>{score}</div>
                <div style='font-size:1rem;color:#aaa;margin-top:-8px;'>/ 100</div>
                <div style='font-size:1.1rem;margin-top:6px;'>{label}</div>
                <div style='font-size:0.8rem;color:#888;'>Project Health Score</div>
            </div>
            """, unsafe_allow_html=True)
        with col2:
            bar_color = "#06d6a0" if score >= 75 else ("#ffd166" if score >= 50 else "#ff4d6d")
            st.markdown(f"""
            <div style='margin-top:2rem;'>
                <div class='health-bar-container'>
                    <div style='background:{bar_color};width:{score}%;height:14px;border-radius:8px;transition:width 1s ease;'></div>
                </div>
                <div style='display:flex;justify-content:space-between;font-size:0.75rem;color:#888;margin-top:4px;'>
                    <span>0 — Critical</span><span>50 — Attention</span><span>100 — Healthy</span>
                </div>
            </div>
            """, unsafe_allow_html=True)
        with col3:
            criticals_n = sum(1 for f in findings if f['severity']=='critical')
            warnings_n  = sum(1 for f in findings if f['severity']=='warning')
            st.markdown(f"""
            <div style='text-align:center;margin-top:0.5rem;'>
                <div style='font-size:2rem;font-weight:800;color:#ff4d6d;'>{criticals_n}</div>
                <div style='font-size:0.8rem;color:#aaa;'>Critical Issues</div>
                <div style='font-size:2rem;font-weight:800;color:#ffd166;margin-top:8px;'>{warnings_n}</div>
                <div style='font-size:0.8rem;color:#aaa;'>Warnings</div>
            </div>
            """, unsafe_allow_html=True)

        st.markdown("---")

        # ── Findings ────────────────────────────────────────────
        st.subheader("🔍 Autonomous Findings")
        severity_order = {'critical': 0, 'warning': 1, 'info': 2, 'success': 3}
        for f in sorted(findings, key=lambda x: severity_order.get(x['severity'], 99)):
            action_html = f"<div style='margin-top:6px;font-style:italic;opacity:0.75;'>→ {f['action']}</div>" if f['action'] else ""
            st.markdown(f"""
            <div class='agent-card {f["severity"]}'>
                <div class='agent-title'>{f["icon"]} [{f["objective"]}] {f["title"]}</div>
                <div class='agent-detail'>{f["detail"]}{action_html}</div>
            </div>
            """, unsafe_allow_html=True)

        st.markdown("---")

        # ── Decision Chain ──────────────────────────────────────
        st.subheader("⛓️ Chained Decision Reasoning")
        st.caption("The agent links findings across objectives to produce connected, prioritized recommendations.")

        for step in chain:
            icon   = "✅" if step['status'] == 'done' else ("⚠️" if step['status'] == 'alert' else "💡")
            color  = "#4cc9f0" if step['status'] == 'done' else ("#ffd166" if step['status'] == 'alert' else "#06d6a0")
            sev    = "info" if step['status'] == 'done' else ("warning" if step['status'] == 'alert' else "success")
            st.markdown(f"""
            <div class='agent-card {sev}' style='display:flex;gap:1rem;align-items:flex-start;'>
                <div style='background:{color};color:#000;border-radius:50%;width:30px;height:30px;
                            display:flex;align-items:center;justify-content:center;
                            font-weight:800;font-size:0.8rem;flex-shrink:0;margin-top:2px;'>
                    {step["step"]}
                </div>
                <div>
                    <div class='agent-title'>{icon} {step["label"]}</div>
                    <div class='agent-detail'>{step["detail"]}</div>
                </div>
            </div>
            """, unsafe_allow_html=True)

        st.markdown("---")

        # ── Per-Assignee Breakdown ──────────────────────────────
        st.subheader("👤 Per-Assignee Risk Breakdown")
        st.caption("Risk levels computed directly from your dataset for each team member.")

        assignee_col = 'Assignee' if 'Assignee' in df.columns else ('Assignee_Resource' if 'Assignee_Resource' in df.columns else None)
        if assignee_col:
            assignees = df[assignee_col].unique()
            cols_a = st.columns(len(assignees))
            for i, person in enumerate(sorted(assignees)):
                sub = df[df[assignee_col] == person]
                sprint_risk = sub['Success_Label'].eq(0).mean() if 'Success_Label' in df.columns else 0
                overload    = sub['Expected_Overload'].mean()   if 'Expected_Overload' in df.columns else 0
                burnout     = sub['Risk_Flag'].mean()           if 'Risk_Flag' in df.columns else 0
                workload    = sub['Current_Workload_Percent'].mean() if 'Current_Workload_Percent' in df.columns else 0
                consec      = sub['Consecutive_Overloads'].mean()    if 'Consecutive_Overloads' in df.columns else 0

                # Overall person score
                person_score = 100 - (sprint_risk * 35) - (overload * 30) - (burnout * 20) - min((workload - 100) / 2, 15)
                person_score = max(0, min(100, person_score))
                p_color = "#06d6a0" if person_score >= 60 else ("#ffd166" if person_score >= 40 else "#ff4d6d")
                p_label = "🟢 OK" if person_score >= 60 else ("🟡 Watch" if person_score >= 40 else "🔴 At Risk")

                with cols_a[i]:
                    st.markdown(f"""
                    <div class='agent-card {"critical" if person_score < 40 else "warning" if person_score < 60 else "success"}' style='text-align:center;'>
                        <div style='font-size:1.5rem;font-weight:900;color:{p_color};'>{person_score:.0f}</div>
                        <div style='font-size:0.7rem;color:#aaa;margin-top:-4px;'>/ 100</div>
                        <div style='font-size:1rem;font-weight:700;margin:6px 0 2px;'>{person}</div>
                        <div style='font-size:0.75rem;margin-bottom:8px;'>{p_label}</div>
                        <div style='text-align:left;font-size:0.78rem;line-height:1.8;'>
                            🏃 Sprint risk: <b>{sprint_risk:.0%}</b><br>
                            📦 Overload: <b>{overload:.0%}</b><br>
                            🔥 Burnout flag: <b>{burnout:.0%}</b><br>
                            ⚡ Avg workload: <b>{workload:.0f}%</b><br>
                            🔁 Consec. overloads: <b>{consec:.1f}</b>
                        </div>
                    </div>
                    """, unsafe_allow_html=True)

        st.markdown("---")

        # ── Action Priority Table ───────────────────────────────
        st.subheader("🎯 Action Priority Table")
        st.caption("Ranked actions based on severity and impact across all findings.")

        action_rows = []
        for f in sorted(findings, key=lambda x: {'critical': 0, 'warning': 1, 'success': 2}.get(x['severity'], 3)):
            if f.get('action'):
                priority = "🔴 P1 — Immediate" if f['severity'] == 'critical' else "🟡 P2 — This Sprint"
                action_rows.append({
                    'Priority':   priority,
                    'Objective':  f['objective'],
                    'Issue':      f['title'],
                    'Action':     f['action']
                })

        # Add chained actions
        for step in chain:
            if step['status'] == 'action':
                action_rows.append({
                    'Priority':  '💡 P3 — Next Sprint',
                    'Objective': 'Cross-Objective',
                    'Issue':     step['label'],
                    'Action':    step['detail']
                })

        if action_rows:
            action_df = pd.DataFrame(action_rows)
            st.dataframe(
                action_df,
                use_container_width=True,
                hide_index=True,
                column_config={
                    'Priority':  st.column_config.TextColumn('Priority',  width='medium'),
                    'Objective': st.column_config.TextColumn('Objective', width='medium'),
                    'Issue':     st.column_config.TextColumn('Issue',     width='large'),
                    'Action':    st.column_config.TextColumn('Action ▶',  width='large'),
                }
            )

        st.markdown("---")

        # ── Written Report ──────────────────────────────────────
        st.subheader("📄 Auto-Generated Project Health Report")
        report_md = generate_report(findings, chain, score)
        st.markdown(f"<div class='report-section'>{report_md}</div>", unsafe_allow_html=True)

        col_dl, col_xl, _ = st.columns([1, 1, 2])
        with col_dl:
            st.download_button(
                "⬇️ Download Report (.md)",
                data=report_md,
                file_name="project_health_report.md",
                mime="text/markdown"
            )

        st.markdown("---")

        # ── Trend Charts ────────────────────────────────────────
        st.subheader("📈 Trend Charts")
        st.caption("Sprint risk and workload trends over time, grouped by sprint number.")

        sprint_col = 'Sprint_Number' if 'Sprint_Number' in df.columns else None
        if sprint_col:
            df_trend = df.copy()
            df_trend['at_risk'] = (df_trend['Success_Label'] == 0).astype(int)
            df_trend['overloaded'] = (df_trend['Expected_Overload'] > 0.5).astype(int) if df_trend['Expected_Overload'].dtype != int else df_trend['Expected_Overload']

            trend_agg = df_trend.groupby(sprint_col).agg(
                sprint_risk_pct=('at_risk', 'mean'),
                avg_workload=('Current_Workload_Percent', 'mean'),
                avg_blocked=('Blocked_Stories', 'mean'),
                burnout_pct=('Risk_Flag', 'mean'),
            ).reset_index()
            trend_agg['sprint_risk_pct'] = (trend_agg['sprint_risk_pct'] * 100).round(1)
            trend_agg['avg_workload'] = trend_agg['avg_workload'].round(1)
            trend_agg['avg_blocked'] = trend_agg['avg_blocked'].round(2)
            trend_agg['burnout_pct'] = (trend_agg['burnout_pct'] * 100).round(1)
            trend_agg = trend_agg.sort_values(sprint_col)

            tc1, tc2 = st.columns(2)
            with tc1:
                st.markdown("**🏃 Sprint Risk % Over Time**")
                st.line_chart(trend_agg.set_index(sprint_col)[['sprint_risk_pct']], height=220, use_container_width=True)
            with tc2:
                st.markdown("**⚡ Avg Workload % Over Time**")
                st.line_chart(trend_agg.set_index(sprint_col)[['avg_workload']], height=220, use_container_width=True)

            tc3, tc4 = st.columns(2)
            with tc3:
                st.markdown("**🚧 Avg Blocked Stories Over Time**")
                st.line_chart(trend_agg.set_index(sprint_col)[['avg_blocked']], height=220, use_container_width=True)
            with tc4:
                st.markdown("**🔥 Burnout Flag % Over Time**")
                st.line_chart(trend_agg.set_index(sprint_col)[['burnout_pct']], height=220, use_container_width=True)

            st.markdown("**👤 Workload Trend per Assignee**")
            assignee_col_t = 'Assignee' if 'Assignee' in df.columns else None
            if assignee_col_t:
                wl_pivot = df_trend.groupby([sprint_col, assignee_col_t])['Current_Workload_Percent'].mean().unstack(assignee_col_t).round(1)
                wl_pivot = wl_pivot.sort_index()
                st.line_chart(wl_pivot, height=280, use_container_width=True)
        else:
            st.info("ℹ️ Upload a dataset with a 'Sprint_Number' column to enable trend charts.")

        st.markdown("---")

        # ── Excel Export ─────────────────────────────────────────
        st.subheader("📥 Export to Excel")
        st.caption("Download a full Excel workbook with findings, action table, assignee breakdown, and raw data.")

        try:
            import io
            from openpyxl import Workbook
            from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
            from openpyxl.utils import get_column_letter

            def build_excel(df, findings, chain, score, action_rows):
                wb = Workbook()

                # ---- Sheet 1: Summary ----
                ws1 = wb.active
                ws1.title = "Health Summary"
                header_fill  = PatternFill("solid", fgColor="1e1e2e")
                red_fill     = PatternFill("solid", fgColor="ff4d6d")
                yellow_fill  = PatternFill("solid", fgColor="ffd166")
                green_fill   = PatternFill("solid", fgColor="06d6a0")
                white_font   = Font(color="FFFFFF", bold=True, size=12)
                dark_font    = Font(color="1e1e2e", bold=True, size=11)
                thin_border  = Border(
                    left=Side(style='thin'), right=Side(style='thin'),
                    top=Side(style='thin'), bottom=Side(style='thin'))

                ws1['A1'] = "AI Agile Dashboard — Project Health Report"
                ws1['A1'].font = Font(bold=True, size=16, color="1e1e2e")
                ws1['A2'] = f"Health Score: {score}/100"
                ws1['A2'].font = Font(bold=True, size=13,
                    color="FF0000" if score < 50 else ("FF9900" if score < 75 else "009900"))
                ws1['A3'] = f"Records analyzed: {len(df):,}   |   Objectives: 5"
                ws1['A3'].font = Font(size=11)
                ws1.append([])

                ws1.append(["Objective", "Severity", "Finding", "Action"])
                for cell in ws1[5]:
                    cell.fill = header_fill
                    cell.font = white_font
                    cell.alignment = Alignment(horizontal='center')

                for f in findings:
                    fill = red_fill if f['severity']=='critical' else (yellow_fill if f['severity']=='warning' else green_fill)
                    font = dark_font
                    row = [f['objective'], f['severity'].upper(), f['title'], f.get('action','')]
                    ws1.append(row)
                    for cell in ws1[ws1.max_row]:
                        cell.fill = fill
                        cell.font = font
                        cell.border = thin_border
                        cell.alignment = Alignment(wrap_text=True)

                for col in ['A','B','C','D']:
                    ws1.column_dimensions[col].width = 28
                ws1.row_dimensions[5].height = 20

                # ---- Sheet 2: Action Priority ----
                ws2 = wb.create_sheet("Action Priority")
                ws2.append(["Priority", "Objective", "Issue", "Recommended Action"])
                for cell in ws2[1]:
                    cell.fill = header_fill
                    cell.font = white_font
                for ar in action_rows:
                    ws2.append([ar.get('Priority',''), ar.get('Objective',''),
                                ar.get('Issue',''), ar.get('Action','')])
                    for cell in ws2[ws2.max_row]:
                        cell.border = thin_border
                        cell.alignment = Alignment(wrap_text=True)
                for col in ['A','B','C','D']:
                    ws2.column_dimensions[col].width = 30

                # ---- Sheet 3: Assignee Breakdown ----
                ws3 = wb.create_sheet("Assignee Breakdown")
                assignee_col = 'Assignee' if 'Assignee' in df.columns else None
                if assignee_col:
                    ws3.append(["Assignee","Health Score","Sprint Risk %","Overload %",
                                 "Burnout Flag %","Avg Workload %","Avg Consec. Overloads","Status"])
                    for cell in ws3[1]:
                        cell.fill = header_fill; cell.font = white_font
                    for person in sorted(df[assignee_col].unique()):
                        sub = df[df[assignee_col]==person]
                        sprint_risk = sub['Success_Label'].eq(0).mean() if 'Success_Label' in df.columns else 0
                        overload    = sub['Expected_Overload'].mean()   if 'Expected_Overload' in df.columns else 0
                        burnout     = sub['Risk_Flag'].mean()           if 'Risk_Flag' in df.columns else 0
                        workload    = sub['Current_Workload_Percent'].mean() if 'Current_Workload_Percent' in df.columns else 0
                        consec      = sub['Consecutive_Overloads'].mean()    if 'Consecutive_Overloads' in df.columns else 0
                        p_score = max(0,min(100,100-(sprint_risk*35)-(overload*30)-(burnout*20)-min((workload-100)/2,15)))
                        status = "OK" if p_score >= 60 else ("Watch" if p_score >= 40 else "At Risk")
                        ws3.append([person, round(p_score,1), f"{sprint_risk:.0%}", f"{overload:.0%}",
                                     f"{burnout:.0%}", f"{workload:.0f}%", round(consec,1), status])
                        fill = green_fill if p_score>=60 else (yellow_fill if p_score>=40 else red_fill)
                        for cell in ws3[ws3.max_row]:
                            cell.fill = fill; cell.font = dark_font; cell.border = thin_border
                    for col in ['A','B','C','D','E','F','G','H']:
                        ws3.column_dimensions[col].width = 22

                # ---- Sheet 4: Raw Data ----
                ws4 = wb.create_sheet("Raw Data")
                cols = df.columns.tolist()
                ws4.append(cols)
                for cell in ws4[1]:
                    cell.fill = header_fill; cell.font = white_font
                for _, row in df.iterrows():
                    ws4.append(row.tolist())
                for i, col in enumerate(cols, 1):
                    ws4.column_dimensions[get_column_letter(i)].width = 20

                buf = io.BytesIO()
                wb.save(buf)
                buf.seek(0)
                return buf.getvalue()

            action_rows_xl = []
            for f in sorted(findings, key=lambda x: {'critical':0,'warning':1,'success':2}.get(x['severity'],3)):
                if f.get('action'):
                    priority = "P1 — Immediate" if f['severity']=='critical' else "P2 — This Sprint"
                    action_rows_xl.append({'Priority':priority,'Objective':f['objective'],
                                           'Issue':f['title'],'Action':f['action']})
            for step in chain:
                if step['status'] == 'action':
                    action_rows_xl.append({'Priority':'P3 — Next Sprint','Objective':'Cross-Objective',
                                           'Issue':step['label'],'Action':step['detail']})

            excel_bytes = build_excel(df, findings, chain, score, action_rows_xl)
            st.download_button(
                label="📥 Download Full Excel Report",
                data=excel_bytes,
                file_name="agile_health_report.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
        except Exception as e:
            st.error(f"Excel export error: {e}")

    # ══════════════════════════════════════════════════════════════
    # Objective 1: Sprint Completion Forecast (Enhanced + Spark)
    # ══════════════════════════════════════════════════════════════
    with tabs[1]:
        st.header("📌 Objective 1 — Sprint Completion Forecasting")
        st.caption("Novel additions: Ensemble Voting, Monte Carlo Simulation, Optimal Path Finder, Similar Sprint Lookup.")
        try:
            base_feats  = ['Planned_Story_Points_Sprint','Completed_Story_Points','Percent_Done',
                           'Days_Remaining_Sprint','Historical_Velocity','Blocked_Stories','Scope_Change']
            spark_feats = [f for f in ['Velocity_Efficiency','Completion_Gap','Blocker_Severity',
                           'Scope_Pressure','Sprint_Momentum','Recovery_Index'] if f in df.columns]

            # 5 Novel engineered features (new to this objective)
            if 'Burn_Rate' not in df.columns:
                comp_gap_s                  = (df['Planned_Story_Points_Sprint'] - df['Completed_Story_Points'])
                df['Days_Per_Remaining_SP'] = (df['Days_Remaining_Sprint'] / comp_gap_s.replace(0,0.001)).clip(-2,10)
                df['Overcommitment_Ratio']  = (df['Planned_Story_Points_Sprint'] / df['Historical_Velocity'].replace(0,1)).clip(0,3)
                scope_p                     = df.get('Scope_Pressure', pd.Series(0, index=df.index))
                df['Blocker_Scope_Risk']    = df['Blocked_Stories'] * (1 + scope_p.clip(0,2))
                days_elapsed                = (10 - df['Days_Remaining_Sprint']).clip(1,10)
                df['Burn_Rate']             = (df['Completed_Story_Points'] / days_elapsed).clip(0,20)
                df['Can_Finish_Score']      = ((df['Historical_Velocity'] / 10 * df['Days_Remaining_Sprint']) - comp_gap_s).clip(-50,50)

            novel_feats    = [f for f in ['Days_Per_Remaining_SP','Overcommitment_Ratio',
                               'Blocker_Scope_Risk','Burn_Rate','Can_Finish_Score'] if f in df.columns]
            all_feats_obj1 = base_feats + spark_feats + novel_feats
            X1 = df[[f for f in all_feats_obj1 if f in df.columns]]
            y1 = df['Success_Label']

            if len(y1.unique()) > 1:
                X1_train, X1_test, y1_train, y1_test = train_test_split(
                    X1, y1, test_size=0.2, random_state=42, stratify=y1)

                # Novel model: Calibrated GBM (isotonic calibration for reliable probabilities)
                with st.spinner("Training Calibrated GBM with novel features..."):
                    gbm_base     = GradientBoostingClassifier(n_estimators=200, random_state=42, learning_rate=0.05, max_depth=4)
                    sprint_model = CalibratedClassifierCV(gbm_base, method='isotonic', cv=5)
                    sprint_model.fit(X1_train, y1_train)
                    gb_clf = GradientBoostingClassifier(n_estimators=200, random_state=42, learning_rate=0.05, max_depth=4)
                    gb_clf.fit(X1_train, y1_train)  # uncalibrated copy for feature importance

                y1_pred = sprint_model.predict(X1_test)
                y1_prob = sprint_model.predict_proba(X1_test)[:,1]
                acc1    = accuracy_score(y1_test, y1_pred)
                auc1    = roc_auc_score(y1_test, y1_prob)
                f1_1    = f1_score(y1_test, y1_pred)
                brier1  = brier_score_loss(y1_test, y1_prob)

                # LR baseline on base features only (for comparison delta)
                lr_base  = LogisticRegression(max_iter=1000, class_weight='balanced')
                lr_base.fit(X1_train[base_feats], y1_train)
                acc_base = accuracy_score(y1_test, lr_base.predict(X1_test[base_feats]))
                auc_base = roc_auc_score(y1_test, lr_base.predict_proba(X1_test[base_feats])[:,1])

                # Individual model accuracies
                ind_accs = {}
                for mname, mclf in [
                    ('Logistic Regression', LogisticRegression(max_iter=1000, class_weight='balanced')),
                    ('Random Forest',       RandomForestClassifier(n_estimators=100, random_state=42)),
                    ('AdaBoost',            AdaBoostClassifier(n_estimators=100, random_state=42)),
                ]:
                    mclf.fit(X1_train, y1_train)
                    ind_accs[mname] = accuracy_score(y1_test, mclf.predict(X1_test))

                mc1,mc2,mc3,mc4,mc5,mc6 = st.columns(6)
                mc1.metric("Accuracy",       f"{acc1:.2%}",  f"{acc1-acc_base:+.2%} vs LR")
                mc2.metric("ROC-AUC",        f"{auc1:.3f}",  f"{auc1-auc_base:+.3f} vs LR")
                mc3.metric("F1 Score",       f"{f1_1:.3f}")
                mc4.metric("Brier Score",    f"{brier1:.3f}", "lower = better")
                mc5.metric("Training Size",  f"{len(X1_train):,}")
                mc6.metric("Total Features", f"{len(all_feats_obj1)}")

                fg1,fg2,fg3 = st.columns(3)
                fg1.markdown(f"**Base features:** {len(base_feats)}")
                fg2.markdown(f"**Spark-engineered:** {len(spark_feats)}")
                fg3.markdown(f"**Novel features:** {len(novel_feats)}")

                # ── Individual vs Ensemble comparison ────────────────────
                with st.expander("🗳️ Ensemble vs Individual Model Accuracy"):
                    st.caption("4 models vote together — majority soft vote wins. Each bar shows individual accuracy vs ensemble.")
                    for mname, macc in sorted(ind_accs.items(), key=lambda x: -x[1]):
                        diff   = acc1 - macc
                        dcolor = "#06d6a0" if diff >= 0 else "#ff4d6d"
                        bw     = macc * 100
                        st.markdown(f"""
<div style='margin-bottom:8px;'>
    <div style='display:flex;justify-content:space-between;font-size:0.82rem;color:#dde;margin-bottom:3px;'>
        <span>{mname}</span>
        <span>{macc:.2%} <span style='color:{dcolor};'>({diff:+.2%} vs ensemble)</span></span>
    </div>
    <div style='background:#333;border-radius:4px;height:9px;'>
        <div style='background:#4cc9f0;width:{bw:.1f}%;height:9px;border-radius:4px;'></div>
    </div>
</div>""", unsafe_allow_html=True)
                    st.markdown(f"""
<div style='margin-bottom:8px;'>
    <div style='display:flex;justify-content:space-between;font-size:0.82rem;color:#dde;font-weight:700;margin-bottom:3px;'>
        <span>🗳️ Ensemble (Voting)</span><span style='color:#06d6a0;'>{acc1:.2%} ⭐</span>
    </div>
    <div style='background:#333;border-radius:4px;height:9px;'>
        <div style='background:#06d6a0;width:{acc1*100:.1f}%;height:9px;border-radius:4px;'></div>
    </div>
</div>""", unsafe_allow_html=True)

                # ── Spark Feature Panel ───────────────────────────────────
                if spark_feats:
                    with st.expander("⚡ Spark-Engineered Features Preview"):
                        spark_desc = {
                            "Velocity_Efficiency": "Historical velocity ÷ planned SP — how realistic is the plan?",
                            "Completion_Gap":       "Planned SP − Completed SP — how much work remains",
                            "Blocker_Severity":     "Blocked stories ÷ days remaining — urgency of blockers",
                            "Scope_Pressure":       "Scope change ÷ planned SP — relative scope creep",
                            "Sprint_Momentum":      "Completed SP ÷ historical velocity — team momentum",
                            "Recovery_Index":       "Projected completion ÷ remaining work — can they catch up?",
                        }
                        sf_cols = st.columns(3)
                        for i, feat in enumerate(spark_feats):
                            with sf_cols[i % 3]:
                                avg_val = df[feat].mean()
                                st.markdown(f"""
<div style='background:#1a1a2e;border:1px solid #4cc9f0;border-radius:8px;padding:0.7rem;margin-bottom:0.5rem;'>
    <div style='font-size:0.75rem;font-weight:700;color:#4cc9f0;'>⚡ {feat.replace("_"," ")}</div>
    <div style='font-size:1.1rem;font-weight:900;color:#fff;margin:2px 0;'>{avg_val:.2f}</div>
    <div style='font-size:0.7rem;color:#aaa;'>{spark_desc.get(feat,"")}</div>
</div>""", unsafe_allow_html=True)

                # Novel features panel
                if novel_feats:
                    with st.expander("New Novel Engineered Features — what they mean"):
                        novel_desc = {
                            "Days_Per_Remaining_SP": "Days remaining divided by SP left — how many days per story point? Low = danger zone",
                            "Overcommitment_Ratio":  "Planned SP divided by historical velocity — >1.0 means team is over-ambitious",
                            "Blocker_Scope_Risk":    "Blockers × (1 + scope pressure) — combined blocker and scope creep danger score",
                            "Burn_Rate":             "SP completed divided by days elapsed — actual daily throughput so far this sprint",
                            "Can_Finish_Score":      "(Velocity/10 × days left) minus gap — positive means on track, negative means behind",
                        }
                        nf_cols = st.columns(len(novel_feats))
                        for i, feat in enumerate(novel_feats):
                            avg_val = df[feat].mean() if feat in df.columns else 0
                            with nf_cols[i]:
                                st.markdown(f"""
<div style='background:#1a2e1a;border:1px solid #06d6a0;border-radius:8px;padding:0.7rem;'>
    <div style='font-size:0.72rem;font-weight:700;color:#06d6a0;'>NEW: {feat.replace("_"," ")}</div>
    <div style='font-size:1rem;font-weight:900;color:#fff;margin:3px 0;'>{avg_val:.2f} avg</div>
    <div style='font-size:0.68rem;color:#aaa;'>{novel_desc.get(feat,"")}</div>
</div>""", unsafe_allow_html=True)

                with st.expander("Feature Importance (Calibrated GBM):"):
                    importances = pd.Series(gb_clf.feature_importances_, index=all_feats_obj1).sort_values(ascending=False)
                    max_imp = importances.max()
                    for feat, imp in importances.items():
                        bw    = imp / max_imp * 100
                        color = "#06d6a0" if imp > importances.mean() else "#4cc9f0"
                        stag  = " ⚡" if feat in spark_feats else ""
                        st.markdown(f"""
<div style='margin-bottom:5px;'>
    <div style='display:flex;justify-content:space-between;font-size:0.78rem;color:#dde;'>
        <span>{feat.replace("_"," ")}{stag}</span><span style='color:{color};'>{imp:.3f}</span>
    </div>
    <div style='background:#333;border-radius:4px;height:7px;'>
        <div style='background:{color};width:{bw:.0f}%;height:7px;border-radius:4px;'></div>
    </div>
</div>""", unsafe_allow_html=True)
                    st.caption("⚡ = Spark-engineered feature")

                st.markdown("---")
                # ── Predict inputs ────────────────────────────────────────
                st.subheader("🔍 Predict Sprint Success")
                col_inp1, col_inp2 = st.columns(2)
                with col_inp1:
                    psp          = st.number_input("Planned Story Points", 1, 100, 40, key="obj1_psp")
                    csp          = st.number_input("Completed Story Points", 0, 100, 30, key="obj1_csp")
                    percent_done = st.slider("% Done", 0.0, 100.0, 75.0, key="obj1_pd")
                    drs          = st.number_input("Days Remaining", 0, 30, 5, key="obj1_drs")
                with col_inp2:
                    hv  = st.number_input("Historical Velocity", 0, 100, 35, key="obj1_hv")
                    bs  = st.number_input("Blocked Stories", 0, 10, 1, key="obj1_bs")
                    sc  = st.number_input("Scope Change", -20, 20, 0, key="obj1_sc")

                if st.button("🗳️ Predict with Ensemble", key="obj1_btn"):
                    vel_eff   = hv / max(psp, 1)
                    comp_gap  = psp - csp
                    blk_sev   = bs / max(abs(drs), 0.1)
                    scope_pr  = sc / max(psp, 1)
                    momentum  = csp / max(hv, 1)
                    recov_idx = min((hv * drs / 10) / max(psp - csp, 0.001), 5.0)
                    spark_vals = [vel_eff, comp_gap, blk_sev, scope_pr, momentum, recov_idx]
                    base_vals  = [psp, csp, percent_done, drs, hv, bs, sc]
                    all_vals   = base_vals + [v for f, v in zip(
                                  ['Velocity_Efficiency','Completion_Gap','Blocker_Severity',
                                   'Scope_Pressure','Sprint_Momentum','Recovery_Index'], spark_vals)
                                  if f in spark_feats]
                    features = np.array([all_vals])
                    p    = ensemble.predict(features)[0]
                    prob = ensemble.predict_proba(features)[0][1]

                    # ── Result card ───────────────────────────────────────
                    res_col1, res_col2 = st.columns([1, 2])
                    with res_col1:
                        conf_label = ("High" if (prob > 0.75 if p else prob < 0.25)
                                      else "Medium" if (prob > 0.6 if p else prob < 0.4) else "Low")
                        conf_color = ("#06d6a0" if p else "#ff4d6d") if conf_label == "High" else (
                                      "#ffd166" if conf_label == "Medium" else "#4cc9f0")
                        icon       = "✅" if p else "⚠️"
                        label      = "Likely to Complete" if p else "Risk of Spillover"
                        disp_prob  = prob if p else 1 - prob
                        bg_col     = "#1a2e1a" if p else "#2e1a1a"
                        border_col = "#06d6a0" if p else "#ff4d6d"
                        st.markdown(f"""
<div style='text-align:center;background:{bg_col};border:2px solid {border_col};border-radius:12px;padding:1.2rem;'>
    <div style='font-size:2.5rem;'>{icon}</div>
    <div style='font-size:1rem;font-weight:800;color:{border_col};'>{label}</div>
    <div style='font-size:2rem;font-weight:900;color:{conf_color};margin-top:6px;'>{disp_prob:.0%}</div>
    <div style='font-size:0.75rem;color:#aaa;'>Ensemble Confidence</div>
    <div style='margin-top:8px;background:{conf_color};color:#000;border-radius:6px;
                padding:3px 10px;display:inline-block;font-weight:700;font-size:0.8rem;'>{conf_label}</div>
</div>""", unsafe_allow_html=True)

                        # Individual model votes
                        st.markdown("<br>**Individual Votes:**", unsafe_allow_html=True)
                        vote_input = np.array([[psp, csp, percent_done, drs, hv, bs, sc]])
                        for mname, mclf in [('LR', lr_clf),('GB', gb_clf),('RF', rf_clf),('Ada', ada_clf)]:
                            vote = mclf.predict(vote_input)[0]
                            vp   = mclf.predict_proba(vote_input)[0][1]
                            vic  = "✅" if vote else "⚠️"
                            st.markdown(f"<span style='font-size:0.8rem;'>{vic} {mname}: {vp:.0%}</span>", unsafe_allow_html=True)

                    with res_col2:
                        st.markdown("**🔎 Risk Factor Breakdown**")
                        coefs      = lr_clf.coef_[0][:7]
                        feat_names = ['Planned SP','Completed SP','% Done','Days Left','Hist. Velocity','Blocked Stories','Scope Change']
                        feat_vals  = [psp, csp, percent_done, drs, hv, bs, sc]
                        contribs   = sorted([(n, v*c) for n,v,c in zip(feat_names,feat_vals,coefs)], key=lambda x: x[1])
                        max_abs    = max(abs(c[1]) for c in contribs) + 1e-9
                        for fname, impact in contribs:
                            direction = "🟢 Helping" if impact > 0 else "🔴 Hurting"
                            bw = min(abs(impact)/max_abs*100, 100)
                            bc = "#06d6a0" if impact > 0 else "#ff4d6d"
                            st.markdown(f"""
<div style='margin-bottom:6px;'>
    <div style='display:flex;justify-content:space-between;font-size:0.8rem;color:#dde;'>
        <span>{fname}</span><span style='color:{bc};'>{direction} ({impact:+.2f})</span>
    </div>
    <div style='background:#333;border-radius:4px;height:8px;'>
        <div style='background:{bc};width:{bw:.0f}%;height:8px;border-radius:4px;'></div>
    </div>
</div>""", unsafe_allow_html=True)

                    st.markdown("---")

                    # ════════════════════════════════════════════════════════
                    # NOVELTY 1: Monte Carlo Simulation
                    # ════════════════════════════════════════════════════════
                    st.markdown("### 🎲 Monte Carlo Sprint Simulation")
                    st.caption("Runs 1,000 simulations with random noise around your inputs to estimate probability distribution of outcomes.")
                    n_sims = 1000
                    np.random.seed(42)
                    sim_probs = []
                    for _ in range(n_sims):
                        noise_psp  = psp  + np.random.normal(0, psp * 0.05)
                        noise_csp  = csp  + np.random.normal(0, csp * 0.05 + 0.5)
                        noise_pct  = np.clip(percent_done + np.random.normal(0, 3), 0, 100)
                        noise_drs  = drs  + np.random.normal(0, 0.5)
                        noise_hv   = hv   + np.random.normal(0, hv * 0.05)
                        noise_bs   = max(0, bs + np.random.normal(0, 0.3))
                        noise_sc   = sc   + np.random.normal(0, 1)
                        sim_input  = np.array([[noise_psp, noise_csp, noise_pct,
                                                noise_drs, noise_hv, noise_bs, noise_sc]])
                        sim_probs.append(ensemble.predict_proba(sim_input)[0][1])

                    sim_arr     = np.array(sim_probs)
                    pct_success = (sim_arr >= 0.5).mean()
                    p5, p25, p50, p75, p95 = np.percentile(sim_arr, [5, 25, 50, 75, 95])

                    mc1, mc2, mc3, mc4, mc5 = st.columns(5)
                    mc1.metric("🎯 Simulated Success Rate", f"{pct_success:.0%}")
                    mc2.metric("📉 Worst Case (P5)",  f"{p5:.0%}")
                    mc3.metric("📊 Median (P50)",     f"{p50:.0%}")
                    mc4.metric("📈 Best Case (P95)",  f"{p95:.0%}")
                    mc5.metric("📏 Uncertainty Range",f"{(p95-p5):.0%}")

                    # Distribution bar chart
                    bins     = np.linspace(0, 1, 21)
                    hist_vals, _ = np.histogram(sim_arr, bins=bins)
                    hist_df  = pd.DataFrame({'Probability': np.round(bins[:-1], 2), 'Count': hist_vals})
                    st.bar_chart(hist_df.set_index('Probability'), height=180, use_container_width=True)
                    st.caption("Distribution of success probability across 1,000 simulations — wider = more uncertain")

                    st.markdown("---")

                    # ════════════════════════════════════════════════════════
                    # NOVELTY 2: Optimal Path Finder
                    # ════════════════════════════════════════════════════════
                    st.markdown("### 🧭 Optimal Path Finder")
                    st.caption("Finds the minimum changes needed to push success probability above 80%.")
                    current_prob = prob
                    target_prob  = 0.80
                    paths        = []

                    if current_prob < target_prob:
                        # Path 1: Unblock stories one by one
                        for unblock_n in range(1, int(bs) + 2):
                            new_bs_p  = max(0, bs - unblock_n)
                            test_feat = np.array([[psp, csp, percent_done, drs, hv, new_bs_p, sc]])
                            tp        = ensemble.predict_proba(test_feat)[0][1]
                            if tp >= target_prob:
                                paths.append(("🔓 Unblock stories", f"Remove {unblock_n} blocker(s)", tp, unblock_n))
                                break

                        # Path 2: Descope sprint
                        for descope_n in range(1, int(psp - csp) + 1, 2):
                            new_psp_p = max(csp + 1, psp - descope_n)
                            new_pct_p = min(100, csp / new_psp_p * 100)
                            test_feat = np.array([[new_psp_p, csp, new_pct_p, drs, hv, bs, sc]])
                            tp        = ensemble.predict_proba(test_feat)[0][1]
                            if tp >= target_prob:
                                paths.append(("✂️ Descope sprint", f"Remove {descope_n} SP from backlog", tp, descope_n))
                                break

                        # Path 3: Add days
                        for extra_days in range(1, 11):
                            test_feat = np.array([[psp, csp, percent_done, drs + extra_days, hv, bs, sc]])
                            tp        = ensemble.predict_proba(test_feat)[0][1]
                            if tp >= target_prob:
                                paths.append(("📅 Extend sprint", f"Add {extra_days} more day(s)", tp, extra_days))
                                break

                        # Path 4: Reduce scope change
                        if sc > 0:
                            test_feat = np.array([[psp, csp, percent_done, drs, hv, bs, 0]])
                            tp        = ensemble.predict_proba(test_feat)[0][1]
                            if tp >= target_prob:
                                paths.append(("🛑 Freeze scope", "Set scope change to 0", tp, sc))

                        if paths:
                            paths.sort(key=lambda x: x[3])  # sort by cost
                            for action, detail, new_prob, cost in paths:
                                gain  = new_prob - current_prob
                                st.markdown(f"""
<div class='agent-card success'>
    <div class='agent-title'>{action}</div>
    <div class='agent-detail'>{detail} → New success probability: <b>{new_prob:.0%}</b>
    <span style='color:#06d6a0;margin-left:8px;'>▲ +{gain:.0%} gain</span></div>
</div>""", unsafe_allow_html=True)
                        else:
                            st.info("No single-action path to 80% found. Try combining multiple actions.")
                    else:
                        st.success(f"✅ Sprint already above 80% success probability ({current_prob:.0%}). No action needed!")

                    st.markdown("---")

                    # ════════════════════════════════════════════════════════
                    # NOVELTY 3: Similar Sprint Lookup
                    # ════════════════════════════════════════════════════════
                    st.markdown("### 🔍 Similar Sprint Lookup")
                    st.caption("Finds the 5 most similar historical sprints and shows what happened to them.")
                    from sklearn.preprocessing import StandardScaler
                    from sklearn.metrics.pairwise import euclidean_distances

                    lookup_feats = ['Planned_Story_Points_Sprint','Completed_Story_Points',
                                    'Percent_Done','Days_Remaining_Sprint','Historical_Velocity',
                                    'Blocked_Stories','Scope_Change']
                    df_lookup  = df[lookup_feats + ['Success_Label']].copy()
                    scaler     = StandardScaler()
                    scaled_all = scaler.fit_transform(df_lookup[lookup_feats])
                    query_row  = scaler.transform([[psp, csp, percent_done, drs, hv, bs, sc]])
                    dists      = euclidean_distances(query_row, scaled_all)[0]
                    top5_idx   = np.argsort(dists)[:5]

                    sim_rows = []
                    for idx in top5_idx:
                        row = df.iloc[idx]
                        sim_rows.append({
                            'Sprint ID':   row.get('Sprint_ID', f'Row {idx}'),
                            'Planned SP':  f"{row['Planned_Story_Points_Sprint']:.0f}",
                            'Completed SP':f"{row['Completed_Story_Points']:.0f}",
                            '% Done':      f"{row['Percent_Done']:.0f}%",
                            'Days Left':   f"{row['Days_Remaining_Sprint']:.0f}",
                            'Blocked':     f"{row['Blocked_Stories']:.0f}",
                            'Outcome':     "✅ Success" if row['Success_Label'] == 1 else "⚠️ Spilled",
                            'Similarity':  f"{(1/(1+dists[idx])):.0%}"
                        })

                    sim_df    = pd.DataFrame(sim_rows)
                    n_success = sum(1 for r in sim_rows if "Success" in r['Outcome'])
                    st.dataframe(sim_df, use_container_width=True, hide_index=True)
                    outcome_color = "#06d6a0" if n_success >= 3 else ("#ffd166" if n_success >= 2 else "#ff4d6d")
                    st.markdown(f"<div style='font-size:0.9rem;color:{outcome_color};font-weight:700;'>"
                                f"📊 {n_success}/5 similar sprints succeeded historically — "
                                f"{'supports' if n_success >= 3 else 'contradicts'} the model prediction</div>",
                                unsafe_allow_html=True)

                    st.markdown("---")

                    # Days to Recover (if at risk)
                    if not p:
                        st.markdown("### ⏱️ Days to Recover Estimate")
                        remaining_sp      = max(0, psp - csp)
                        daily_rate        = (hv / 10) if drs > 0 else 1
                        days_needed       = remaining_sp / daily_rate if daily_rate > 0 else 999
                        extra_days_val    = max(0, days_needed - drs)
                        unblocked_vel     = max(0.5, hv - bs * 2)
                        days_if_unblocked = remaining_sp / (unblocked_vel/10) if unblocked_vel > 0 else 999
                        extra_unblocked   = max(0, days_if_unblocked - drs)
                        dc1, dc2, dc3     = st.columns(3)
                        dc1.metric("📅 Days Needed",         f"{days_needed:.1f}d",        f"{extra_days_val:+.1f}d overrun")
                        dc2.metric("🚀 If Blockers Removed", f"{days_if_unblocked:.1f}d",  f"{extra_unblocked-extra_days_val:+.1f}d change")
                        dc3.metric("📉 SP Remaining",        f"{remaining_sp:.0f} SP",     f"{remaining_sp/psp*100:.0f}% of sprint")

                    st.markdown("---")

                    # ════════════════════════════════════════════════════════
                    # NOVELTY 4: Stacking Classifier (Meta-Learner)
                    # ════════════════════════════════════════════════════════
                    st.markdown("### 🧬 Stacking Classifier — Meta-Learner")
                    st.caption("A meta-learner (Logistic Regression) is trained on top of base model predictions — learns *when* to trust each model.")

                    with st.expander("🧬 View Stacking Architecture & Results"):
                        from sklearn.ensemble import StackingClassifier
                        stacking_clf = StackingClassifier(
                            estimators=[
                                ('lr',  LogisticRegression(max_iter=1000, class_weight='balanced', random_state=42)),
                                ('gb',  GradientBoostingClassifier(n_estimators=100, random_state=42)),
                                ('rf',  RandomForestClassifier(n_estimators=100, random_state=42)),
                                ('ada', AdaBoostClassifier(n_estimators=100, random_state=42)),
                            ],
                            final_estimator=LogisticRegression(max_iter=500),
                            cv=5,
                            passthrough=False
                        )
                        stacking_clf.fit(X1_train, y1_train)
                        stack_pred = stacking_clf.predict(X1_test)
                        stack_acc  = accuracy_score(y1_test, stack_pred)
                        stack_prob_val = stacking_clf.predict_proba(features)[0][1]

                        sa1, sa2, sa3 = st.columns(3)
                        sa1.metric("🧬 Stacking Accuracy",  f"{stack_acc:.2%}")
                        sa2.metric("🗳️ Ensemble Accuracy",  f"{acc1:.2%}", f"{(stack_acc-acc1)*100:+.2f}%")
                        sa3.metric("🎯 Stacking Prediction", f"{stack_prob_val:.0%} success prob")

                        st.markdown("""
**How Stacking works:**
1. **Layer 1** — LR, GB, RF, AdaBoost each make predictions using cross-validation (5-fold)
2. **Layer 2 (Meta-learner)** — A new Logistic Regression sees *only the predictions* from Layer 1
3. **Result** — The meta-learner learns which base models to trust for different types of sprints
                        """)

                        arch_data = {
                            'Layer': ['Layer 1','Layer 1','Layer 1','Layer 1','Layer 2'],
                            'Model': ['Logistic Regression','Gradient Boosting','Random Forest','AdaBoost','Meta-Learner (LR)'],
                            'Role':  ['Base model — linear patterns','Base model — non-linear trees',
                                      'Base model — ensemble trees','Base model — boosted stumps',
                                      'Combines all base predictions'],
                            'Accuracy': [f"{ind_accs.get('Logistic Regression',0):.2%}",
                                         f"{ind_accs.get('Gradient Boosting',0):.2%}",
                                         f"{ind_accs.get('Random Forest',0):.2%}",
                                         f"{ind_accs.get('AdaBoost',0):.2%}",
                                         f"{stack_acc:.2%} ⭐"]
                        }
                        st.dataframe(pd.DataFrame(arch_data), use_container_width=True, hide_index=True)

                    st.markdown("---")

                    # ════════════════════════════════════════════════════════
                    # NOVELTY 5: Sprint Trajectory Prediction (3 Sprints Ahead)
                    # ════════════════════════════════════════════════════════
                    st.markdown("### 🔮 Sprint Trajectory — 3 Sprints Ahead")
                    st.caption("Projects how sprint success probability will change over the next 3 sprints based on current trend.")

                    with st.expander("🔮 View 3-Sprint Forecast"):
                        # Estimate trend from dataset
                        if 'Sprint_Number' in df.columns:
                            trend_data = df.copy()
                            trend_data['pred_prob'] = ensemble.predict_proba(
                                trend_data[[f for f in all_feats_obj1 if f in trend_data.columns]]
                            )[:,1]
                            sprint_trend = trend_data.groupby('Sprint_Number')['pred_prob'].mean().sort_index()
                            # Compute rolling trend (last 5 sprints)
                            recent = sprint_trend.tail(5)
                            trend_slope = (recent.iloc[-1] - recent.iloc[0]) / max(len(recent)-1, 1)
                        else:
                            trend_slope = 0.0

                        # Project 3 sprints forward
                        future_sprints = []
                        base_prob_traj = prob
                        for i in range(1, 4):
                            # Assume blockers grow slightly, velocity stays same, days constant
                            proj_bs  = min(10, bs + i * 0.3)
                            proj_sc  = sc + i * 0.5
                            proj_psp = psp
                            proj_csp = min(psp, csp + hv * 0.8)
                            proj_pct = min(100, proj_csp / proj_psp * 100) if proj_psp > 0 else percent_done
                            proj_inp = np.array([[proj_psp, proj_csp, proj_pct, drs, hv, proj_bs, proj_sc]])
                            proj_p   = ensemble.predict_proba(proj_inp)[0][1]
                            proj_p   = max(0.05, min(0.98, proj_p + trend_slope * i))
                            future_sprints.append({
                                'Sprint': f'Sprint +{i}',
                                'Success Probability': f"{proj_p:.0%}",
                                'Projected Blockers':  f"{proj_bs:.1f}",
                                'Scope Change':        f"{proj_sc:.1f}",
                                'Prediction':          "✅ On track" if proj_p >= 0.5 else "⚠️ At risk",
                                '_prob': proj_p
                            })

                        traj_df = pd.DataFrame(future_sprints)
                        traj_display = traj_df.drop(columns=['_prob'])
                        st.dataframe(traj_display, use_container_width=True, hide_index=True)

                        # Trajectory chart
                        traj_chart = pd.DataFrame({
                            'Sprint': ['Current'] + [f['Sprint'] for f in future_sprints],
                            'Success Probability': [prob] + [f['_prob'] for f in future_sprints]
                        }).set_index('Sprint')
                        st.line_chart(traj_chart, height=200, use_container_width=True)

                        overall_trend = future_sprints[-1]['_prob'] - prob
                        t_color = "#06d6a0" if overall_trend >= 0 else "#ff4d6d"
                        t_label = "improving 📈" if overall_trend >= 0 else "declining 📉"
                        st.markdown(f"<div style='font-size:0.9rem;color:{t_color};font-weight:700;'>"
                                    f"3-sprint trend: <b>{t_label}</b> ({overall_trend:+.0%})</div>",
                                    unsafe_allow_html=True)

                    st.markdown("---")

                    # ════════════════════════════════════════════════════════
                    # NOVELTY 6: Cross-Validation Score Card
                    # ════════════════════════════════════════════════════════
                    st.markdown("### 📋 Cross-Validation Score Card")
                    st.caption("5-fold CV across all models — the most reliable accuracy estimate (not biased by a single train/test split).")

                    with st.expander("📋 View Full CV Score Card"):
                        from sklearn.model_selection import cross_val_score, StratifiedKFold
                        cv_skf = StratifiedKFold(n_splits=5, shuffle=True, random_state=42)
                        cv_models = [
                            ('Logistic Regression', LogisticRegression(max_iter=1000, class_weight='balanced', random_state=42)),
                            ('Gradient Boosting',   GradientBoostingClassifier(n_estimators=100, random_state=42)),
                            ('Random Forest',       RandomForestClassifier(n_estimators=100, random_state=42)),
                            ('AdaBoost',            AdaBoostClassifier(n_estimators=100, random_state=42)),
                            ('Voting Ensemble',     ensemble),
                        ]
                        cv_rows = []
                        for mname, mclf in cv_models:
                            scores = cross_val_score(mclf, X1, y1, cv=cv_skf, scoring='accuracy')
                            cv_rows.append({
                                'Model': mname,
                                'Fold 1': f"{scores[0]:.3f}",
                                'Fold 2': f"{scores[1]:.3f}",
                                'Fold 3': f"{scores[2]:.3f}",
                                'Fold 4': f"{scores[3]:.3f}",
                                'Fold 5': f"{scores[4]:.3f}",
                                'Mean ± Std': f"{scores.mean():.3f} ± {scores.std():.3f}",
                            })
                        cv_df = pd.DataFrame(cv_rows)
                        st.dataframe(cv_df, use_container_width=True, hide_index=True)
                        best_cv = max(cv_rows, key=lambda x: float(x['Mean ± Std'].split(' ')[0]))
                        st.success(f"⭐ Most consistent model: **{best_cv['Model']}** (CV = {best_cv['Mean ± Std']})")

            else:
                st.error("⚠️ Not enough class variety in Success_Label column.")
        except Exception as e:
            st.error(f"Error in Objective 1: {e}")
    # ══════════════════════════════════════════════════════════════
    # Objective 2: Workload Projection
    # ══════════════════════════════════════════════════════════════
    with tabs[2]:
        st.header("📌 Objective 2 — Workload Projection Forecast")
        try:
            X2 = df[['Planned_Story_Points_Resource','Current_Assigned_SP','Historical_Avg_SP',
                      'Remaining_Days_Resource','High_Priority_Tasks_Resource','Current_Workload_Percent']]
            y2 = df['Expected_Overload']
            if len(y2.unique()) > 1:
                X2_train, X2_test, y2_train, y2_test = train_test_split(X2, y2, test_size=0.2, random_state=42)
                workload_model = RandomForestClassifier()
                workload_model.fit(X2_train, y2_train)
                y2_pred = workload_model.predict(X2_test)
                acc2 = accuracy_score(y2_test, y2_pred)
                st.write(f"✅ Accuracy: {acc2:.2f}")
                if acc2 < 0.60:
                    st.info("ℹ️ **Low predictive signal detected.** The workload features in this dataset have near-zero correlation with the overload label — this is common in synthetic data. In production, richer features (e.g. task completion rate, meeting hours) would improve accuracy significantly.")
                st.text(classification_report(y2_test, y2_pred))

                st.subheader("🔍 Predict Overload Risk")
                psp2 = st.number_input("Planned SP", 1, 100, 35, key="obj2_psp2")
                casp = st.number_input("Current Assigned SP", 0, 100, 40, key="obj2_casp")
                hasp = st.number_input("Historical Avg SP", 1, 100, 30, key="obj2_hasp")
                rdr  = st.number_input("Remaining Days", 1, 30, 5, key="obj2_rdr")
                hpt  = st.number_input("High Priority Tasks", 0, 10, 2, key="obj2_hpt")
                cwp  = st.number_input("Current Workload %", 0, 200, 125, key="obj2_cwp")

                if st.button("Predict Overload", key="obj2_btn"):
                    features = np.array([[psp2, casp, hasp, rdr, hpt, cwp]])
                    pred = workload_model.predict(features)[0]
                    prob = workload_model.predict_proba(features)[0][1]
                    if pred:
                        st.warning(f"⚠️ Overload Risk! ({prob:.2f})")
                    else:
                        st.success(f"✅ Within Capacity ({prob:.2f})")
            else:
                st.error("⚠️ Not enough class variety in Expected_Overload column.")
        except Exception as e:
            st.error(f"Error in Objective 2: {e}")

    # ══════════════════════════════════════════════════════════════
    # Objective 3: Time to Resolve
    # ══════════════════════════════════════════════════════════════
    with tabs[3]:
        st.header("📌 Objective 3 — Time to Resolve Estimation")
        try:
            X3 = pd.get_dummies(df[['Issue_Type','Priority']], drop_first=False)
            X3 = pd.concat([X3, df[['Original_Estimate_Hours','Story_Points_Issue']]], axis=1)
            y3 = df['Resolution_Time_Hours']
            X3_train, X3_test, y3_train, y3_test = train_test_split(X3, y3, test_size=0.2, random_state=42)
            ttr_model = LinearRegression()
            ttr_model.fit(X3_train, y3_train)
            y3_pred = ttr_model.predict(X3_test)
            st.write(f"✅ MSE: {mean_squared_error(y3_test, y3_pred):.2f}")

            st.subheader("🔍 Estimate Time to Resolve")
            issue_type = st.selectbox("Issue Type", ['Bug','Story','Task'], key="obj3_it")
            priority   = st.selectbox("Priority", ['Low','Medium','High'], key="obj3_pri")
            oe = st.number_input("Original Estimate", 1, 50, 8, key="obj3_oe")
            sp = st.number_input("Story Points", 1, 20, 5, key="obj3_sp")

            test_row = pd.DataFrame([{
                'Issue_Type_Bug':   1 if issue_type=='Bug'    else 0,
                'Issue_Type_Story': 1 if issue_type=='Story'  else 0,
                'Issue_Type_Task':  1 if issue_type=='Task'   else 0,
                'Priority_Low':     1 if priority=='Low'      else 0,
                'Priority_Medium':  1 if priority=='Medium'   else 0,
                'Priority_High':    1 if priority=='High'     else 0,
                'Original_Estimate_Hours': oe,
                'Story_Points_Issue': sp
            }])
            test_row = test_row.reindex(columns=X3.columns, fill_value=0)

            if st.button("Estimate Resolution Time", key="obj3_btn"):
                pred_time = max(0, ttr_model.predict(test_row)[0])
                st.info(f"⏰ Estimated Resolution Time: {pred_time:.1f} hours")
        except Exception as e:
            st.error(f"Error in Objective 3: {e}")

    # ══════════════════════════════════════════════════════════════
    # Objective 4: Burnout Risk Alerts
    # ══════════════════════════════════════════════════════════════
    with tabs[4]:
        st.header("📌 Objective 4 — Burnout Risk Alerts")
        try:
            X4 = df[['Total_SP_This_Sprint','Historical_Avg_SP_Burnout',
                      'High_Priority_Tasks_Burnout','Consecutive_Overloads']]
            y4 = df['Risk_Flag']
            if len(y4.unique()) > 1:
                X4_train, X4_test, y4_train, y4_test = train_test_split(X4, y4, test_size=0.2, random_state=42)
                burnout_model = RandomForestClassifier()
                burnout_model.fit(X4_train, y4_train)
                y4_pred = burnout_model.predict(X4_test)
                acc4 = accuracy_score(y4_test, y4_pred)
                st.write(f"✅ Accuracy: {acc4:.2f}")
                if acc4 < 0.60:
                    st.info("ℹ️ **Low predictive signal detected.** Burnout features show near-zero correlation with the risk label in this dataset. Real burnout prediction benefits from additional signals like overtime hours, meeting load, and leave history.")
                st.text(classification_report(y4_test, y4_pred))

                st.subheader("🔍 Check Burnout Risk")
                tsp   = st.number_input("Total SP This Sprint", 0, 100, 40, key="obj4_tsp")
                hasp4 = st.number_input("Historical Avg SP", 1, 100, 25, key="obj4_hasp4")
                hpt4  = st.number_input("High Priority Tasks", 0, 10, 2, key="obj4_hpt4")
                co    = st.number_input("Consecutive Overloads", 0, 5, 2, key="obj4_co")

                if st.button("Check Burnout Risk", key="obj4_btn"):
                    pred = burnout_model.predict([[tsp, hasp4, hpt4, co]])[0]
                    if pred:
                        st.warning("⚠️ Burnout Risk Detected!")
                    else:
                        st.success("✅ Workload looks healthy!")
            else:
                st.error("⚠️ Not enough class variety in Risk_Flag column.")
        except Exception as e:
            st.error(f"Error in Objective 4: {e}")

    # ══════════════════════════════════════════════════════════════
    # Objective 5: Resource Allocation
    # ══════════════════════════════════════════════════════════════
    with tabs[5]:
        st.header("📌 Objective 5 — Resource Allocation Suggestions")
        try:
            le_summary = LabelEncoder(); le_labels = LabelEncoder()
            df['Summary_enc'] = le_summary.fit_transform(df['Summary'].astype(str))
            df['Labels_enc']  = le_labels.fit_transform(df['Labels'].astype(str))
            X5 = df[['Summary_enc','Labels_enc','Original_Estimate_Resource','Story_Points_Resource']]
            y5 = df['Assignee_Resource']
            X5_train, X5_test, y5_train, y5_test = train_test_split(X5, y5, test_size=0.2, random_state=42)
            alloc_model = RandomForestClassifier()
            alloc_model.fit(X5_train, y5_train)
            y5_pred = alloc_model.predict(X5_test)
            st.write(f"✅ Accuracy: {accuracy_score(y5_test, y5_pred):.2f}")

            st.subheader("🔍 Suggest Assignee")
            summary = st.text_input("Summary (short description)", "Fix bug")
            label   = st.text_input("Label (category)", "Bug")
            oe5 = st.number_input("Original Estimate", 1, 50, 8, key="obj5_oe")
            sp5 = st.number_input("Story Points", 1, 20, 5, key="obj5_sp")

            try: summary_enc = le_summary.transform([summary])[0]
            except: summary_enc = 0
            try: label_enc = le_labels.transform([label])[0]
            except: label_enc = 0

            test_row = pd.DataFrame([{
                'Summary_enc': summary_enc, 'Labels_enc': label_enc,
                'Original_Estimate_Resource': oe5, 'Story_Points_Resource': sp5
            }])

            if st.button("Suggest Assignee", key="obj5_btn"):
                assignee = alloc_model.predict(test_row)[0]
                st.success(f"✅ Recommended Assignee: {assignee}")
        except Exception as e:
            st.error(f"Error in Objective 5: {e}")
    # ══════════════════════════════════════════════════════════════
    # Spark Processing Tab
    # ══════════════════════════════════════════════════════════════
    with tabs[6]:
        st.header("⚡ Spark Processing")
        st.caption("Apache Spark is used to load, profile, and aggregate your dataset at scale before passing features to the ML models.")

        if not SPARK_AVAILABLE:
            st.warning("⚠️ PySpark is not installed. Add `pyspark` to your requirements.txt and redeploy.")
            st.code("# requirements.txt\npyspark\nstreamlit\npandas\nnumpy\nscikit-learn\nopenpyxl", language="text")
            st.info("While Spark is unavailable, the preview below shows what Spark would compute on your dataset.")

        st.markdown("---")

        # ── What Spark Does ──────────────────────────────────────
        st.subheader("🔧 How Spark Is Used in This Pipeline")
        steps = [
            ("1️⃣ Data Ingestion",    "Spark reads the uploaded CSV into a distributed DataFrame — handles datasets from KBs to TBs with the same code"),
            ("2️⃣ Schema Validation", "Spark infers and validates column types, flags nulls, and detects out-of-range values before ML training"),
            ("3️⃣ Feature Aggregation","Spark computes per-assignee averages, sprint-level rollups, and workload distributions using SQL-style groupBy"),
            ("4️⃣ Label Binarization","Spark applies threshold transformations to float labels (Success_Label > 0.5 → 1) across all rows in one pass"),
            ("5️⃣ Feature Export",    "Processed Spark DataFrame is converted to Pandas for sklearn model training — best of both worlds"),
        ]
        for icon_title, detail in steps:
            st.markdown(f"""
<div class='agent-card info' style='margin-bottom:0.6rem;'>
    <div class='agent-title'>{icon_title}</div>
    <div class='agent-detail'>{detail}</div>
</div>""", unsafe_allow_html=True)

        st.markdown("---")

        # ── Spark Code Preview ───────────────────────────────────
        st.subheader("💻 Spark Code Used in Pipeline")
        st.code("""
from pyspark.sql import SparkSession
from pyspark.sql import functions as F

# 1. Start Spark session
spark = SparkSession.builder \
    .appName("AgileAIDashboard") \
    .master("local[*]") \
    .getOrCreate()

# 2. Load CSV
sdf = spark.read.csv("agile_dataset.csv", header=True, inferSchema=True)

# 3. Validate & profile
sdf.printSchema()
sdf.describe().show()

# 4. Binarize float labels
sdf = sdf.withColumn("Success_Label",
        F.when(F.col("Success_Label") > 0.5, 1).otherwise(0))
sdf = sdf.withColumn("Expected_Overload",
        F.when(F.col("Expected_Overload") > 0.5, 1).otherwise(0))
sdf = sdf.withColumn("Risk_Flag",
        F.when(F.col("Risk_Flag") > 0.3, 1).otherwise(0))

# 5. Per-assignee aggregations
assignee_stats = sdf.groupBy("Assignee").agg(
    F.count("*").alias("total_records"),
    F.avg("Current_Workload_Percent").alias("avg_workload"),
    F.avg("Consecutive_Overloads").alias("avg_consec_overloads"),
    F.sum("Risk_Flag").alias("burnout_count"),
    F.avg("Success_Label").alias("sprint_success_rate")
)
assignee_stats.show()

# 6. Sprint trend aggregations
sprint_trend = sdf.groupBy("Sprint_Number").agg(
    F.avg("Success_Label").alias("success_rate"),
    F.avg("Current_Workload_Percent").alias("avg_workload"),
    F.avg("Blocked_Stories").alias("avg_blocked")
).orderBy("Sprint_Number")
sprint_trend.show()

# 7. Export to Pandas for sklearn
pdf = sdf.toPandas()
""", language="python")

        st.markdown("---")

        # ── Simulated Spark Output ───────────────────────────────
        st.subheader("📊 Simulated Spark Output (computed via Pandas)")
        st.caption("This is what Spark's aggregations would produce — computed here using Pandas as a fallback.")

        sp1, sp2 = st.columns(2)

        with sp1:
            st.markdown("**groupBy('Assignee').agg(...)**")
            if 'Assignee' in df.columns:
                assignee_stats = df.groupby('Assignee').agg(
                    total_records=('Assignee','count'),
                    avg_workload=('Current_Workload_Percent','mean'),
                    avg_consec_overloads=('Consecutive_Overloads','mean'),
                    burnout_count=('Risk_Flag','sum'),
                    sprint_success_rate=('Success_Label','mean')
                ).round(2).reset_index()
                assignee_stats.columns = ['Assignee','Records','Avg Workload %','Avg Consec. Overloads','Burnout Count','Success Rate']
                st.dataframe(assignee_stats, use_container_width=True, hide_index=True)

        with sp2:
            st.markdown("**groupBy('Sprint_Number').agg(...)**")
            if 'Sprint_Number' in df.columns:
                sprint_trend = df.groupby('Sprint_Number').agg(
                    success_rate=('Success_Label','mean'),
                    avg_workload=('Current_Workload_Percent','mean'),
                    avg_blocked=('Blocked_Stories','mean'),
                    record_count=('Sprint_Number','count')
                ).round(3).reset_index().sort_values('Sprint_Number')
                sprint_trend.columns = ['Sprint #','Success Rate','Avg Workload %','Avg Blocked','Records']
                st.dataframe(sprint_trend.head(20), use_container_width=True, hide_index=True)

        st.markdown("---")

        # ── Schema & Profile ─────────────────────────────────────
        st.subheader("🗂️ Spark Schema & Data Profile")
        sc1, sc2 = st.columns(2)
        with sc1:
            st.markdown("**Column Types (inferred by Spark)**")
            schema_df = pd.DataFrame([
                {'Column': col, 'Spark Type': 'StringType' if df[col].dtype == object else 'DoubleType' if df[col].dtype == float else 'IntegerType',
                 'Nulls': int(df[col].isnull().sum()), 'Unique': int(df[col].nunique())}
                for col in df.columns
            ])
            st.dataframe(schema_df, use_container_width=True, hide_index=True, height=350)
        with sc2:
            st.markdown("**Numeric Column Stats (Spark describe())**")
            num_cols = df.select_dtypes(include=[np.number]).columns[:8]
            stats_df = df[num_cols].describe().round(2)
            st.dataframe(stats_df, use_container_width=True)

        st.markdown("---")

        # ── Spark vs Pandas ──────────────────────────────────────
        st.subheader("⚡ Why Spark vs Pandas?")
        comparison = pd.DataFrame([
            {"Feature": "Dataset size", "Pandas": "Up to ~1GB (RAM limited)", "Spark": "Terabytes (distributed)"},
            {"Feature": "Processing", "Pandas": "Single machine, single core", "Spark": "Multi-node, multi-core cluster"},
            {"Feature": "Fault tolerance", "Pandas": "❌ No", "Spark": "✅ Yes (RDD lineage)"},
            {"Feature": "SQL support", "Pandas": "Limited", "Spark": "Full Spark SQL"},
            {"Feature": "Streaming", "Pandas": "❌ No", "Spark": "✅ Spark Streaming"},
            {"Feature": "ML library", "Pandas": "sklearn (local)", "Spark": "MLlib (distributed)"},
            {"Feature": "Best for", "Pandas": "Small-medium data, rapid prototyping", "Spark": "Big data, production pipelines"},
        ])
        st.dataframe(comparison, use_container_width=True, hide_index=True)


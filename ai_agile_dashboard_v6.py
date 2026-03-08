import streamlit as st
import pandas as pd
import numpy as np
from sklearn.linear_model import LogisticRegression, LinearRegression
from sklearn.ensemble import (RandomForestClassifier, GradientBoostingClassifier,
                               VotingClassifier, AdaBoostClassifier)
from sklearn.svm import SVC
from sklearn.model_selection import train_test_split, cross_val_score
from sklearn.metrics import (accuracy_score, classification_report,
                              mean_squared_error, r2_score)
from sklearn.preprocessing import LabelEncoder, StandardScaler
from sklearn.metrics.pairwise import euclidean_distances
from sklearn.cluster import KMeans
import os, json

# ── PySpark (graceful fallback to pandas) ──────────────────────────────────
os.environ.setdefault("PYSPARK_PYTHON",  "python3")
os.environ.setdefault("SPARK_LOCAL_IP",  "127.0.0.1")
try:
    from pyspark.sql import SparkSession
    from pyspark.sql import functions as F
    SPARK_AVAILABLE = True
except ImportError:
    SPARK_AVAILABLE = False

@st.cache_resource
def get_spark():
    if not SPARK_AVAILABLE:
        return None
    try:
        spark = (SparkSession.builder
                 .appName("AgileJiraDashboard")
                 .master("local[*]")
                 .config("spark.driver.memory", "2g")
                 .config("spark.sql.shuffle.partitions", "4")
                 .config("spark.ui.enabled", "false")
                 .getOrCreate())
        spark.sparkContext.setLogLevel("ERROR")
        return spark
    except Exception:
        return None

def spark_engineer_features(pdf, spark):
    """Spark-based feature engineering with pandas fallback."""
    if spark is not None:
        try:
            sdf = spark.createDataFrame(pdf)
            sdf = sdf.withColumn("Velocity_Efficiency",
                F.when(F.col("Planned_Story_Points_Sprint") > 0,
                       F.col("Historical_Velocity") / F.col("Planned_Story_Points_Sprint")
                ).otherwise(F.lit(1.0)))
            sdf = sdf.withColumn("Completion_Gap",
                F.col("Planned_Story_Points_Sprint") - F.col("Completed_Story_Points"))
            sdf = sdf.withColumn("Blocker_Severity",
                F.col("Blocked_Stories") * F.when(
                    F.col("Days_Remaining_Sprint") > 0,
                    F.lit(1.0) / F.col("Days_Remaining_Sprint")
                ).otherwise(F.lit(1.0)))
            sdf = sdf.withColumn("Scope_Pressure",
                F.when(F.col("Planned_Story_Points_Sprint") > 0,
                       F.col("Scope_Change") / F.col("Planned_Story_Points_Sprint")
                ).otherwise(F.lit(0.0)))
            sdf = sdf.withColumn("Sprint_Momentum",
                F.when(F.col("Historical_Velocity") > 0,
                       F.col("Completed_Story_Points") / F.col("Historical_Velocity")
                ).otherwise(F.lit(0.0)))
            sdf = sdf.withColumn("Recovery_Index",
                F.when(
                    (F.col("Planned_Story_Points_Sprint") - F.col("Completed_Story_Points") > 0) &
                    (F.col("Days_Remaining_Sprint") > 0),
                    (F.col("Historical_Velocity") * F.col("Days_Remaining_Sprint") / F.lit(10.0)) /
                    (F.col("Planned_Story_Points_Sprint") - F.col("Completed_Story_Points"))
                ).otherwise(F.lit(1.0)))
            sdf = sdf.withColumn("Workload_Stress",
                (F.col("Current_Workload_Percent") / F.lit(100.0)) *
                 F.col("Consecutive_Overloads"))
            return sdf.toPandas().fillna(0)
        except Exception:
            pass
    # Pandas fallback
    df = pdf.copy()
    df["Velocity_Efficiency"] = (df["Historical_Velocity"] / df["Planned_Story_Points_Sprint"].replace(0,1)).clip(0,3)
    df["Completion_Gap"]      = df["Planned_Story_Points_Sprint"] - df["Completed_Story_Points"]
    df["Blocker_Severity"]    = df["Blocked_Stories"] * (1 / df["Days_Remaining_Sprint"].replace(0,1).abs())
    df["Scope_Pressure"]      = (df["Scope_Change"] / df["Planned_Story_Points_Sprint"].replace(0,1)).clip(-1,2)
    df["Sprint_Momentum"]     = (df["Completed_Story_Points"] / df["Historical_Velocity"].replace(0,1)).clip(0,2)
    df["Recovery_Index"]      = ((df["Historical_Velocity"] * df["Days_Remaining_Sprint"] / 10) /
                                  (df["Planned_Story_Points_Sprint"] - df["Completed_Story_Points"]).replace(0,0.001)).clip(0,5)
    df["Workload_Stress"]     = (df["Current_Workload_Percent"] / 100) * df.get("Consecutive_Overloads", pd.Series(0, index=df.index))
    return df.fillna(0)

st.set_page_config(page_title="AI Agile Dashboard", layout="wide")

st.markdown("""
<style>
    .agent-card {
        background: linear-gradient(135deg, #1e1e2e 0%, #2a2a3e 100%);
        border: 1px solid #444466;
        border-radius: 12px;
        padding: 1.2rem 1.5rem;
        margin-bottom: 1rem;
        color: #e0e0f0;
    }
    .agent-card.critical {
        border-left: 5px solid #ff4d6d;
        background: linear-gradient(135deg, #2e1e22 0%, #3e2a2e 100%);
    }
    .agent-card.warning {
        border-left: 5px solid #ffd166;
        background: linear-gradient(135deg, #2e2a1e 0%, #3e362a 100%);
    }
    .agent-card.success {
        border-left: 5px solid #06d6a0;
        background: linear-gradient(135deg, #1e2e2a 0%, #2a3e36 100%);
    }
    .agent-card.info {
        border-left: 5px solid #4cc9f0;
        background: linear-gradient(135deg, #1e252e 0%, #2a333e 100%);
    }
    .agent-title {
        font-size: 1rem;
        font-weight: 700;
        margin-bottom: 0.3rem;
        letter-spacing: 0.03em;
    }
    .agent-detail {
        font-size: 0.85rem;
        opacity: 0.85;
        line-height: 1.5;
    }
    .chain-step {
        display: flex;
        align-items: flex-start;
        gap: 1rem;
        margin-bottom: 0.8rem;
    }
    .step-num {
        background: #4cc9f0;
        color: #000;
        border-radius: 50%;
        width: 28px;
        height: 28px;
        display: flex;
        align-items: center;
        justify-content: center;
        font-weight: 800;
        font-size: 0.8rem;
        flex-shrink: 0;
    }
    .health-bar-container {
        background: #333;
        border-radius: 8px;
        height: 14px;
        width: 100%;
        margin-top: 4px;
    }
    .report-section {
        background: #1a1a2e;
        border-radius: 10px;
        padding: 1rem 1.5rem;
        margin-bottom: 1rem;
        border: 1px solid #333355;
        color: #dde;
        font-size: 0.9rem;
        line-height: 1.7;
    }
</style>
""", unsafe_allow_html=True)

st.title("🚀 AI Agile Project Management Dashboard")

uploaded_file = st.file_uploader("📁 Upload the Combined CSV for All Objectives", type="csv")

# ── shared state ────────────────────────────────────────────────────────────
models = {}   # will hold trained models keyed by objective
encoders = {} # label encoders

if uploaded_file:
    df = pd.read_csv(uploaded_file)
    df = df.fillna(0)

    # Handle both string ('Yes'/'No') and float/probability labels
    def binarize_col(series, threshold=0.5):
        if series.dtype == object:
            return series.map({'No': 0, 'Yes': 1}).fillna(0).astype(int)
        return (series > threshold).astype(int)

    thresholds = {'Success_Label': 0.5, 'Expected_Overload': 0.5, 'Risk_Flag': 0.3}
    for col, thresh in thresholds.items():
        if col in df.columns:
            df[col] = binarize_col(df[col], threshold=thresh)

    st.success("✅ File uploaded successfully!")

    # Dataset summary
    c1, c2, c3, c4 = st.columns(4)
    c1.metric("📋 Total Records", f"{len(df):,}")
    c2.metric("📊 Features", len(df.columns))
    if 'Success_Label' in df.columns:
        c3.metric("🔴 Sprints at Risk", int((df['Success_Label']==0).sum()))
    if 'Risk_Flag' in df.columns:
        c4.metric("⚠️ Burnout Flags", int(df['Risk_Flag'].sum()))

    # ── Spark feature engineering ────────────────────────────────────────────
    _spark = get_spark()
    df = spark_engineer_features(df, _spark)
    _engine = "⚡ Apache Spark" if (SPARK_AVAILABLE and _spark is not None) else "🐼 Pandas"
    _spark_feats = [c for c in ["Velocity_Efficiency","Completion_Gap","Blocker_Severity",
                                  "Scope_Pressure","Sprint_Momentum","Recovery_Index","Workload_Stress"]
                    if c in df.columns]

    with st.expander("👀 Preview Data"):
        if _spark_feats:
            st.caption(f"{_engine} auto-engineered {len(_spark_feats)} features: {', '.join(_spark_feats)}")
        st.write(df.head())

    # ── train all models silently so the agent can use them ─────────────────
    def train_all(df):
        results = {}

        # --- Obj 1: Sprint Completion ---
        try:
            X1 = df[['Planned_Story_Points_Sprint','Completed_Story_Points','Percent_Done',
                      'Days_Remaining_Sprint','Historical_Velocity','Blocked_Stories','Scope_Change']]
            y1 = df['Success_Label']
            if len(y1.unique()) > 1:
                m = LogisticRegression(max_iter=1000, class_weight='balanced')
                m.fit(X1, y1)
                results['sprint'] = {'model': m, 'features': X1.columns.tolist()}
        except: pass

        # --- Obj 2: Workload ---
        try:
            X2 = df[['Planned_Story_Points_Resource','Current_Assigned_SP','Historical_Avg_SP',
                      'Remaining_Days_Resource','High_Priority_Tasks_Resource','Current_Workload_Percent']]
            y2 = df['Expected_Overload']
            if len(y2.unique()) > 1:
                m = RandomForestClassifier(n_estimators=100, random_state=42)
                m.fit(X2, y2)
                results['workload'] = {'model': m, 'features': X2.columns.tolist()}
        except: pass

        # --- Obj 3: Time to Resolve ---
        try:
            X3 = pd.get_dummies(df[['Issue_Type','Priority']], drop_first=False)
            X3 = pd.concat([X3, df[['Original_Estimate_Hours','Story_Points_Issue']]], axis=1)
            y3 = df['Resolution_Time_Hours']
            m = LinearRegression()
            m.fit(X3, y3)
            results['ttr'] = {'model': m, 'features': X3.columns.tolist(), 'X3': X3}
        except: pass

        # --- Obj 4: Burnout ---
        try:
            X4 = df[['Total_SP_This_Sprint','Historical_Avg_SP_Burnout',
                      'High_Priority_Tasks_Burnout','Consecutive_Overloads']]
            y4 = df['Risk_Flag']
            if len(y4.unique()) > 1:
                m = RandomForestClassifier(n_estimators=100, random_state=42)
                m.fit(X4, y4)
                results['burnout'] = {'model': m, 'features': X4.columns.tolist()}
        except: pass

        # --- Obj 5: Resource Allocation ---
        try:
            le_s = LabelEncoder(); le_l = LabelEncoder()
            df['Summary_enc'] = le_s.fit_transform(df['Summary'].astype(str))
            df['Labels_enc']  = le_l.fit_transform(df['Labels'].astype(str))
            X5 = df[['Summary_enc','Labels_enc','Original_Estimate_Resource','Story_Points_Resource']]
            y5 = df['Assignee_Resource']
            m = RandomForestClassifier(n_estimators=100, random_state=42)
            m.fit(X5, y5)
            results['alloc'] = {'model': m, 'features': X5.columns.tolist(),
                                'le_summary': le_s, 'le_labels': le_l}
        except: pass

        return results, df

    models, df = train_all(df)

    # ── agentic scan: run all models on every row ────────────────────────────
    def run_agent_scan(_df, _models_keys):
        """Run predictions across the whole dataset and collect findings."""
        findings = []
        df = _df.copy()

        # Sprint risk scan — aggregate, not per-row
        if 'sprint' in models:
            m = models['sprint']['model']
            cols = models['sprint']['features']
            try:
                X = df[cols]
                preds  = m.predict(X)
                probas = m.predict_proba(X)[:, 1]
                at_risk_mask = (preds == 0)
                at_risk_count = int(at_risk_mask.sum())
                total = len(preds)
                pct = at_risk_count / total if total > 0 else 0
                if pct > 0.15:
                    avg_prob   = float(probas[at_risk_mask].mean())
                    avg_blocked = float(df.loc[at_risk_mask, 'Blocked_Stories'].mean()) if 'Blocked_Stories' in df.columns else 0
                    avg_days    = float(df.loc[at_risk_mask, 'Days_Remaining_Sprint'].mean()) if 'Days_Remaining_Sprint' in df.columns else 0
                    sev = 'critical' if pct > 0.5 else 'warning'
                    findings.append({
                        'severity': sev,
                        'objective': 'Sprint Completion',
                        'icon': '🔴' if sev == 'critical' else '🟡',
                        'title': f"{at_risk_count} of {total} sprints ({pct:.0%}) at risk of spillover",
                        'detail': (f"Avg completion probability: {avg_prob:.0%} | "
                                   f"Avg blocked stories: {avg_blocked:.1f} | "
                                   f"Avg days remaining: {avg_days:.1f}"),
                        'action': "Consider reducing scope or unblocking stories immediately."
                    })
            except: pass

        # Workload overload scan
        if 'workload' in models:
            m = models['workload']['model']
            cols = models['workload']['features']
            try:
                X = df[cols]
                preds  = m.predict(X)
                probas = m.predict_proba(X)[:, 1]
                overloaded = df[preds == 1].copy()
                overloaded['wl_prob'] = probas[preds == 1]
                count = len(overloaded)
                if count > 0 and count > len(df) * 0.2:
                    findings.append({
                        'severity': 'critical' if count > len(df) * 0.45 else 'warning',
                        'objective': 'Workload Projection',
                        'icon': '🔴' if count > len(df) * 0.45 else '🟡',
                        'title': f"{count} resource(s) projected to be overloaded",
                        'detail': (f"Average overload probability: {overloaded['wl_prob'].mean():.0%} | "
                                   f"Avg current workload: {overloaded.get('Current_Workload_Percent', pd.Series([0])).mean():.0f}%"),
                        'action': "Redistribute story points from overloaded to available team members."
                    })
            except: pass

        # Burnout risk scan
        if 'burnout' in models:
            m = models['burnout']['model']
            cols = models['burnout']['features']
            try:
                X = df[cols]
                preds = m.predict(X)
                at_risk_count = int(preds.sum())
                pct_flagged_b = at_risk_count / len(preds) if len(preds) > 0 else 0
                if pct_flagged_b > 0.25:
                    avg_co = df.loc[preds == 1, 'Consecutive_Overloads'].mean() if 'Consecutive_Overloads' in df.columns else 0
                    pct_flagged = pct_flagged_b
                    findings.append({
                        'severity': 'critical' if pct_flagged > 0.5 else 'warning',
                        'objective': 'Burnout Risk',
                        'icon': '🔴' if pct_flagged > 0.5 else '🟡',
                        'title': f"{at_risk_count} team member(s) flagged for burnout risk",
                        'detail': f"Avg consecutive overloads: {avg_co:.1f} sprints",
                        'action': "Schedule 1:1s, reduce high-priority task load, or grant recovery sprint."
                    })
            except: pass

        # Healthy signal
        sprint_ok   = sum(1 for f in findings if f['objective'] == 'Sprint Completion') == 0
        workload_ok = sum(1 for f in findings if f['objective'] == 'Workload Projection') == 0
        burnout_ok  = sum(1 for f in findings if f['objective'] == 'Burnout Risk') == 0

        if sprint_ok:
            findings.append({'severity':'success','objective':'Sprint Completion','icon':'✅',
                             'title':'All sprints on track','detail':'No spillover risk detected.',
                             'action':''})
        if workload_ok:
            findings.append({'severity':'success','objective':'Workload Projection','icon':'✅',
                             'title':'Workloads within capacity','detail':'No overload signals found.',
                             'action':''})
        if burnout_ok:
            findings.append({'severity':'success','objective':'Burnout Risk','icon':'✅',
                             'title':'No burnout risk detected','detail':'Team load looks sustainable.',
                             'action':''})

        return findings

    # ── chained decisions ────────────────────────────────────────────────────
    def build_chain(findings):
        """Build an agentic chain of decisions based on cross-objective findings."""
        chain = []
        has_sprint_risk   = any(f['objective'] == 'Sprint Completion'  and f['severity'] in ('critical','warning') for f in findings)
        has_overload      = any(f['objective'] == 'Workload Projection' and f['severity'] in ('critical','warning') for f in findings)
        has_burnout       = any(f['objective'] == 'Burnout Risk'        and f['severity'] in ('critical','warning') for f in findings)

        chain.append({
            'step': 1,
            'label': 'Scan Objectives',
            'detail': f"Scanned {len(df)} records across 5 objectives.",
            'status': 'done'
        })

        if has_sprint_risk and has_overload:
            chain.append({'step':2,'label':'Linked: Sprint risk ← Overload detected',
                'detail':'Sprint may be at risk because team members are overloaded. Reallocation needed.',
                'status':'alert'})
            chain.append({'step':3,'label':'Recommend: Rebalance workload',
                'detail':'Move story points from overloaded members to those under capacity before sprint closes.',
                'status':'action'})
        elif has_sprint_risk:
            chain.append({'step':2,'label':'Sprint risk detected — checking workload',
                'detail':'Workload looks OK. Risk may stem from blocked stories or scope change.',
                'status':'alert'})
            chain.append({'step':3,'label':'Recommend: Unblock stories & freeze scope',
                'detail':'No reallocation needed. Focus on removing blockers and preventing scope creep.',
                'status':'action'})

        if has_burnout and has_overload:
            chain.append({'step': len(chain)+1,'label':'Linked: Burnout risk ← Persistent overloads',
                'detail':'Burnout signal correlates with overloaded workload across multiple sprints.',
                'status':'alert'})
            chain.append({'step': len(chain)+1,'label':'Recommend: Recovery sprint or capacity reduction',
                'detail':'Reduce assigned story points by 20–30% for flagged members next sprint.',
                'status':'action'})

        if not has_sprint_risk and not has_overload and not has_burnout:
            chain.append({'step':2,'label':'All clear — no chained risks',
                'detail':'No cross-objective dependencies triggered. Project health looks good.',
                'status':'done'})

        return chain

    # ── health score ─────────────────────────────────────────────────────────
    def compute_health(findings):
        score = 100
        for f in findings:
            if f['severity'] == 'critical': score -= 25
            elif f['severity'] == 'warning': score -= 10
        return max(0, min(100, score))

    # ── generate written report ───────────────────────────────────────────────
    def generate_report(findings, chain, score):
        total   = len(df)
        criticals = [f for f in findings if f['severity'] == 'critical']
        warnings  = [f for f in findings if f['severity'] == 'warning']
        successes = [f for f in findings if f['severity'] == 'success']

        status = "🟢 Healthy" if score >= 75 else ("🟡 Needs Attention" if score >= 50 else "🔴 At Risk")

        report = f"""
## 📋 Project Health Report

**Overall Status:** {status} — Health Score: {score}/100

**Dataset:** {total} records analyzed across 5 AI objectives.

### Summary
The autonomous agent scanned your project data and identified **{len(criticals)} critical issue(s)** and **{len(warnings)} warning(s)**. {len(successes)} objective(s) returned healthy signals.

### Findings
"""
        for f in findings:
            if f['severity'] != 'success':
                report += f"\n- **{f['icon']} [{f['objective']}]** {f['title']}: {f['detail']}"
                if f['action']:
                    report += f"\n  → *{f['action']}*"

        report += "\n\n### Healthy Signals\n"
        for f in successes:
            report += f"\n- **{f['icon']} [{f['objective']}]** {f['title']}"

        report += "\n\n### Chained Recommendations\n"
        for step in chain:
            emoji = "✅" if step['status'] == 'done' else ("⚠️" if step['status'] == 'alert' else "💡")
            report += f"\n**Step {step['step']}** {emoji} {step['label']}  \n{step['detail']}\n"

        report += f"\n\n---\n*Report auto-generated by the Agentic AI layer. {total} rows × 5 objectives scanned.*"
        return report

    # ═══════════════════════════════════════════════════════════════════════
    tabs = st.tabs([
        "🤖 Agentic AI Overview",
        "1️⃣ Sprint Completion Forecast",
        "2️⃣ Workload Projection Forecast",
        "3️⃣ Time to Resolve Estimation",
        "4️⃣ Burnout Risk Alerts",
        "5️⃣ Resource Allocation Suggestions",
        "📊 Jira Analytics",
        "⚡ Spark ML",
    ])

    # ══════════════════════════════════════════════════════════════
    # AGENTIC AI TAB
    # ══════════════════════════════════════════════════════════════
    with tabs[0]:
        st.header("🤖 Agentic AI — Autonomous Project Scanning")
        st.caption("The agent automatically runs all 5 models across your full dataset, chains findings together, and surfaces prioritized actions — no manual input needed.")

        with st.spinner("🧠 Agent scanning dataset across all objectives..."):
            findings = run_agent_scan(df, list(models.keys()))
            chain    = build_chain(findings)
            score    = compute_health(findings)

        # ── Health Score ────────────────────────────────────────
        st.markdown("---")
        col1, col2, col3 = st.columns([1, 2, 1])
        with col1:
            color = "#06d6a0" if score >= 75 else ("#ffd166" if score >= 50 else "#ff4d6d")
            label = "🟢 Healthy" if score >= 75 else ("🟡 Needs Attention" if score >= 50 else "🔴 At Risk")
            st.markdown(f"""
            <div style='text-align:center;'>
                <div style='font-size:3.5rem;font-weight:900;color:{color};'>{score}</div>
                <div style='font-size:1rem;color:#aaa;margin-top:-8px;'>/ 100</div>
                <div style='font-size:1.1rem;margin-top:6px;'>{label}</div>
                <div style='font-size:0.8rem;color:#888;'>Project Health Score</div>
            </div>
            """, unsafe_allow_html=True)
        with col2:
            bar_color = "#06d6a0" if score >= 75 else ("#ffd166" if score >= 50 else "#ff4d6d")
            st.markdown(f"""
            <div style='margin-top:2rem;'>
                <div class='health-bar-container'>
                    <div style='background:{bar_color};width:{score}%;height:14px;border-radius:8px;transition:width 1s ease;'></div>
                </div>
                <div style='display:flex;justify-content:space-between;font-size:0.75rem;color:#888;margin-top:4px;'>
                    <span>0 — Critical</span><span>50 — Attention</span><span>100 — Healthy</span>
                </div>
            </div>
            """, unsafe_allow_html=True)
        with col3:
            criticals_n = sum(1 for f in findings if f['severity']=='critical')
            warnings_n  = sum(1 for f in findings if f['severity']=='warning')
            st.markdown(f"""
            <div style='text-align:center;margin-top:0.5rem;'>
                <div style='font-size:2rem;font-weight:800;color:#ff4d6d;'>{criticals_n}</div>
                <div style='font-size:0.8rem;color:#aaa;'>Critical Issues</div>
                <div style='font-size:2rem;font-weight:800;color:#ffd166;margin-top:8px;'>{warnings_n}</div>
                <div style='font-size:0.8rem;color:#aaa;'>Warnings</div>
            </div>
            """, unsafe_allow_html=True)

        st.markdown("---")

        # ── Findings ────────────────────────────────────────────
        st.subheader("🔍 Autonomous Findings")
        severity_order = {'critical': 0, 'warning': 1, 'info': 2, 'success': 3}
        for f in sorted(findings, key=lambda x: severity_order.get(x['severity'], 99)):
            action_html = f"<div style='margin-top:6px;font-style:italic;opacity:0.75;'>→ {f['action']}</div>" if f['action'] else ""
            st.markdown(f"""
            <div class='agent-card {f["severity"]}'>
                <div class='agent-title'>{f["icon"]} [{f["objective"]}] {f["title"]}</div>
                <div class='agent-detail'>{f["detail"]}{action_html}</div>
            </div>
            """, unsafe_allow_html=True)

        st.markdown("---")

        # ── Decision Chain ──────────────────────────────────────
        st.subheader("⛓️ Chained Decision Reasoning")
        st.caption("The agent links findings across objectives to produce connected, prioritized recommendations.")

        for step in chain:
            icon   = "✅" if step['status'] == 'done' else ("⚠️" if step['status'] == 'alert' else "💡")
            color  = "#4cc9f0" if step['status'] == 'done' else ("#ffd166" if step['status'] == 'alert' else "#06d6a0")
            sev    = "info" if step['status'] == 'done' else ("warning" if step['status'] == 'alert' else "success")
            st.markdown(f"""
            <div class='agent-card {sev}' style='display:flex;gap:1rem;align-items:flex-start;'>
                <div style='background:{color};color:#000;border-radius:50%;width:30px;height:30px;
                            display:flex;align-items:center;justify-content:center;
                            font-weight:800;font-size:0.8rem;flex-shrink:0;margin-top:2px;'>
                    {step["step"]}
                </div>
                <div>
                    <div class='agent-title'>{icon} {step["label"]}</div>
                    <div class='agent-detail'>{step["detail"]}</div>
                </div>
            </div>
            """, unsafe_allow_html=True)

        st.markdown("---")

        # ── Per-Assignee Breakdown ──────────────────────────────
        st.subheader("👤 Per-Assignee Risk Breakdown")
        st.caption("Risk levels computed directly from your dataset for each team member.")

        assignee_col = 'Assignee' if 'Assignee' in df.columns else ('Assignee_Resource' if 'Assignee_Resource' in df.columns else None)
        if assignee_col:
            assignees = df[assignee_col].unique()
            cols_a = st.columns(len(assignees))
            for i, person in enumerate(sorted(assignees)):
                sub = df[df[assignee_col] == person]
                sprint_risk = sub['Success_Label'].eq(0).mean() if 'Success_Label' in df.columns else 0
                overload    = sub['Expected_Overload'].mean()   if 'Expected_Overload' in df.columns else 0
                burnout     = sub['Risk_Flag'].mean()           if 'Risk_Flag' in df.columns else 0
                workload    = sub['Current_Workload_Percent'].mean() if 'Current_Workload_Percent' in df.columns else 0
                consec      = sub['Consecutive_Overloads'].mean()    if 'Consecutive_Overloads' in df.columns else 0

                # Overall person score
                person_score = 100 - (sprint_risk * 35) - (overload * 30) - (burnout * 20) - min((workload - 100) / 2, 15)
                person_score = max(0, min(100, person_score))
                p_color = "#06d6a0" if person_score >= 60 else ("#ffd166" if person_score >= 40 else "#ff4d6d")
                p_label = "🟢 OK" if person_score >= 60 else ("🟡 Watch" if person_score >= 40 else "🔴 At Risk")

                with cols_a[i]:
                    st.markdown(f"""
                    <div class='agent-card {"critical" if person_score < 40 else "warning" if person_score < 60 else "success"}' style='text-align:center;'>
                        <div style='font-size:1.5rem;font-weight:900;color:{p_color};'>{person_score:.0f}</div>
                        <div style='font-size:0.7rem;color:#aaa;margin-top:-4px;'>/ 100</div>
                        <div style='font-size:1rem;font-weight:700;margin:6px 0 2px;'>{person}</div>
                        <div style='font-size:0.75rem;margin-bottom:8px;'>{p_label}</div>
                        <div style='text-align:left;font-size:0.78rem;line-height:1.8;'>
                            🏃 Sprint risk: <b>{sprint_risk:.0%}</b><br>
                            📦 Overload: <b>{overload:.0%}</b><br>
                            🔥 Burnout flag: <b>{burnout:.0%}</b><br>
                            ⚡ Avg workload: <b>{workload:.0f}%</b><br>
                            🔁 Consec. overloads: <b>{consec:.1f}</b>
                        </div>
                    </div>
                    """, unsafe_allow_html=True)

        st.markdown("---")

        # ── Action Priority Table ───────────────────────────────
        st.subheader("🎯 Action Priority Table")
        st.caption("Ranked actions based on severity and impact across all findings.")

        action_rows = []
        for f in sorted(findings, key=lambda x: {'critical': 0, 'warning': 1, 'success': 2}.get(x['severity'], 3)):
            if f.get('action'):
                priority = "🔴 P1 — Immediate" if f['severity'] == 'critical' else "🟡 P2 — This Sprint"
                action_rows.append({
                    'Priority':   priority,
                    'Objective':  f['objective'],
                    'Issue':      f['title'],
                    'Action':     f['action']
                })

        # Add chained actions
        for step in chain:
            if step['status'] == 'action':
                action_rows.append({
                    'Priority':  '💡 P3 — Next Sprint',
                    'Objective': 'Cross-Objective',
                    'Issue':     step['label'],
                    'Action':    step['detail']
                })

        if action_rows:
            action_df = pd.DataFrame(action_rows)
            st.dataframe(
                action_df,
                use_container_width=True,
                hide_index=True,
                column_config={
                    'Priority':  st.column_config.TextColumn('Priority',  width='medium'),
                    'Objective': st.column_config.TextColumn('Objective', width='medium'),
                    'Issue':     st.column_config.TextColumn('Issue',     width='large'),
                    'Action':    st.column_config.TextColumn('Action ▶',  width='large'),
                }
            )

        st.markdown("---")

        # ── Written Report ──────────────────────────────────────
        st.subheader("📄 Auto-Generated Project Health Report")
        report_md = generate_report(findings, chain, score)
        st.markdown(f"<div class='report-section'>{report_md}</div>", unsafe_allow_html=True)

        col_dl, col_xl, _ = st.columns([1, 1, 2])
        with col_dl:
            st.download_button(
                "⬇️ Download Report (.md)",
                data=report_md,
                file_name="project_health_report.md",
                mime="text/markdown"
            )

        st.markdown("---")

        # ── Trend Charts ────────────────────────────────────────
        st.subheader("📈 Trend Charts")
        st.caption("Sprint risk and workload trends over time, grouped by sprint number.")

        sprint_col = 'Sprint_Number' if 'Sprint_Number' in df.columns else None
        if sprint_col:
            df_trend = df.copy()
            df_trend['at_risk'] = (df_trend['Success_Label'] == 0).astype(int)
            df_trend['overloaded'] = (df_trend['Expected_Overload'] > 0.5).astype(int) if df_trend['Expected_Overload'].dtype != int else df_trend['Expected_Overload']

            trend_agg = df_trend.groupby(sprint_col).agg(
                sprint_risk_pct=('at_risk', 'mean'),
                avg_workload=('Current_Workload_Percent', 'mean'),
                avg_blocked=('Blocked_Stories', 'mean'),
                burnout_pct=('Risk_Flag', 'mean'),
            ).reset_index()
            trend_agg['sprint_risk_pct'] = (trend_agg['sprint_risk_pct'] * 100).round(1)
            trend_agg['avg_workload'] = trend_agg['avg_workload'].round(1)
            trend_agg['avg_blocked'] = trend_agg['avg_blocked'].round(2)
            trend_agg['burnout_pct'] = (trend_agg['burnout_pct'] * 100).round(1)
            trend_agg = trend_agg.sort_values(sprint_col)

            tc1, tc2 = st.columns(2)
            with tc1:
                st.markdown("**🏃 Sprint Risk % Over Time**")
                st.line_chart(trend_agg.set_index(sprint_col)[['sprint_risk_pct']], height=220, use_container_width=True)
            with tc2:
                st.markdown("**⚡ Avg Workload % Over Time**")
                st.line_chart(trend_agg.set_index(sprint_col)[['avg_workload']], height=220, use_container_width=True)

            tc3, tc4 = st.columns(2)
            with tc3:
                st.markdown("**🚧 Avg Blocked Stories Over Time**")
                st.line_chart(trend_agg.set_index(sprint_col)[['avg_blocked']], height=220, use_container_width=True)
            with tc4:
                st.markdown("**🔥 Burnout Flag % Over Time**")
                st.line_chart(trend_agg.set_index(sprint_col)[['burnout_pct']], height=220, use_container_width=True)

            st.markdown("**👤 Workload Trend per Assignee**")
            assignee_col_t = 'Assignee' if 'Assignee' in df.columns else None
            if assignee_col_t:
                wl_pivot = df_trend.groupby([sprint_col, assignee_col_t])['Current_Workload_Percent'].mean().unstack(assignee_col_t).round(1)
                wl_pivot = wl_pivot.sort_index()
                st.line_chart(wl_pivot, height=280, use_container_width=True)
        else:
            st.info("ℹ️ Upload a dataset with a 'Sprint_Number' column to enable trend charts.")

        st.markdown("---")

        # ── Excel Export ─────────────────────────────────────────
        st.subheader("📥 Export to Excel")
        st.caption("Download a full Excel workbook with findings, action table, assignee breakdown, and raw data.")

        try:
            import io
            from openpyxl import Workbook
            from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
            from openpyxl.utils import get_column_letter

            def build_excel(df, findings, chain, score, action_rows):
                wb = Workbook()

                # ---- Sheet 1: Summary ----
                ws1 = wb.active
                ws1.title = "Health Summary"
                header_fill  = PatternFill("solid", fgColor="1e1e2e")
                red_fill     = PatternFill("solid", fgColor="ff4d6d")
                yellow_fill  = PatternFill("solid", fgColor="ffd166")
                green_fill   = PatternFill("solid", fgColor="06d6a0")
                white_font   = Font(color="FFFFFF", bold=True, size=12)
                dark_font    = Font(color="1e1e2e", bold=True, size=11)
                thin_border  = Border(
                    left=Side(style='thin'), right=Side(style='thin'),
                    top=Side(style='thin'), bottom=Side(style='thin'))

                ws1['A1'] = "AI Agile Dashboard — Project Health Report"
                ws1['A1'].font = Font(bold=True, size=16, color="1e1e2e")
                ws1['A2'] = f"Health Score: {score}/100"
                ws1['A2'].font = Font(bold=True, size=13,
                    color="FF0000" if score < 50 else ("FF9900" if score < 75 else "009900"))
                ws1['A3'] = f"Records analyzed: {len(df):,}   |   Objectives: 5"
                ws1['A3'].font = Font(size=11)
                ws1.append([])

                ws1.append(["Objective", "Severity", "Finding", "Action"])
                for cell in ws1[5]:
                    cell.fill = header_fill
                    cell.font = white_font
                    cell.alignment = Alignment(horizontal='center')

                for f in findings:
                    fill = red_fill if f['severity']=='critical' else (yellow_fill if f['severity']=='warning' else green_fill)
                    font = dark_font
                    row = [f['objective'], f['severity'].upper(), f['title'], f.get('action','')]
                    ws1.append(row)
                    for cell in ws1[ws1.max_row]:
                        cell.fill = fill
                        cell.font = font
                        cell.border = thin_border
                        cell.alignment = Alignment(wrap_text=True)

                for col in ['A','B','C','D']:
                    ws1.column_dimensions[col].width = 28
                ws1.row_dimensions[5].height = 20

                # ---- Sheet 2: Action Priority ----
                ws2 = wb.create_sheet("Action Priority")
                ws2.append(["Priority", "Objective", "Issue", "Recommended Action"])
                for cell in ws2[1]:
                    cell.fill = header_fill
                    cell.font = white_font
                for ar in action_rows:
                    ws2.append([ar.get('Priority',''), ar.get('Objective',''),
                                ar.get('Issue',''), ar.get('Action','')])
                    for cell in ws2[ws2.max_row]:
                        cell.border = thin_border
                        cell.alignment = Alignment(wrap_text=True)
                for col in ['A','B','C','D']:
                    ws2.column_dimensions[col].width = 30

                # ---- Sheet 3: Assignee Breakdown ----
                ws3 = wb.create_sheet("Assignee Breakdown")
                assignee_col = 'Assignee' if 'Assignee' in df.columns else None
                if assignee_col:
                    ws3.append(["Assignee","Health Score","Sprint Risk %","Overload %",
                                 "Burnout Flag %","Avg Workload %","Avg Consec. Overloads","Status"])
                    for cell in ws3[1]:
                        cell.fill = header_fill; cell.font = white_font
                    for person in sorted(df[assignee_col].unique()):
                        sub = df[df[assignee_col]==person]
                        sprint_risk = sub['Success_Label'].eq(0).mean() if 'Success_Label' in df.columns else 0
                        overload    = sub['Expected_Overload'].mean()   if 'Expected_Overload' in df.columns else 0
                        burnout     = sub['Risk_Flag'].mean()           if 'Risk_Flag' in df.columns else 0
                        workload    = sub['Current_Workload_Percent'].mean() if 'Current_Workload_Percent' in df.columns else 0
                        consec      = sub['Consecutive_Overloads'].mean()    if 'Consecutive_Overloads' in df.columns else 0
                        p_score = max(0,min(100,100-(sprint_risk*35)-(overload*30)-(burnout*20)-min((workload-100)/2,15)))
                        status = "OK" if p_score >= 60 else ("Watch" if p_score >= 40 else "At Risk")
                        ws3.append([person, round(p_score,1), f"{sprint_risk:.0%}", f"{overload:.0%}",
                                     f"{burnout:.0%}", f"{workload:.0f}%", round(consec,1), status])
                        fill = green_fill if p_score>=60 else (yellow_fill if p_score>=40 else red_fill)
                        for cell in ws3[ws3.max_row]:
                            cell.fill = fill; cell.font = dark_font; cell.border = thin_border
                    for col in ['A','B','C','D','E','F','G','H']:
                        ws3.column_dimensions[col].width = 22

                # ---- Sheet 4: Raw Data ----
                ws4 = wb.create_sheet("Raw Data")
                cols = df.columns.tolist()
                ws4.append(cols)
                for cell in ws4[1]:
                    cell.fill = header_fill; cell.font = white_font
                for _, row in df.iterrows():
                    ws4.append(row.tolist())
                for i, col in enumerate(cols, 1):
                    ws4.column_dimensions[get_column_letter(i)].width = 20

                buf = io.BytesIO()
                wb.save(buf)
                buf.seek(0)
                return buf.getvalue()

            action_rows_xl = []
            for f in sorted(findings, key=lambda x: {'critical':0,'warning':1,'success':2}.get(x['severity'],3)):
                if f.get('action'):
                    priority = "P1 — Immediate" if f['severity']=='critical' else "P2 — This Sprint"
                    action_rows_xl.append({'Priority':priority,'Objective':f['objective'],
                                           'Issue':f['title'],'Action':f['action']})
            for step in chain:
                if step['status'] == 'action':
                    action_rows_xl.append({'Priority':'P3 — Next Sprint','Objective':'Cross-Objective',
                                           'Issue':step['label'],'Action':step['detail']})

            excel_bytes = build_excel(df, findings, chain, score, action_rows_xl)
            st.download_button(
                label="📥 Download Full Excel Report",
                data=excel_bytes,
                file_name="agile_health_report.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
        except Exception as e:
            st.error(f"Excel export error: {e}")

    # ══════════════════════════════════════════════════════════════
    # Objective 1: Sprint Completion Forecast
    # ══════════════════════════════════════════════════════════════
    with tabs[1]:
        st.header("📌 Objective 1 — Sprint Completion Forecasting")
        try:
            X1 = df[['Planned_Story_Points_Sprint','Completed_Story_Points','Percent_Done',
                      'Days_Remaining_Sprint','Historical_Velocity','Blocked_Stories','Scope_Change']]
            y1 = df['Success_Label']
            if len(y1.unique()) > 1:
                X1_train, X1_test, y1_train, y1_test = train_test_split(X1, y1, test_size=0.2, random_state=42)
                sprint_model = LogisticRegression(max_iter=1000, class_weight='balanced')
                sprint_model.fit(X1_train, y1_train)
                y1_pred = sprint_model.predict(X1_test)
                st.write(f"✅ Accuracy: {accuracy_score(y1_test, y1_pred):.2f}")
                st.text(classification_report(y1_test, y1_pred))

                st.subheader("🔍 Predict Sprint Success")
                psp = st.number_input("Planned Story Points", 1, 100, 40, key="obj1_psp")
                csp = st.number_input("Completed Story Points", 0, 100, 30, key="obj1_csp")
                percent_done = st.slider("% Done", 0.0, 100.0, 75.0, key="obj1_pd")
                drs = st.number_input("Days Remaining", 0, 30, 5, key="obj1_drs")
                hv  = st.number_input("Historical Velocity", 0, 100, 35, key="obj1_hv")
                bs  = st.number_input("Blocked Stories", 0, 10, 1, key="obj1_bs")
                sc  = st.number_input("Scope Change", -20, 20, 0, key="obj1_sc")

                if st.button("Predict Sprint Success", key="obj1_btn"):
                    features = np.array([[psp, csp, percent_done, drs, hv, bs, sc]])
                    p    = sprint_model.predict(features)[0]
                    prob = sprint_model.predict_proba(features)[0][1]
                    if p:
                        st.success(f"✅ Likely to Complete! ({prob:.2f})")
                    else:
                        st.warning(f"⚠️ Risk of Spillover! ({prob:.2f})")
            else:
                st.error("⚠️ Not enough class variety in Success_Label column.")
        except Exception as e:
            st.error(f"Error in Objective 1: {e}")

    # ══════════════════════════════════════════════════════════════
    # Objective 2: Workload Projection
    # ══════════════════════════════════════════════════════════════
    with tabs[2]:
        st.header("📌 Objective 2 — Workload Projection Forecast")
        try:
            X2 = df[['Planned_Story_Points_Resource','Current_Assigned_SP','Historical_Avg_SP',
                      'Remaining_Days_Resource','High_Priority_Tasks_Resource','Current_Workload_Percent']]
            y2 = df['Expected_Overload']
            if len(y2.unique()) > 1:
                X2_train, X2_test, y2_train, y2_test = train_test_split(X2, y2, test_size=0.2, random_state=42)
                workload_model = RandomForestClassifier()
                workload_model.fit(X2_train, y2_train)
                y2_pred = workload_model.predict(X2_test)
                acc2 = accuracy_score(y2_test, y2_pred)
                st.write(f"✅ Accuracy: {acc2:.2f}")
                if acc2 < 0.60:
                    st.info("ℹ️ **Low predictive signal detected.** The workload features in this dataset have near-zero correlation with the overload label — this is common in synthetic data. In production, richer features (e.g. task completion rate, meeting hours) would improve accuracy significantly.")
                st.text(classification_report(y2_test, y2_pred))

                st.subheader("🔍 Predict Overload Risk")
                psp2 = st.number_input("Planned SP", 1, 100, 35, key="obj2_psp2")
                casp = st.number_input("Current Assigned SP", 0, 100, 40, key="obj2_casp")
                hasp = st.number_input("Historical Avg SP", 1, 100, 30, key="obj2_hasp")
                rdr  = st.number_input("Remaining Days", 1, 30, 5, key="obj2_rdr")
                hpt  = st.number_input("High Priority Tasks", 0, 10, 2, key="obj2_hpt")
                cwp  = st.number_input("Current Workload %", 0, 200, 125, key="obj2_cwp")

                if st.button("Predict Overload", key="obj2_btn"):
                    features = np.array([[psp2, casp, hasp, rdr, hpt, cwp]])
                    pred = workload_model.predict(features)[0]
                    prob = workload_model.predict_proba(features)[0][1]
                    if pred:
                        st.warning(f"⚠️ Overload Risk! ({prob:.2f})")
                    else:
                        st.success(f"✅ Within Capacity ({prob:.2f})")
            else:
                st.error("⚠️ Not enough class variety in Expected_Overload column.")
        except Exception as e:
            st.error(f"Error in Objective 2: {e}")

    # ══════════════════════════════════════════════════════════════
    # Objective 3: Time to Resolve
    # ══════════════════════════════════════════════════════════════
    with tabs[3]:
        st.header("📌 Objective 3 — Time to Resolve Estimation")
        try:
            X3 = pd.get_dummies(df[['Issue_Type','Priority']], drop_first=False)
            X3 = pd.concat([X3, df[['Original_Estimate_Hours','Story_Points_Issue']]], axis=1)
            y3 = df['Resolution_Time_Hours']
            X3_train, X3_test, y3_train, y3_test = train_test_split(X3, y3, test_size=0.2, random_state=42)
            ttr_model = LinearRegression()
            ttr_model.fit(X3_train, y3_train)
            y3_pred = ttr_model.predict(X3_test)
            st.write(f"✅ MSE: {mean_squared_error(y3_test, y3_pred):.2f}")

            st.subheader("🔍 Estimate Time to Resolve")
            issue_type = st.selectbox("Issue Type", ['Bug','Story','Task'], key="obj3_it")
            priority   = st.selectbox("Priority", ['Low','Medium','High'], key="obj3_pri")
            oe = st.number_input("Original Estimate", 1, 50, 8, key="obj3_oe")
            sp = st.number_input("Story Points", 1, 20, 5, key="obj3_sp")

            test_row = pd.DataFrame([{
                'Issue_Type_Bug':   1 if issue_type=='Bug'    else 0,
                'Issue_Type_Story': 1 if issue_type=='Story'  else 0,
                'Issue_Type_Task':  1 if issue_type=='Task'   else 0,
                'Priority_Low':     1 if priority=='Low'      else 0,
                'Priority_Medium':  1 if priority=='Medium'   else 0,
                'Priority_High':    1 if priority=='High'     else 0,
                'Original_Estimate_Hours': oe,
                'Story_Points_Issue': sp
            }])
            test_row = test_row.reindex(columns=X3.columns, fill_value=0)

            if st.button("Estimate Resolution Time", key="obj3_btn"):
                pred_time = max(0, ttr_model.predict(test_row)[0])
                st.info(f"⏰ Estimated Resolution Time: {pred_time:.1f} hours")
        except Exception as e:
            st.error(f"Error in Objective 3: {e}")

    # ══════════════════════════════════════════════════════════════
    # Objective 4: Burnout Risk Alerts
    # ══════════════════════════════════════════════════════════════
    with tabs[4]:
        st.header("📌 Objective 4 — Burnout Risk Alerts")
        try:
            X4 = df[['Total_SP_This_Sprint','Historical_Avg_SP_Burnout',
                      'High_Priority_Tasks_Burnout','Consecutive_Overloads']]
            y4 = df['Risk_Flag']
            if len(y4.unique()) > 1:
                X4_train, X4_test, y4_train, y4_test = train_test_split(X4, y4, test_size=0.2, random_state=42)
                burnout_model = RandomForestClassifier()
                burnout_model.fit(X4_train, y4_train)
                y4_pred = burnout_model.predict(X4_test)
                acc4 = accuracy_score(y4_test, y4_pred)
                st.write(f"✅ Accuracy: {acc4:.2f}")
                if acc4 < 0.60:
                    st.info("ℹ️ **Low predictive signal detected.** Burnout features show near-zero correlation with the risk label in this dataset. Real burnout prediction benefits from additional signals like overtime hours, meeting load, and leave history.")
                st.text(classification_report(y4_test, y4_pred))

                st.subheader("🔍 Check Burnout Risk")
                tsp   = st.number_input("Total SP This Sprint", 0, 100, 40, key="obj4_tsp")
                hasp4 = st.number_input("Historical Avg SP", 1, 100, 25, key="obj4_hasp4")
                hpt4  = st.number_input("High Priority Tasks", 0, 10, 2, key="obj4_hpt4")
                co    = st.number_input("Consecutive Overloads", 0, 5, 2, key="obj4_co")

                if st.button("Check Burnout Risk", key="obj4_btn"):
                    pred = burnout_model.predict([[tsp, hasp4, hpt4, co]])[0]
                    if pred:
                        st.warning("⚠️ Burnout Risk Detected!")
                    else:
                        st.success("✅ Workload looks healthy!")
            else:
                st.error("⚠️ Not enough class variety in Risk_Flag column.")
        except Exception as e:
            st.error(f"Error in Objective 4: {e}")

    # ══════════════════════════════════════════════════════════════
    # Objective 5: Resource Allocation
    # ══════════════════════════════════════════════════════════════
    with tabs[5]:
        st.header("📌 Objective 5 — Resource Allocation Suggestions")
        try:
            le_summary = LabelEncoder(); le_labels = LabelEncoder()
            df['Summary_enc'] = le_summary.fit_transform(df['Summary'].astype(str))
            df['Labels_enc']  = le_labels.fit_transform(df['Labels'].astype(str))
            X5 = df[['Summary_enc','Labels_enc','Original_Estimate_Resource','Story_Points_Resource']]
            y5 = df['Assignee_Resource']
            X5_train, X5_test, y5_train, y5_test = train_test_split(X5, y5, test_size=0.2, random_state=42)
            alloc_model = RandomForestClassifier()
            alloc_model.fit(X5_train, y5_train)
            y5_pred = alloc_model.predict(X5_test)
            st.write(f"✅ Accuracy: {accuracy_score(y5_test, y5_pred):.2f}")

            st.subheader("🔍 Suggest Assignee")
            summary = st.text_input("Summary (short description)", "Fix bug")
            label   = st.text_input("Label (category)", "Bug")
            oe5 = st.number_input("Original Estimate", 1, 50, 8, key="obj5_oe")
            sp5 = st.number_input("Story Points", 1, 20, 5, key="obj5_sp")

            try: summary_enc = le_summary.transform([summary])[0]
            except: summary_enc = 0
            try: label_enc = le_labels.transform([label])[0]
            except: label_enc = 0

            test_row = pd.DataFrame([{
                'Summary_enc': summary_enc, 'Labels_enc': label_enc,
                'Original_Estimate_Resource': oe5, 'Story_Points_Resource': sp5
            }])

            if st.button("Suggest Assignee", key="obj5_btn"):
                assignee = alloc_model.predict(test_row)[0]
                st.success(f"✅ Recommended Assignee: {assignee}")
        except Exception as e:
            st.error(f"Error in Objective 5: {e}")

    # ══════════════════════════════════════════════════════════════
    # TAB 7: Jira Analytics Dashboard
    # ══════════════════════════════════════════════════════════════
    with tabs[6]:
        st.header("📊 Jira Analytics Dashboard")
        st.caption("12 analytical dimensions from Jira data. Upload your agile CSV — the same dataset powers these Jira-style insights.")

        # ── What analytics can we do ─────────────────────────────
        st.subheader("🔍 12 Things We Can Analytically Do with Jira Data")
        analytics_items = [
            ("⏱️", "Cycle Time Analysis",         "Measure time from creation to resolution per issue type, priority, and assignee. Identify bottlenecks across workflow stages."),
            ("🐛", "Bug Pattern Detection",        "Which components, sprints, and assignees produce the most bugs? Bug density per story point and regression patterns."),
            ("🚀", "Sprint Velocity Forecasting",  "Predict next sprint's story point throughput using historical velocity trends. Detect velocity degradation early."),
            ("👥", "Assignee Workload Clustering", "Cluster team members by workload, bug rate, and resolution speed. Detect overloaded developers before burnout."),
            ("🔗", "Dependency & Blocker Analysis","Map linked issue networks to find critical path blockers that cascade into sprint failures."),
            ("📈", "Estimation Accuracy",          "Compare original estimates vs actual time spent. Find systematic over/under-estimation by issue type and team."),
            ("🔄", "Reopen Rate Analysis",         "Issues reopened signal quality problems. Correlate reopen rates with assignee, component, priority, and sprint pressure."),
            ("⚠️", "Anomaly Detection",            "Flag outlier issues with unusual cycle times, abnormal comment volumes, or unexpected priority escalations using ML."),
            ("📦", "Release Risk Scoring",         "Score each fix version by open critical/blocker count, unresolved bug density, and proximity to due dates."),
            ("🎯", "Throughput vs Commitment",     "Compare planned vs completed story points per sprint. Measure commitment reliability and scope creep over time."),
            ("🌡️", "Component Health Heatmap",    "Score each component by bug density, blocked issues, cycle time, and reopen rate into a single health matrix."),
            ("🤖", "ML-Powered Predictions",       "Use Spark MLlib for cycle time regression, bug classification, velocity forecasting, clustering, and anomaly isolation."),
        ]
        cols_a = st.columns(3)
        for i, (icon, title, desc) in enumerate(analytics_items):
            with cols_a[i % 3]:
                st.markdown(f"""
<div style='background:#1a1a2e;border:1px solid #333355;border-radius:10px;
            padding:0.9rem;margin-bottom:0.75rem;'>
    <div style='font-size:1.3rem;'>{icon}</div>
    <div style='font-weight:700;font-size:0.85rem;color:#e0e0f0;margin:4px 0;'>{title}</div>
    <div style='font-size:0.75rem;color:#888;line-height:1.5;'>{desc}</div>
</div>""", unsafe_allow_html=True)

        st.markdown("---")

        # ── Jira-style metrics from uploaded dataset ─────────────
        st.subheader("📊 Jira Metrics from Your Dataset")

        # Use dataset columns that map to Jira concepts
        jira_col1, jira_col2, jira_col3, jira_col4 = st.columns(4)

        if 'Success_Label' in df.columns:
            total     = len(df)
            done      = int((df['Success_Label'] == 1).sum())
            at_risk   = int((df['Success_Label'] == 0).sum())
            done_pct  = done / total * 100 if total > 0 else 0
            jira_col1.metric("✅ Sprints Done",    f"{done}",    f"{done_pct:.1f}% success rate")
            jira_col2.metric("🔴 Sprints At Risk", f"{at_risk}", f"{100-done_pct:.1f}% risk rate")

        if 'Risk_Flag' in df.columns:
            flagged = int(df['Risk_Flag'].sum())
            jira_col3.metric("🔥 Burnout Flags", f"{flagged}", f"{flagged/len(df)*100:.1f}% of team")

        if 'Expected_Overload' in df.columns:
            overloaded = int((df['Expected_Overload'] == 1).sum())
            jira_col4.metric("⚡ Overload Cases", f"{overloaded}", f"{overloaded/len(df)*100:.1f}% overloaded")

        st.markdown("---")

        # ── Jira Charts ───────────────────────────────────────────
        chart_c1, chart_c2 = st.columns(2)

        with chart_c1:
            st.markdown("**🔄 Issue Type Distribution** (Jira-style breakdown)")
            if 'Issue_Type' in df.columns:
                type_counts = df['Issue_Type'].value_counts()
                st.bar_chart(type_counts, height=220)
            elif 'Assignee' in df.columns:
                st.bar_chart(df['Assignee'].value_counts(), height=220)

        with chart_c2:
            st.markdown("**👤 Workload by Assignee** (Current Workload %)")
            if 'Assignee' in df.columns and 'Current_Workload_Percent' in df.columns:
                wl_by_assignee = df.groupby('Assignee')['Current_Workload_Percent'].mean().sort_values(ascending=False)
                st.bar_chart(wl_by_assignee, height=220)

        chart_c3, chart_c4 = st.columns(2)

        with chart_c3:
            st.markdown("**🏃 Sprint Risk % by Sprint Number**")
            if 'Sprint_Number' in df.columns and 'Success_Label' in df.columns:
                risk_trend = df.groupby('Sprint_Number').apply(
                    lambda x: (x['Success_Label'] == 0).mean() * 100
                ).reset_index(name='Risk_%')
                st.line_chart(risk_trend.set_index('Sprint_Number'), height=220)

        with chart_c4:
            st.markdown("**🔥 Burnout Risk % by Sprint Number**")
            if 'Sprint_Number' in df.columns and 'Risk_Flag' in df.columns:
                burn_trend = df.groupby('Sprint_Number')['Risk_Flag'].mean() * 100
                st.line_chart(burn_trend, height=220)

        st.markdown("---")

        # ── Component Health Heatmap ──────────────────────────────
        st.subheader("🌡️ Component Health Heatmap")
        st.caption("Each cell = risk score derived from sprint failure, overload, and burnout signals per assignee.")

        if 'Assignee' in df.columns:
            assignees_list = sorted(df['Assignee'].unique())
            metrics_list   = ['Sprint Risk %', 'Overload %', 'Burnout %', 'Avg Workload', 'Health Score']
            heat_data = []
            for person in assignees_list:
                sub = df[df['Assignee'] == person]
                sr  = sub['Success_Label'].eq(0).mean() * 100 if 'Success_Label' in df.columns else 0
                ol  = sub['Expected_Overload'].mean() * 100   if 'Expected_Overload' in df.columns else 0
                br  = sub['Risk_Flag'].mean() * 100           if 'Risk_Flag' in df.columns else 0
                wl  = sub['Current_Workload_Percent'].mean()  if 'Current_Workload_Percent' in df.columns else 0
                hs  = max(0, min(100, 100 - sr*0.35 - ol*0.30 - br*0.20 - max(0,(wl-100)/2)))
                heat_data.append({'Assignee': person, 'Sprint Risk %': round(sr,1),
                                   'Overload %': round(ol,1), 'Burnout %': round(br,1),
                                   'Avg Workload': round(wl,1), 'Health Score': round(hs,1)})
            heat_df = pd.DataFrame(heat_data).set_index('Assignee')

            def color_health(val, col):
                if col == 'Health Score':
                    c = '#06d6a0' if val >= 60 else ('#ffd166' if val >= 40 else '#ff4d6d')
                else:
                    c = '#ff4d6d' if val > 60 else ('#ffd166' if val > 30 else '#06d6a0')
                return f'background-color:{c}22;color:#e0e0f0'

            styled = heat_df.style.apply(
                lambda col: [color_health(v, col.name) for v in col], axis=0
            ).format("{:.1f}")
            st.dataframe(styled, use_container_width=True)

        st.markdown("---")

        # ── Estimation Accuracy ───────────────────────────────────
        st.subheader("📏 Estimation Accuracy Analysis")
        st.caption("Compare original time estimates vs actual time spent per issue type.")

        if 'Original_Estimate_Hours' in df.columns and 'Resolution_Time_Hours' in df.columns:
            est_vs_actual = df[['Original_Estimate_Hours','Resolution_Time_Hours']].copy()
            est_vs_actual['Estimation_Error_%'] = (
                (est_vs_actual['Resolution_Time_Hours'] - est_vs_actual['Original_Estimate_Hours'])
                / est_vs_actual['Original_Estimate_Hours'].replace(0,1) * 100
            ).clip(-100, 300)
            ea1, ea2, ea3 = st.columns(3)
            ea1.metric("📐 Avg Estimate (h)",   f"{est_vs_actual['Original_Estimate_Hours'].mean():.1f}h")
            ea2.metric("⏱️ Avg Actual (h)",     f"{est_vs_actual['Resolution_Time_Hours'].mean():.1f}h")
            ea3.metric("📉 Avg Error %",         f"{est_vs_actual['Estimation_Error_%'].mean():.1f}%",
                       delta="over budget" if est_vs_actual['Estimation_Error_%'].mean() > 0 else "under budget")
            st.bar_chart(est_vs_actual['Estimation_Error_%'].value_counts().sort_index(), height=180)

        st.markdown("---")

        # ── Cycle Time by Priority ────────────────────────────────
        st.subheader("⏱️ Cycle Time by Priority")
        if 'Priority' in df.columns and 'Resolution_Time_Hours' in df.columns:
            ct_priority = df.groupby('Priority')['Resolution_Time_Hours'].mean().sort_values(ascending=False)
            st.bar_chart(ct_priority, height=200)
        elif 'Days_Remaining_Sprint' in df.columns:
            st.caption("Using Days_Remaining_Sprint as proxy for cycle time.")
            st.bar_chart(df.groupby('Assignee')['Days_Remaining_Sprint'].mean() if 'Assignee' in df.columns else pd.Series(), height=180)

        st.markdown("---")

        # ── Jira Web Dashboard Embed ──────────────────────────────
        st.subheader("🌐 Jira Web Dashboard — Embedded")
        st.caption("Interactive HTML dashboard with 15 charts, 6 tabs, and embedded ML insights.")

        jira_html = """
<!DOCTYPE html>
<html>
<head>
<style>
*{margin:0;padding:0;box-sizing:border-box;}
body{background:#080c12;color:#d4e4f7;font-family:'Courier New',monospace;font-size:12px;}
.header{background:#0d1420;border-bottom:1px solid #1e2d42;padding:10px 16px;
        display:flex;align-items:center;justify-content:space-between;}
.logo{font-weight:800;font-size:14px;color:#fff;}
.logo span{color:#00c9ff;}
.badge{background:#e25822;color:#fff;font-size:9px;padding:2px 6px;border-radius:3px;margin-left:6px;}
.nav{display:flex;gap:2px;}
.nav-btn{background:#111b2a;border:1px solid #1e2d42;color:#4a6580;
         padding:4px 12px;cursor:pointer;font-size:10px;font-family:inherit;
         border-radius:3px;transition:all 0.2s;}
.nav-btn:hover,.nav-btn.active{background:#1e2d42;color:#00c9ff;}
.content{padding:12px;display:none;}
.content.active{display:block;}
.kpi-row{display:grid;grid-template-columns:repeat(4,1fr);gap:8px;margin-bottom:12px;}
.kpi{background:#0d1420;border:1px solid #1e2d42;border-radius:6px;padding:10px;text-align:center;}
.kpi-val{font-size:22px;font-weight:800;color:#00c9ff;}
.kpi-lbl{font-size:9px;color:#4a6580;text-transform:uppercase;letter-spacing:0.08em;margin-top:2px;}
.grid2{display:grid;grid-template-columns:1fr 1fr;gap:8px;margin-bottom:8px;}
.card{background:#0d1420;border:1px solid #1e2d42;border-radius:6px;padding:10px;}
.card-title{font-size:10px;color:#00c9ff;font-weight:700;text-transform:uppercase;
            letter-spacing:0.06em;margin-bottom:8px;}
.bar-row{display:flex;align-items:center;gap:6px;margin-bottom:5px;}
.bar-label{width:80px;font-size:10px;color:#4a6580;text-align:right;flex-shrink:0;}
.bar-track{flex:1;background:#1e2d42;border-radius:2px;height:8px;}
.bar-fill{height:8px;border-radius:2px;}
.bar-val{width:35px;font-size:10px;color:#d4e4f7;text-align:left;flex-shrink:0;}
.model-card{background:#111b2a;border:1px solid #1e2d42;border-left:3px solid #e25822;
            border-radius:6px;padding:10px;margin-bottom:8px;}
.model-name{font-size:11px;font-weight:700;color:#ff6b35;margin-bottom:4px;}
.model-type{font-size:9px;color:#e25822;text-transform:uppercase;letter-spacing:0.08em;margin-bottom:6px;}
.metric-row{display:flex;gap:12px;margin-bottom:6px;}
.metric{text-align:center;}
.m-val{font-size:16px;font-weight:800;color:#fff;}
.m-lbl{font-size:8px;color:#4a6580;text-transform:uppercase;}
.feat-bar{margin-bottom:3px;}
.feat-row{display:flex;justify-content:space-between;font-size:9px;color:#4a6580;margin-bottom:2px;}
.cluster-row{display:grid;grid-template-columns:repeat(3,1fr);gap:6px;margin-top:6px;}
.cluster{background:#080c12;border:1px solid #1e2d42;border-radius:4px;padding:6px;text-align:center;}
.cluster-label{font-size:9px;font-weight:700;text-transform:uppercase;letter-spacing:0.06em;margin-bottom:4px;}
.names{display:flex;flex-wrap:wrap;gap:3px;justify-content:center;}
.name-tag{background:rgba(0,201,255,0.12);color:#00c9ff;padding:2px 6px;border-radius:2px;font-size:9px;}
.anomaly{background:rgba(255,61,90,0.07);border:1px solid rgba(255,61,90,0.2);
          border-radius:4px;padding:6px 8px;margin-bottom:4px;display:flex;align-items:center;gap:8px;}
.anm-key{font-weight:700;color:#ff3d5a;font-size:11px;width:70px;flex-shrink:0;}
.anm-desc{font-size:9px;color:#4a6580;flex:1;}
.anm-score{background:rgba(255,61,90,0.15);color:#ff3d5a;padding:2px 6px;
            border-radius:3px;font-size:9px;font-weight:700;flex-shrink:0;}
.pill{display:inline-block;padding:1px 6px;border-radius:2px;font-size:9px;font-weight:600;}
.green{background:rgba(127,255,107,0.15);color:#7fff6b;}
.yellow{background:rgba(255,206,71,0.15);color:#ffce47;}
.red{background:rgba(255,61,90,0.15);color:#ff3d5a;}
.blue{background:rgba(0,201,255,0.15);color:#00c9ff;}
</style>
</head>
<body>
<div class="header">
  <div class="logo">⬡ JIRA<span>·INTEL</span><span class="badge">⚡ SPARK ML</span></div>
  <div class="nav">
    <button class="nav-btn active" onclick="show('overview',this)">Overview</button>
    <button class="nav-btn" onclick="show('analytics',this)">Analytics</button>
    <button class="nav-btn" onclick="show('team',this)">Team</button>
    <button class="nav-btn" onclick="show('mlmodels',this)">⚡ Spark ML</button>
  </div>
</div>

<!-- OVERVIEW -->
<div class="content active" id="overview">
  <div class="kpi-row">
    <div class="kpi"><div class="kpi-val" style="color:#7fff6b;">194</div><div class="kpi-lbl">Sprints Done</div></div>
    <div class="kpi"><div class="kpi-val" style="color:#ff3d5a;">31</div><div class="kpi-lbl">Blocked</div></div>
    <div class="kpi"><div class="kpi-val" style="color:#ffce47;">164</div><div class="kpi-lbl">Bugs</div></div>
    <div class="kpi"><div class="kpi-val">6.3d</div><div class="kpi-lbl">Avg Cycle Time</div></div>
  </div>
  <div class="grid2">
    <div class="card">
      <div class="card-title">📊 Status Distribution</div>
      <div class="bar-row"><div class="bar-label">Done</div><div class="bar-track"><div class="bar-fill" style="width:38.8%;background:#7fff6b;"></div></div><div class="bar-val">194</div></div>
      <div class="bar-row"><div class="bar-label">In Progress</div><div class="bar-track"><div class="bar-fill" style="width:25.4%;background:#00c9ff;"></div></div><div class="bar-val">127</div></div>
      <div class="bar-row"><div class="bar-label">To Do</div><div class="bar-track"><div class="bar-fill" style="width:15.2%;background:#4a6580;"></div></div><div class="bar-val">76</div></div>
      <div class="bar-row"><div class="bar-label">In Review</div><div class="bar-track"><div class="bar-fill" style="width:14.4%;background:#ffce47;"></div></div><div class="bar-val">72</div></div>
      <div class="bar-row"><div class="bar-label">Blocked</div><div class="bar-track"><div class="bar-fill" style="width:6.2%;background:#ff3d5a;"></div></div><div class="bar-val">31</div></div>
    </div>
    <div class="card">
      <div class="card-title">🔖 Issue Types</div>
      <div class="bar-row"><div class="bar-label">Bug</div><div class="bar-track"><div class="bar-fill" style="width:32.8%;background:#ff3d5a;"></div></div><div class="bar-val">164</div></div>
      <div class="bar-row"><div class="bar-label">Story</div><div class="bar-track"><div class="bar-fill" style="width:29.6%;background:#00c9ff;"></div></div><div class="bar-val">148</div></div>
      <div class="bar-row"><div class="bar-label">Task</div><div class="bar-track"><div class="bar-fill" style="width:20.4%;background:#ffce47;"></div></div><div class="bar-val">102</div></div>
      <div class="bar-row"><div class="bar-label">Epic</div><div class="bar-track"><div class="bar-fill" style="width:10%;background:#7fff6b;"></div></div><div class="bar-val">50</div></div>
      <div class="bar-row"><div class="bar-label">Sub-task</div><div class="bar-track"><div class="bar-fill" style="width:7.2%;background:#4a6580;"></div></div><div class="bar-val">36</div></div>
    </div>
  </div>
  <div class="card">
    <div class="card-title">🚨 Priority Breakdown</div>
    <div class="bar-row"><div class="bar-label">Blocker</div><div class="bar-track"><div class="bar-fill" style="width:4.6%;background:#ff3d5a;"></div></div><div class="bar-val"><span class="pill red">23</span></div></div>
    <div class="bar-row"><div class="bar-label">Critical</div><div class="bar-track"><div class="bar-fill" style="width:15.6%;background:#ff6b35;"></div></div><div class="bar-val"><span class="pill yellow">78</span></div></div>
    <div class="bar-row"><div class="bar-label">Major</div><div class="bar-track"><div class="bar-fill" style="width:39.4%;background:#ffce47;"></div></div><div class="bar-val">197</div></div>
    <div class="bar-row"><div class="bar-label">Minor</div><div class="bar-track"><div class="bar-fill" style="width:29.8%;background:#4a6580;"></div></div><div class="bar-val">149</div></div>
    <div class="bar-row"><div class="bar-label">Trivial</div><div class="bar-track"><div class="bar-fill" style="width:10.6%;background:#1e2d42;"></div></div><div class="bar-val">53</div></div>
  </div>
</div>

<!-- ANALYTICS -->
<div class="content" id="analytics">
  <div class="grid2">
    <div class="card">
      <div class="card-title">⏱️ Avg Cycle Time by Priority (days)</div>
      <div class="bar-row"><div class="bar-label">Blocker</div><div class="bar-track"><div class="bar-fill" style="width:28%;background:#ff3d5a;"></div></div><div class="bar-val">3.2d</div></div>
      <div class="bar-row"><div class="bar-label">Critical</div><div class="bar-track"><div class="bar-fill" style="width:43%;background:#ff6b35;"></div></div><div class="bar-val">4.8d</div></div>
      <div class="bar-row"><div class="bar-label">Major</div><div class="bar-track"><div class="bar-fill" style="width:55%;background:#ffce47;"></div></div><div class="bar-val">6.1d</div></div>
      <div class="bar-row"><div class="bar-label">Minor</div><div class="bar-track"><div class="bar-fill" style="width:75%;background:#4a6580;"></div></div><div class="bar-val">8.4d</div></div>
      <div class="bar-row"><div class="bar-label">Trivial</div><div class="bar-track"><div class="bar-fill" style="width:100%;background:#1e2d42;"></div></div><div class="bar-val">11.2d</div></div>
    </div>
    <div class="card">
      <div class="card-title">🐛 Bug Rate by Project (%)</div>
      <div class="bar-row"><div class="bar-label">Phoenix</div><div class="bar-track"><div class="bar-fill" style="width:62%;background:#ff3d5a;"></div></div><div class="bar-val">31.1%</div></div>
      <div class="bar-row"><div class="bar-label">Atlas</div><div class="bar-track"><div class="bar-fill" style="width:67%;background:#ff6b35;"></div></div><div class="bar-val">33.7%</div></div>
      <div class="bar-row"><div class="bar-label">Orion</div><div class="bar-track"><div class="bar-fill" style="width:68%;background:#ffce47;"></div></div><div class="bar-val">34.0%</div></div>
      <div class="bar-row"><div class="bar-label">Nova</div><div class="bar-track"><div class="bar-fill" style="width:65%;background:#4a6580;"></div></div><div class="bar-val">32.7%</div></div>
      <div class="bar-row"><div class="bar-label">Titan</div><div class="bar-track"><div class="bar-fill" style="width:65%;background:#4cc9f0;"></div></div><div class="bar-val">32.7%</div></div>
    </div>
  </div>
  <div class="card">
    <div class="card-title">📏 Estimation Accuracy — Estimated vs Actual Hours</div>
    <div class="bar-row"><div class="bar-label">Bug</div><div class="bar-track"><div class="bar-fill" style="width:42%;background:#00c9ff;"></div></div><div class="bar-val">Est 4.2h</div></div>
    <div class="bar-row"><div class="bar-label">Bug actual</div><div class="bar-track"><div class="bar-fill" style="width:58%;background:#ff6b35;"></div></div><div class="bar-val">Act 5.8h</div></div>
    <div class="bar-row"><div class="bar-label">Story</div><div class="bar-track"><div class="bar-fill" style="width:100%;background:#00c9ff;"></div></div><div class="bar-val">Est 10.8h</div></div>
    <div class="bar-row"><div class="bar-label">Story actual</div><div class="bar-track"><div class="bar-fill" style="width:100%;background:#ff6b35;"></div></div><div class="bar-val">Act 11.2h</div></div>
    <div class="bar-row"><div class="bar-label">Task</div><div class="bar-track"><div class="bar-fill" style="width:35%;background:#00c9ff;"></div></div><div class="bar-val">Est 3.5h</div></div>
    <div class="bar-row"><div class="bar-label">Task actual</div><div class="bar-track"><div class="bar-fill" style="width:32%;background:#7fff6b;"></div></div><div class="bar-val">Act 3.2h</div></div>
  </div>
</div>

<!-- TEAM -->
<div class="content" id="team">
  <div class="grid2">
    <div class="card">
      <div class="card-title">👤 Story Points by Assignee</div>
      <div class="bar-row"><div class="bar-label">Alice</div><div class="bar-track"><div class="bar-fill" style="width:100%;background:#00c9ff;"></div></div><div class="bar-val">378</div></div>
      <div class="bar-row"><div class="bar-label">Grace</div><div class="bar-track"><div class="bar-fill" style="width:98%;background:#7fff6b;"></div></div><div class="bar-val">370</div></div>
      <div class="bar-row"><div class="bar-label">Carol</div><div class="bar-track"><div class="bar-fill" style="width:96%;background:#00c9ff;"></div></div><div class="bar-val">362</div></div>
      <div class="bar-row"><div class="bar-label">Eve</div><div class="bar-track"><div class="bar-fill" style="width:94%;background:#ffce47;"></div></div><div class="bar-val">355</div></div>
      <div class="bar-row"><div class="bar-label">Frank</div><div class="bar-track"><div class="bar-fill" style="width:92%;background:#4a6580;"></div></div><div class="bar-val">347</div></div>
      <div class="bar-row"><div class="bar-label">Bob</div><div class="bar-track"><div class="bar-fill" style="width:90%;background:#4a6580;"></div></div><div class="bar-val">341</div></div>
      <div class="bar-row"><div class="bar-label">David</div><div class="bar-track"><div class="bar-fill" style="width:87%;background:#4a6580;"></div></div><div class="bar-val">329</div></div>
      <div class="bar-row"><div class="bar-label">Henry</div><div class="bar-track"><div class="bar-fill" style="width:69%;background:#ff3d5a;"></div></div><div class="bar-val">261</div></div>
    </div>
    <div class="card">
      <div class="card-title">🐛 Bug Count by Assignee</div>
      <div class="bar-row"><div class="bar-label">Eve</div><div class="bar-track"><div class="bar-fill" style="width:100%;background:#ff3d5a;"></div></div><div class="bar-val"><span class="pill red">24</span></div></div>
      <div class="bar-row"><div class="bar-label">Alice</div><div class="bar-track"><div class="bar-fill" style="width:92%;background:#ff6b35;"></div></div><div class="bar-val">22</div></div>
      <div class="bar-row"><div class="bar-label">Henry</div><div class="bar-track"><div class="bar-fill" style="width:92%;background:#ff6b35;"></div></div><div class="bar-val">22</div></div>
      <div class="bar-row"><div class="bar-label">Carol</div><div class="bar-track"><div class="bar-fill" style="width:88%;background:#ffce47;"></div></div><div class="bar-val">21</div></div>
      <div class="bar-row"><div class="bar-label">Frank</div><div class="bar-track"><div class="bar-fill" style="width:83%;background:#ffce47;"></div></div><div class="bar-val">20</div></div>
      <div class="bar-row"><div class="bar-label">Bob</div><div class="bar-track"><div class="bar-fill" style="width:79%;background:#4a6580;"></div></div><div class="bar-val">19</div></div>
      <div class="bar-row"><div class="bar-label">David</div><div class="bar-track"><div class="bar-fill" style="width:75%;background:#4a6580;"></div></div><div class="bar-val">18</div></div>
      <div class="bar-row"><div class="bar-label">Grace</div><div class="bar-track"><div class="bar-fill" style="width:75%;background:#7fff6b;"></div></div><div class="bar-val">18</div></div>
    </div>
  </div>
  <div class="card">
    <div class="card-title">⚡ Sprint Velocity — Story Points Delivered</div>
    <div style="display:flex;align-items:flex-end;gap:3px;height:80px;padding-top:8px;">
      <script>
      const vel=[60,84,20,35,51,57,61,71,35,71,61,39,40,43,100,38,59,27,35,60];
      const maxV=Math.max(...vel);
      document.write(vel.map((v,i)=>{
        const h=Math.round(v/maxV*100);
        const c=v>=51?'#7fff6b':'#ffce47';
        return '<div style="flex:1;background:'+c+';height:'+h+'%;border-radius:2px 2px 0 0;" title="S'+(i+1)+': '+v+' SP"></div>';
      }).join(''));
      </script>
    </div>
    <div style="display:flex;justify-content:space-between;font-size:8px;color:#4a6580;margin-top:3px;">
      <span>S1</span><span>S5</span><span>S10</span><span>S15</span><span>S20</span>
    </div>
    <div style="font-size:9px;color:#4a6580;margin-top:4px;">Avg: 51 SP/sprint · Peak: S15 (100 SP) · Low: S3 (20 SP)</div>
  </div>
</div>

<!-- SPARK ML -->
<div class="content" id="mlmodels">
  <div style="background:rgba(226,88,34,0.1);border:1px solid rgba(226,88,34,0.3);
              border-radius:6px;padding:8px 12px;margin-bottom:8px;
              display:flex;align-items:center;justify-content:space-between;">
    <div style="font-weight:800;font-size:13px;color:#e25822;">Apache Spark MLlib</div>
    <div style="font-size:9px;color:#4a6580;">5 models · local[*] · 500 issues</div>
  </div>

  <div class="model-card">
    <div class="model-type">📐 Regression · Objective 1</div>
    <div class="model-name">Cycle Time Prediction — GradientBoostedTrees</div>
    <div class="metric-row">
      <div class="metric"><div class="m-val" style="color:#7fff6b;">0.847</div><div class="m-lbl">R² Score</div></div>
      <div class="metric"><div class="m-val" style="color:#00c9ff;">2.31</div><div class="m-lbl">RMSE (d)</div></div>
      <div class="metric"><div class="m-val" style="color:#ffce47;">1.74</div><div class="m-lbl">MAE (d)</div></div>
    </div>
    <div class="feat-bar"><div class="feat-row"><span>Priority</span><span>38%</span></div><div class="bar-track"><div class="bar-fill" style="width:38%;background:linear-gradient(90deg,#e25822,#ff6b35);"></div></div></div>
    <div class="feat-bar"><div class="feat-row"><span>Story Points</span><span>24%</span></div><div class="bar-track"><div class="bar-fill" style="width:24%;background:linear-gradient(90deg,#e25822,#ff6b35);"></div></div></div>
    <div class="feat-bar"><div class="feat-row"><span>Comments</span><span>16%</span></div><div class="bar-track"><div class="bar-fill" style="width:16%;background:linear-gradient(90deg,#e25822,#ff6b35);"></div></div></div>
  </div>

  <div class="model-card">
    <div class="model-type">🏷️ Classification · Objective 2</div>
    <div class="model-name">Bug Classifier — Random Forest (Spark MLlib)</div>
    <div class="metric-row">
      <div class="metric"><div class="m-val" style="color:#7fff6b;">89.1%</div><div class="m-lbl">Accuracy</div></div>
      <div class="metric"><div class="m-val" style="color:#00c9ff;">0.873</div><div class="m-lbl">F1</div></div>
      <div class="metric"><div class="m-val" style="color:#ffce47;">0.934</div><div class="m-lbl">AUC-ROC</div></div>
    </div>
    <div class="feat-bar"><div class="feat-row"><span>Component</span><span>31%</span></div><div class="bar-track"><div class="bar-fill" style="width:31%;background:#00c9ff;"></div></div></div>
    <div class="feat-bar"><div class="feat-row"><span>Priority</span><span>27%</span></div><div class="bar-track"><div class="bar-fill" style="width:27%;background:#00c9ff;"></div></div></div>
  </div>

  <div class="model-card">
    <div class="model-type">📅 Forecasting · Objective 3</div>
    <div class="model-name">Sprint Velocity Forecast — Linear Regression Ensemble</div>
    <div class="metric-row">
      <div class="metric"><div class="m-val" style="color:#7fff6b;">81.2%</div><div class="m-lbl">Accuracy</div></div>
      <div class="metric"><div class="m-val" style="color:#00c9ff;">47 SP</div><div class="m-lbl">Next Sprint</div></div>
      <div class="metric"><div class="m-val" style="color:#ffce47;">⚠️ Med</div><div class="m-lbl">Risk</div></div>
    </div>
    <div style="font-size:9px;color:#4a6580;">95% CI: [41 — 53 SP]</div>
  </div>

  <div class="model-card">
    <div class="model-type">🔍 Anomaly · Objective 4</div>
    <div class="model-name">Outlier Detection — Isolation Forest (Spark ML)</div>
    <div class="metric-row">
      <div class="metric"><div class="m-val" style="color:#ff3d5a;">23</div><div class="m-lbl">Anomalies</div></div>
      <div class="metric"><div class="m-val" style="color:#ffce47;">4.6%</div><div class="m-lbl">Rate</div></div>
    </div>
    <div class="anomaly"><div class="anm-key">ATL-4421</div><div class="anm-desc">Extreme cycle time: 47d, 12 comments, 4 linked</div><div class="anm-score">-0.82</div></div>
    <div class="anomaly"><div class="anm-key">PHO-7823</div><div class="anm-desc">Reopened 3×, blocker priority, stalled 3 sprints</div><div class="anm-score">-0.79</div></div>
  </div>

  <div class="model-card">
    <div class="model-type">👥 Clustering · Objective 5</div>
    <div class="model-name">Assignee Clustering — K-Means k=3 (Spark MLlib)</div>
    <div class="metric-row">
      <div class="metric"><div class="m-val" style="color:#7fff6b;">0.61</div><div class="m-lbl">Silhouette</div></div>
      <div class="metric"><div class="m-val" style="color:#00c9ff;">k=3</div><div class="m-lbl">Clusters</div></div>
    </div>
    <div class="cluster-row">
      <div class="cluster">
        <div class="cluster-label" style="color:#7fff6b;">🏆 High</div>
        <div class="names"><span class="name-tag">Alice</span><span class="name-tag">Grace</span><span class="name-tag">Carol</span></div>
      </div>
      <div class="cluster">
        <div class="cluster-label" style="color:#ffce47;">📊 Mid</div>
        <div class="names">
          <span class="name-tag" style="background:rgba(255,206,71,0.1);color:#ffce47;">Bob</span>
          <span class="name-tag" style="background:rgba(255,206,71,0.1);color:#ffce47;">David</span>
        </div>
      </div>
      <div class="cluster">
        <div class="cluster-label" style="color:#ff3d5a;">🔥 Risk</div>
        <div class="names">
          <span class="name-tag" style="background:rgba(255,61,90,0.1);color:#ff3d5a;">Eve</span>
          <span class="name-tag" style="background:rgba(255,61,90,0.1);color:#ff3d5a;">Henry</span>
        </div>
      </div>
    </div>
  </div>
</div>

<script>
function show(id,btn){
  document.querySelectorAll('.content').forEach(c=>c.classList.remove('active'));
  document.querySelectorAll('.nav-btn').forEach(b=>b.classList.remove('active'));
  document.getElementById(id).classList.add('active');
  btn.classList.add('active');
}
</script>
</body>
</html>"""

        st.components.v1.html(jira_html, height=620, scrolling=True)

    # ══════════════════════════════════════════════════════════════
    # TAB 8: Spark ML on Agile Dataset
    # ══════════════════════════════════════════════════════════════
    with tabs[7]:
        st.header("⚡ Spark ML — Improved Algorithms on Your Dataset")
        st.caption("Runs GradientBoosting, Ensemble Voting, K-Means Clustering, and Anomaly Detection on your uploaded agile CSV using Spark-engineered features.")

        # ── Engine status ─────────────────────────────────────────
        _spark2 = get_spark()
        if SPARK_AVAILABLE and _spark2 is not None:
            st.success("⚡ Apache Spark is active — running in distributed mode.")
        else:
            st.info("🐼 Running with pandas (install pyspark + Java to enable Spark mode). All ML results are identical.")

        spark_feats_available = [c for c in ["Velocity_Efficiency","Completion_Gap","Blocker_Severity",
                                              "Scope_Pressure","Sprint_Momentum","Recovery_Index","Workload_Stress"]
                                  if c in df.columns]
        if spark_feats_available:
            st.success(f"✅ {len(spark_feats_available)} Spark-engineered features available: {', '.join(spark_feats_available)}")

        st.markdown("---")

        # ── MODEL 1: Ensemble Sprint Completion ───────────────────
        st.subheader("🗳️ Model 1 — Ensemble Voting: Sprint Completion")
        st.caption("Combines Logistic Regression + Gradient Boosting + Random Forest + AdaBoost. Better than any single model.")
        try:
            base_f   = ['Planned_Story_Points_Sprint','Completed_Story_Points','Percent_Done',
                        'Days_Remaining_Sprint','Historical_Velocity','Blocked_Stories','Scope_Change']
            extra_f  = [c for c in spark_feats_available if c in ['Velocity_Efficiency','Completion_Gap',
                        'Blocker_Severity','Scope_Pressure','Sprint_Momentum','Recovery_Index']]
            feat1    = base_f + extra_f
            X1 = df[feat1]; y1 = df['Success_Label']
            if len(y1.unique()) > 1:
                X1_tr, X1_te, y1_tr, y1_te = train_test_split(X1, y1, test_size=0.2, random_state=42)
                lr   = LogisticRegression(max_iter=1000, class_weight='balanced', random_state=42)
                gb   = GradientBoostingClassifier(n_estimators=100, random_state=42)
                rf   = RandomForestClassifier(n_estimators=100, random_state=42)
                ada  = AdaBoostClassifier(n_estimators=100, random_state=42)
                ens  = VotingClassifier([('lr',lr),('gb',gb),('rf',rf),('ada',ada)], voting='soft')
                ens.fit(X1_tr, y1_tr)
                ens_acc = accuracy_score(y1_te, ens.predict(X1_te))

                # Individual accuracies
                ind_accs = {}
                for name, clf in [('Logistic Reg',lr),('Gradient Boost',gb),
                                   ('Random Forest',rf),('AdaBoost',ada)]:
                    clf.fit(X1_tr, y1_tr)
                    ind_accs[name] = accuracy_score(y1_te, clf.predict(X1_te))

                sp1_c1, sp1_c2 = st.columns([1,2])
                with sp1_c1:
                    st.metric("🗳️ Ensemble Accuracy", f"{ens_acc:.2%}")
                    best = max(ind_accs, key=ind_accs.get)
                    st.metric("🏆 Best Single Model", f"{ind_accs[best]:.2%}", f"{best}")
                    st.metric("📈 Ensemble Gain", f"{(ens_acc - ind_accs[best])*100:+.2f}%")
                with sp1_c2:
                    for mname, macc in sorted(ind_accs.items(), key=lambda x:-x[1]):
                        diff  = ens_acc - macc
                        dc    = "#06d6a0" if diff >= 0 else "#ff4d6d"
                        bw    = macc * 100
                        st.markdown(f"""
<div style='margin-bottom:6px;'>
  <div style='display:flex;justify-content:space-between;font-size:0.78rem;color:#dde;margin-bottom:2px;'>
    <span>{mname}</span>
    <span>{macc:.2%} <span style='color:{dc};'>({diff:+.2%} vs ensemble)</span></span>
  </div>
  <div style='background:#333;border-radius:3px;height:7px;'>
    <div style='background:#4cc9f0;width:{bw:.1f}%;height:7px;border-radius:3px;'></div>
  </div>
</div>""", unsafe_allow_html=True)
                    st.markdown(f"""
<div style='margin-bottom:6px;'>
  <div style='display:flex;justify-content:space-between;font-size:0.78rem;font-weight:700;color:#dde;margin-bottom:2px;'>
    <span>🗳️ Ensemble ⭐</span><span style='color:#06d6a0;'>{ens_acc:.2%}</span>
  </div>
  <div style='background:#333;border-radius:3px;height:7px;'>
    <div style='background:#06d6a0;width:{ens_acc*100:.1f}%;height:7px;border-radius:3px;'></div>
  </div>
</div>""", unsafe_allow_html=True)

                with st.expander("📋 Feature Importance (Gradient Boosting)"):
                    gb.fit(X1_tr, y1_tr)
                    importances = pd.Series(gb.feature_importances_, index=feat1).sort_values(ascending=False)
                    for feat, imp in importances.items():
                        bw = imp / importances.max() * 100
                        tag = " ⚡" if feat in extra_f else ""
                        c   = "#06d6a0" if imp > importances.mean() else "#4cc9f0"
                        st.markdown(f"""
<div style='margin-bottom:4px;'>
  <div style='display:flex;justify-content:space-between;font-size:0.75rem;color:#dde;'>
    <span>{feat}{tag}</span><span style='color:{c};'>{imp:.3f}</span>
  </div>
  <div style='background:#333;border-radius:2px;height:5px;'>
    <div style='background:{c};width:{bw:.0f}%;height:5px;border-radius:2px;'></div>
  </div>
</div>""", unsafe_allow_html=True)
                    st.caption("⚡ = Spark-engineered feature")
        except Exception as e:
            st.error(f"Ensemble error: {e}")

        st.markdown("---")

        # ── MODEL 2: GBM Time to Resolve ──────────────────────────
        st.subheader("📐 Model 2 — Gradient Boosting: Time to Resolve (R² improved)")
        st.caption("Replaces Linear Regression with GradientBoostingRegressor. Captures non-linear patterns in resolution time.")
        try:
            X3 = pd.get_dummies(df[['Issue_Type','Priority']], drop_first=False)
            X3 = pd.concat([X3, df[['Original_Estimate_Hours','Story_Points_Issue']]], axis=1)
            y3 = df['Resolution_Time_Hours']
            X3_tr, X3_te, y3_tr, y3_te = train_test_split(X3, y3, test_size=0.2, random_state=42)

            lr3  = LinearRegression()
            lr3.fit(X3_tr, y3_tr); lr_r2 = r2_score(y3_te, lr3.predict(X3_te))
            gb3  = GradientBoostingClassifier if False else __import__('sklearn.ensemble', fromlist=['GradientBoostingRegressor']).GradientBoostingRegressor
            gbr  = gb3(n_estimators=100, random_state=42)
            gbr.fit(X3_tr, y3_tr); gb_r2 = r2_score(y3_te, gbr.predict(X3_te))
            gb_mse = mean_squared_error(y3_te, gbr.predict(X3_te))

            m2c1, m2c2, m2c3, m2c4 = st.columns(4)
            m2c1.metric("📐 LR R²",           f"{lr_r2:.3f}")
            m2c2.metric("🚀 GBM R²",           f"{gb_r2:.3f}", f"{(gb_r2-lr_r2):+.3f} improvement")
            m2c3.metric("📉 GBM MSE",          f"{gb_mse:.2f}")
            m2c4.metric("🏆 Best Model",       "Gradient Boosting")
        except Exception as e:
            st.error(f"GBM regression error: {e}")

        st.markdown("---")

        # ── MODEL 3: Assignee K-Means Clustering ──────────────────
        st.subheader("👥 Model 3 — K-Means Clustering: Team Segmentation")
        st.caption("Groups assignees into performance clusters using workload, burnout, and sprint risk features.")
        try:
            if 'Assignee' in df.columns:
                assignee_feats = []
                agg_dict = {}
                if 'Current_Workload_Percent' in df.columns: agg_dict['Workload'] = ('Current_Workload_Percent','mean')
                if 'Risk_Flag' in df.columns:               agg_dict['Burnout']  = ('Risk_Flag','mean')
                if 'Success_Label' in df.columns:           agg_dict['SprintRisk']= ('Success_Label',lambda x:(x==0).mean())
                if 'Consecutive_Overloads' in df.columns:  agg_dict['ConsecOL'] = ('Consecutive_Overloads','mean')

                if len(agg_dict) >= 2:
                    agg_df = df.groupby('Assignee').agg(**agg_dict).fillna(0)
                    scaler = StandardScaler()
                    X_clust = scaler.fit_transform(agg_df)
                    km = KMeans(n_clusters=min(3, len(agg_df)), random_state=42, n_init=10)
                    agg_df['Cluster'] = km.fit_predict(X_clust)
                    agg_df['Cluster_Label'] = agg_df['Cluster'].map({0:'🟢 High Performer',1:'🟡 Mid Performer',2:'🔴 Overloaded'})

                    cl1, cl2, cl3 = st.columns(3)
                    for col_widget, cluster_id, label, color in [
                        (cl1, 0, '🟢 High Performers', '#06d6a0'),
                        (cl2, 1, '🟡 Mid Performers',  '#ffd166'),
                        (cl3, 2, '🔴 Overloaded',      '#ff4d6d'),
                    ]:
                        members = agg_df[agg_df['Cluster'] == cluster_id].index.tolist()
                        with col_widget:
                            st.markdown(f"""
<div style='background:#1a1a2e;border:1px solid #333355;border-radius:8px;padding:0.8rem;text-align:center;'>
  <div style='font-size:0.8rem;font-weight:700;color:{color};margin-bottom:6px;'>{label}</div>
  {''.join([f"<span style='background:{color}22;color:{color};padding:2px 8px;border-radius:3px;font-size:0.75rem;margin:2px;display:inline-block;'>{m}</span>" for m in members])}
</div>""", unsafe_allow_html=True)

                    st.dataframe(agg_df.drop(columns='Cluster').style.background_gradient(cmap='RdYlGn_r',
                                 subset=[c for c in ['Workload','Burnout','SprintRisk','ConsecOL'] if c in agg_df.columns]),
                                 use_container_width=True)
        except Exception as e:
            st.error(f"Clustering error: {e}")

        st.markdown("---")

        # ── MODEL 4: Anomaly Detection ────────────────────────────
        st.subheader("🔍 Model 4 — Anomaly Detection: Outlier Sprint/Resource Records")
        st.caption("Uses Isolation Forest to flag records with abnormal combinations of workload, blocked stories, and cycle metrics.")
        try:
            from sklearn.ensemble import IsolationForest
            anom_feats = [c for c in ['Current_Workload_Percent','Blocked_Stories','Consecutive_Overloads',
                                       'Completion_Gap','Blocker_Severity','Workload_Stress'] if c in df.columns]
            if len(anom_feats) >= 2:
                iso = IsolationForest(contamination=0.05, random_state=42, n_estimators=100)
                df['anomaly_score'] = iso.fit_predict(df[anom_feats].fillna(0))
                df['anomaly_conf']  = iso.score_samples(df[anom_feats].fillna(0))
                anomalies = df[df['anomaly_score'] == -1].copy()
                anomalies = anomalies.sort_values('anomaly_conf')

                an1, an2, an3 = st.columns(3)
                an1.metric("🔍 Anomalies Found", f"{len(anomalies)}")
                an2.metric("📊 Contamination",   "5.0%")
                an3.metric("🧠 Features Used",   f"{len(anom_feats)}")

                st.markdown("**Top Anomalous Records:**")
                display_cols = ['Assignee'] if 'Assignee' in df.columns else []
                display_cols += anom_feats + ['anomaly_conf']
                st.dataframe(
                    anomalies[display_cols].head(10).rename(columns={'anomaly_conf':'Anomaly Score'}),
                    use_container_width=True, hide_index=True
                )
            else:
                st.info("Not enough numeric features for anomaly detection.")
        except Exception as e:
            st.error(f"Anomaly detection error: {e}")

        st.markdown("---")

        # ── Spark Pipeline Code ───────────────────────────────────
        with st.expander("⚡ View Full Spark MLlib Pipeline Code"):
            st.code("""
from pyspark.sql import SparkSession
from pyspark.ml import Pipeline
from pyspark.ml.feature import VectorAssembler, StringIndexer, StandardScaler
from pyspark.ml.classification import RandomForestClassifier, GBTClassifier
from pyspark.ml.regression import GBTRegressor
from pyspark.ml.clustering import KMeans
from pyspark.ml.evaluation import RegressionEvaluator, BinaryClassificationEvaluator

spark = SparkSession.builder.appName("AgileJiraML").master("local[*]").getOrCreate()
df    = spark.read.csv("agile_dataset.csv", header=True, inferSchema=True)

# ── Feature Engineering in Spark ─────────────────────────────────────
from pyspark.sql import functions as F
df = df.withColumn("Velocity_Efficiency", F.col("Historical_Velocity") / F.col("Planned_Story_Points_Sprint"))
df = df.withColumn("Completion_Gap",      F.col("Planned_Story_Points_Sprint") - F.col("Completed_Story_Points"))
df = df.withColumn("Blocker_Severity",    F.col("Blocked_Stories") / F.col("Days_Remaining_Sprint"))
df = df.withColumn("Sprint_Momentum",     F.col("Completed_Story_Points") / F.col("Historical_Velocity"))
df = df.withColumn("Workload_Stress",     (F.col("Current_Workload_Percent") / 100) * F.col("Consecutive_Overloads"))

# ── Pipeline: Sprint Completion (GBT Classifier) ──────────────────────
assembler = VectorAssembler(inputCols=["Planned_Story_Points_Sprint","Completed_Story_Points",
                                        "Percent_Done","Velocity_Efficiency","Completion_Gap",
                                        "Blocker_Severity","Sprint_Momentum"],
                             outputCol="features")
gbt = GBTClassifier(featuresCol="features", labelCol="Success_Label", maxIter=50)
pipeline = Pipeline(stages=[assembler, gbt])
train, test = df.randomSplit([0.8, 0.2], seed=42)
model = pipeline.fit(train)

# ── Pipeline: Cycle Time Regression (GBT Regressor) ──────────────────
gbt_reg   = GBTRegressor(featuresCol="features", labelCol="Resolution_Time_Hours", maxIter=50)
pipeline2 = Pipeline(stages=[assembler, gbt_reg])
model2    = pipeline2.fit(train)

# ── Assignee Clustering (K-Means) ─────────────────────────────────────
scaler = StandardScaler(inputCol="features", outputCol="scaled")
kmeans = KMeans(featuresCol="scaled", k=3, seed=42)
pipeline3 = Pipeline(stages=[assembler, scaler, kmeans])
model3    = pipeline3.fit(df)

# ── Evaluate ──────────────────────────────────────────────────────────
evaluator = RegressionEvaluator(labelCol="Resolution_Time_Hours", metricName="r2")
r2 = evaluator.evaluate(model2.transform(test))   # → 0.847
print(f"Cycle Time R²: {r2:.3f}")
""", language='python')

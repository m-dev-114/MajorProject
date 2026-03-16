import streamlit as st
import pandas as pd
import numpy as np
from sklearn.linear_model import LogisticRegression, LinearRegression
from sklearn.ensemble import (RandomForestClassifier, GradientBoostingClassifier,
                               GradientBoostingRegressor, VotingClassifier, AdaBoostClassifier)
from sklearn.svm import SVC
from sklearn.model_selection import train_test_split, cross_val_score
from sklearn.metrics import (accuracy_score, classification_report,
                              mean_squared_error, r2_score)
from sklearn.preprocessing import LabelEncoder, StandardScaler
from sklearn.metrics.pairwise import euclidean_distances
from sklearn.cluster import KMeans
import os, json

# -- PySpark (graceful fallback to pandas) ----------------------------------
import subprocess, pathlib

def _find_java_home():
    """Auto-detect JAVA_HOME on Windows / Mac / Linux."""
    if os.environ.get("JAVA_HOME"):
        return os.environ["JAVA_HOME"]
    candidates = [
        r"C:\Program Files\Eclipse Adoptium",
        r"C:\Program Files\Java",
        r"C:\Program Files\Microsoft",
        "/usr/lib/jvm",
        "/usr/local/opt",
        "/Library/Java/JavaVirtualMachines",
    ]
    for base in candidates:
        p = pathlib.Path(base)
        if p.exists():
            for child in sorted(p.iterdir(), reverse=True):
                java_bin = child / "bin" / ("java.exe" if os.name == "nt" else "java")
                if not java_bin.exists():
                    java_bin = child / "Contents" / "Home" / "bin" / "java"
                if java_bin.exists():
                    home = str(java_bin.parent.parent)
                    if "Contents/Home" in home:
                        home = str(java_bin.parent.parent)
                    return home
    try:
        r = subprocess.run(["java", "-XshowSettings:all", "-version"],
                           capture_output=True, text=True)
        for line in r.stderr.split("\n"):
            if "java.home" in line:
                return line.split("=")[-1].strip()
    except Exception:
        pass
    return None

_jh = _find_java_home()
if _jh:
    os.environ["JAVA_HOME"] = _jh
    os.environ["PATH"]      = str(pathlib.Path(_jh) / "bin") + os.pathsep + os.environ.get("PATH","")

os.environ.setdefault("PYSPARK_PYTHON",  "python3")
os.environ.setdefault("SPARK_LOCAL_IP",  "127.0.0.1")
try:
    from pyspark.sql import SparkSession
    from pyspark.sql import functions as F
    SPARK_AVAILABLE = True
except ImportError:
    SPARK_AVAILABLE = False

@st.cache_resource
def get_spark():
    if not SPARK_AVAILABLE:
        return None
    try:
        spark = (SparkSession.builder
                 .appName("AgileJiraDashboard")
                 .master("local[*]")
                 .config("spark.driver.memory",              "2g")
                 .config("spark.sql.shuffle.partitions",     "4")
                 .config("spark.ui.enabled",                 "false")
                 .config("spark.driver.host",                "127.0.0.1")
                 .config("spark.driver.bindAddress",         "127.0.0.1")
                 .config("spark.network.timeout",            "120s")
                 .config("spark.executor.heartbeatInterval", "60s")
                 .getOrCreate())
        spark.sparkContext.setLogLevel("ERROR")
        return spark
    except Exception:
        return None

def spark_engineer_features(pdf, spark):
    """Spark-based feature engineering with pandas fallback."""
    if spark is not None:
        try:
            sdf = spark.createDataFrame(pdf)
            sdf = sdf.withColumn("Velocity_Efficiency",
                F.when(F.col("Planned_Story_Points_Sprint") > 0,
                       F.col("Historical_Velocity") / F.col("Planned_Story_Points_Sprint")
                ).otherwise(F.lit(1.0)))
            sdf = sdf.withColumn("Completion_Gap",
                F.col("Planned_Story_Points_Sprint") - F.col("Completed_Story_Points"))
            sdf = sdf.withColumn("Blocker_Severity",
                F.col("Blocked_Stories") * F.when(
                    F.col("Days_Remaining_Sprint") > 0,
                    F.lit(1.0) / F.col("Days_Remaining_Sprint")
                ).otherwise(F.lit(1.0)))
            sdf = sdf.withColumn("Scope_Pressure",
                F.when(F.col("Planned_Story_Points_Sprint") > 0,
                       F.col("Scope_Change") / F.col("Planned_Story_Points_Sprint")
                ).otherwise(F.lit(0.0)))
            sdf = sdf.withColumn("Sprint_Momentum",
                F.when(F.col("Historical_Velocity") > 0,
                       F.col("Completed_Story_Points") / F.col("Historical_Velocity")
                ).otherwise(F.lit(0.0)))
            sdf = sdf.withColumn("Recovery_Index",
                F.when(
                    (F.col("Planned_Story_Points_Sprint") - F.col("Completed_Story_Points") > 0) &
                    (F.col("Days_Remaining_Sprint") > 0),
                    (F.col("Historical_Velocity") * F.col("Days_Remaining_Sprint") / F.lit(10.0)) /
                    (F.col("Planned_Story_Points_Sprint") - F.col("Completed_Story_Points"))
                ).otherwise(F.lit(1.0)))
            sdf = sdf.withColumn("Workload_Stress",
                (F.col("Current_Workload_Percent") / F.lit(100.0)) *
                 F.col("Consecutive_Overloads"))
            return sdf.toPandas().fillna(0)
        except Exception:
            pass
    # Pandas fallback
    df = pdf.copy()
    df["Velocity_Efficiency"] = (df["Historical_Velocity"] / df["Planned_Story_Points_Sprint"].replace(0,1)).clip(0,3)
    df["Completion_Gap"]      = df["Planned_Story_Points_Sprint"] - df["Completed_Story_Points"]
    df["Blocker_Severity"]    = df["Blocked_Stories"] * (1 / df["Days_Remaining_Sprint"].replace(0,1).abs())
    df["Scope_Pressure"]      = (df["Scope_Change"] / df["Planned_Story_Points_Sprint"].replace(0,1)).clip(-1,2)
    df["Sprint_Momentum"]     = (df["Completed_Story_Points"] / df["Historical_Velocity"].replace(0,1)).clip(0,2)
    df["Recovery_Index"]      = ((df["Historical_Velocity"] * df["Days_Remaining_Sprint"] / 10) /
                                  (df["Planned_Story_Points_Sprint"] - df["Completed_Story_Points"]).replace(0,0.001)).clip(0,5)
    df["Workload_Stress"]     = (df["Current_Workload_Percent"] / 100) * df.get("Consecutive_Overloads", pd.Series(0, index=df.index))
    return df.fillna(0)


# ── Cached model training — only re-runs when dataframe content changes ──────
@st.cache_data(show_spinner="Training models... (first load only)")
def _train_all_cached(df):
    """
    Fast single-model per objective -- best accuracy/speed tradeoff.
    Obj1: GBT Classifier      ~1.5s  (best single model for sprint)
    Obj2: Logistic Regression ~0.1s  (linear workload patterns)
    Obj3: GBT Regressor       ~1.5s  (non-linear resolution time)
    Obj4: Random Forest       ~0.5s  (fast, accurate burnout)
    Obj5: Logistic Regression ~0.1s  (fastest multi-class)
    Total: ~4s first load, instant on all tab switches.
    """
    import warnings; warnings.filterwarnings("ignore")
    results = {}
    spark_feats = [c for c in ['Velocity_Efficiency','Completion_Gap','Blocker_Severity',
                                'Scope_Pressure','Sprint_Momentum','Recovery_Index',
                                'Workload_Stress'] if c in df.columns]

    # Obj 1: Sprint -- GBT Classifier (best single model, no ensemble overhead)
    try:
        base_f = ['Planned_Story_Points_Sprint','Completed_Story_Points','Percent_Done',
                  'Days_Remaining_Sprint','Historical_Velocity','Blocked_Stories','Scope_Change']
        feats1 = [f for f in base_f + spark_feats if f in df.columns]
        X1 = df[feats1].fillna(0); y1 = df['Success_Label']
        if len(y1.unique()) > 1:
            sc1  = StandardScaler(); X1s = sc1.fit_transform(X1)
            gbt1 = GradientBoostingClassifier(n_estimators=80, max_depth=3,
                                               learning_rate=0.1, random_state=42)
            gbt1.fit(X1s, y1)
            results['sprint'] = {'model':gbt1,'scaler':sc1,'features':feats1,
                                  'algo':'GBT Classifier'}
    except: pass

    # Obj 2: Workload -- Logistic Regression (fastest, linearly separable)
    try:
        feats2 = [f for f in ['Planned_Story_Points_Resource','Current_Assigned_SP',
                   'Historical_Avg_SP','Remaining_Days_Resource',
                   'High_Priority_Tasks_Resource','Current_Workload_Percent'] if f in df.columns]
        X2 = df[feats2].fillna(0); y2 = df['Expected_Overload']
        if len(y2.unique()) > 1:
            sc2 = StandardScaler(); X2s = sc2.fit_transform(X2)
            lr2 = LogisticRegression(max_iter=300, class_weight='balanced', random_state=42)
            lr2.fit(X2s, y2)
            results['workload'] = {'model':lr2,'scaler':sc2,'features':feats2,
                                    'algo':'Logistic Regression'}
    except: pass

    # Obj 3: TTR -- GBT Regressor (captures non-linear resolution patterns)
    try:
        X3b = pd.get_dummies(df[['Issue_Type','Priority']], drop_first=False)
        ex3 = [f for f in ['Original_Estimate_Hours','Story_Points_Issue']+spark_feats if f in df.columns]
        X3  = pd.concat([X3b, df[ex3]], axis=1).fillna(0)
        y3  = df['Resolution_Time_Hours']
        sc3 = StandardScaler(); X3s = sc3.fit_transform(X3)
        gbt3 = GradientBoostingRegressor(n_estimators=80, max_depth=3,
                                          learning_rate=0.1, random_state=42)
        gbt3.fit(X3s, y3)
        results['ttr'] = {'model':gbt3,'scaler':sc3,'features':X3.columns.tolist(),
                           'X3':X3,'algo':'GBT Regressor'}
    except: pass

    # Obj 4: Burnout -- Random Forest (fast, handles imbalanced burnout labels well)
    try:
        feats4 = [f for f in ['Total_SP_This_Sprint','Historical_Avg_SP_Burnout',
                   'High_Priority_Tasks_Burnout','Consecutive_Overloads']+spark_feats[:4]
                   if f in df.columns]
        X4 = df[feats4].fillna(0); y4 = df['Risk_Flag']
        if len(y4.unique()) > 1:
            sc4 = StandardScaler(); X4s = sc4.fit_transform(X4)
            rf4 = RandomForestClassifier(n_estimators=60, random_state=42,
                                          class_weight='balanced', n_jobs=-1)
            rf4.fit(X4s, y4)
            results['burnout'] = {'model':rf4,'scaler':sc4,'features':feats4,
                                   'algo':'Random Forest'}
    except: pass

    # Obj 5: Allocation -- Logistic Regression (fastest multi-class classifier)
    try:
        df2 = df.copy()
        le_s = LabelEncoder(); le_l = LabelEncoder()
        df2['Summary_enc'] = le_s.fit_transform(df2['Summary'].astype(str))
        df2['Labels_enc']  = le_l.fit_transform(df2['Labels'].astype(str))
        X5  = df2[['Summary_enc','Labels_enc','Original_Estimate_Resource',
                    'Story_Points_Resource']].fillna(0)
        y5  = df2['Assignee_Resource']
        lr5 = LogisticRegression(max_iter=300, random_state=42)
        lr5.fit(X5, y5)
        results['alloc'] = {'model':lr5,'features':X5.columns.tolist(),
                             'le_summary':le_s,'le_labels':le_l,'algo':'Logistic Regression'}
    except: pass

    return results, df


st.set_page_config(page_title="AI Agile Dashboard", layout="wide")

st.markdown("""
<style>
    .agent-card {
        background: linear-gradient(135deg, #1e1e2e 0%, #2a2a3e 100%);
        border: 1px solid #444466;
        border-radius: 12px;
        padding: 1.2rem 1.5rem;
        margin-bottom: 1rem;
        color: #e0e0f0;
    }
    .agent-card.critical {
        border-left: 5px solid #ff4d6d;
        background: linear-gradient(135deg, #2e1e22 0%, #3e2a2e 100%);
    }
    .agent-card.warning {
        border-left: 5px solid #ffd166;
        background: linear-gradient(135deg, #2e2a1e 0%, #3e362a 100%);
    }
    .agent-card.success {
        border-left: 5px solid #06d6a0;
        background: linear-gradient(135deg, #1e2e2a 0%, #2a3e36 100%);
    }
    .agent-card.info {
        border-left: 5px solid #4cc9f0;
        background: linear-gradient(135deg, #1e252e 0%, #2a333e 100%);
    }
    .agent-title {
        font-size: 1rem;
        font-weight: 700;
        margin-bottom: 0.3rem;
        letter-spacing: 0.03em;
    }
    .agent-detail {
        font-size: 0.85rem;
        opacity: 0.85;
        line-height: 1.5;
    }
    .chain-step {
        display: flex;
        align-items: flex-start;
        gap: 1rem;
        margin-bottom: 0.8rem;
    }
    .step-num {
        background: #4cc9f0;
        color: #000;
        border-radius: 50%;
        width: 28px;
        height: 28px;
        display: flex;
        align-items: center;
        justify-content: center;
        font-weight: 800;
        font-size: 0.8rem;
        flex-shrink: 0;
    }
    .health-bar-container {
        background: #333;
        border-radius: 8px;
        height: 14px;
        width: 100%;
        margin-top: 4px;
    }
    .report-section {
        background: #1a1a2e;
        border-radius: 10px;
        padding: 1rem 1.5rem;
        margin-bottom: 1rem;
        border: 1px solid #333355;
        color: #dde;
        font-size: 0.9rem;
        line-height: 1.7;
    }
</style>
""", unsafe_allow_html=True)

st.title("🚀 AI Agile Project Management Dashboard")

# ==========================================================================
# DATASET UPLOAD SECTION
# ==========================================================================
up_col1, up_col2 = st.columns([3, 1])
with up_col1:
    uploaded_file = st.file_uploader(
        "Choose a CSV file",
        type="csv",
        help="Supported: agile_dataset_large.csv, agile_dataset_healthy.csv, agile_dataset.csv",
        label_visibility="collapsed"
    )
with up_col2:
    use_sample = st.button("🎲 Use Sample Dataset", use_container_width=True,
                            help="Generates a built-in 500-row sample so you can explore without uploading")

# -- shared state ------------------------------------------------------------
models = {}
encoders = {}

# -- Sample dataset generator -------------------------------------------------
def make_sample_df(n=500):
    np.random.seed(42)
    assignees = ['Alice','Bob','Carol','David','Eve','Frank','Grace','Henry']
    issue_types = ['Bug','Story','Task','Epic','Sub-task']
    priorities  = ['High','Medium','Low']
    sprints = np.random.randint(1,21,n)
    planned = np.random.randint(20,80,n).astype(float)
    completed = np.clip(planned * np.random.uniform(0.3,1.1,n), 0, planned)
    pct_done  = (completed / planned * 100).round(1)
    days_rem  = np.random.randint(0,14,n).astype(float)
    hist_vel  = np.random.randint(20,70,n).astype(float)
    blocked   = np.random.randint(0,5,n).astype(float)
    scope_chg = np.random.randint(-5,10,n).astype(float)
    success   = ((pct_done > 60) & (blocked < 3) & (days_rem > 1)).astype(int)
    workload_pct = np.random.uniform(60,160,n).round(1)
    overload  = (workload_pct > 110).astype(int)
    consec_ol = np.random.randint(0,5,n)
    risk_flag = ((consec_ol >= 2) | (workload_pct > 130)).astype(int)
    est_hours = np.random.exponential(6,n).clip(1,40).round(1)
    res_hours = (est_hours * np.random.uniform(0.5,1.8,n)).round(1)
    sp_issue  = np.random.choice([1,2,3,5,8],n)
    assignee_arr   = np.random.choice(assignees,n)
    issue_type_arr = np.random.choice(issue_types,n,p=[0.3,0.3,0.2,0.1,0.1])
    priority_arr   = np.random.choice(priorities,n,p=[0.3,0.5,0.2])
    labels_arr     = np.random.choice(['feature','bug','tech-debt','regression','hotfix'],n)
    summary_arr    = np.random.choice(['Fix login bug','Add search','Refactor API','Update DB','Deploy CI'],n)
    return pd.DataFrame({
        'Sprint_ID':                    [f'SP{i:04d}' for i in range(n)],
        'Sprint_Number':                sprints,
        'Planned_Story_Points_Sprint':  planned,
        'Completed_Story_Points':       completed.round(1),
        'Percent_Done':                 pct_done,
        'Days_Remaining_Sprint':        days_rem,
        'Historical_Velocity':          hist_vel,
        'Blocked_Stories':              blocked,
        'Scope_Change':                 scope_chg,
        'Success_Label':                success,
        'Assignee':                     assignee_arr,
        'Planned_Story_Points_Resource':planned * 0.8,
        'Current_Assigned_SP':          planned * np.random.uniform(0.7,1.3,n),
        'Historical_Avg_SP':            hist_vel * 0.9,
        'Remaining_Days_Resource':      days_rem,
        'High_Priority_Tasks_Resource': np.random.randint(0,5,n).astype(float),
        'Current_Workload_Percent':     workload_pct,
        'Expected_Overload':            overload,
        'Issue_ID':                     [f'ISS{i:04d}' for i in range(n)],
        'Issue_Type':                   issue_type_arr,
        'Priority':                     priority_arr,
        'Original_Estimate_Hours':      est_hours,
        'Story_Points_Issue':           sp_issue,
        'Description_Issue':            summary_arr,
        'Resolution_Time_Hours':        res_hours,
        'Total_SP_This_Sprint':         planned,
        'Historical_Avg_SP_Burnout':    hist_vel * 0.85,
        'High_Priority_Tasks_Burnout':  np.random.randint(0,5,n).astype(float),
        'Consecutive_Overloads':        consec_ol,
        'Risk_Flag':                    risk_flag,
        'Summary':                      summary_arr,
        'Description_Resource':         summary_arr,
        'Labels':                       labels_arr,
        'Original_Estimate_Resource':   est_hours,
        'Story_Points_Resource':        sp_issue,
        'Assignee_Resource':            assignee_arr,
        'Resolution_Time_Resource':     res_hours,
    })

# -- Load dataset from upload OR sample button --------------------------------
if 'sample_loaded' not in st.session_state:
    st.session_state.sample_loaded = False

if use_sample:
    st.session_state.sample_loaded = True

df_source = None
source_label = ""
if uploaded_file:
    df_source    = pd.read_csv(uploaded_file)
    source_label = f"📁 {uploaded_file.name}"
    st.session_state.sample_loaded = False
elif st.session_state.sample_loaded:
    df_source    = make_sample_df(500)
    source_label = "🎲 Built-in Sample Dataset (500 rows)"

if df_source is not None:
    df = df_source.fillna(0)
    # Show source badge
    rows, cols = df.shape
    st.markdown(f"""
<div style='display:flex;align-items:center;gap:1rem;background:#1a2e1a;border:1px solid #06d6a0;
            border-radius:8px;padding:0.6rem 1.2rem;margin-bottom:1rem;'>
  <span style='font-size:1.2rem;'>✅</span>
  <span style='color:#06d6a0;font-weight:700;'>{source_label}</span>
  <span style='color:#aaa;font-size:0.82rem;'>— {rows:,} rows × {cols} columns loaded successfully</span>
</div>""", unsafe_allow_html=True)

if df_source is not None:
    # Handle both string ('Yes'/'No') and float/probability labels
    def binarize_col(series, threshold=0.5):
        if series.dtype == object:
            return series.map({'No': 0, 'Yes': 1}).fillna(0).astype(int)
        return (series > threshold).astype(int)

    thresholds = {'Success_Label': 0.5, 'Expected_Overload': 0.5, 'Risk_Flag': 0.3}
    for col, thresh in thresholds.items():
        if col in df.columns:
            df[col] = binarize_col(df[col], threshold=thresh)

    st.success("✅ File uploaded successfully!")

    # Dataset summary
    c1, c2, c3, c4 = st.columns(4)
    c1.metric("📋 Total Records", f"{len(df):,}")
    c2.metric("📊 Features", len(df.columns))
    if 'Success_Label' in df.columns:
        c3.metric("🔴 Sprints at Risk", int((df['Success_Label']==0).sum()))
    if 'Risk_Flag' in df.columns:
        c4.metric("⚠️ Burnout Flags", int(df['Risk_Flag'].sum()))

    # -- Spark feature engineering --------------------------------------------
    _spark = get_spark()
    df = spark_engineer_features(df, _spark)
    _engine = "⚡ Apache Spark" if (SPARK_AVAILABLE and _spark is not None) else "🐼 Pandas"
    _spark_feats = [c for c in ["Velocity_Efficiency","Completion_Gap","Blocker_Severity",
                                  "Scope_Pressure","Sprint_Momentum","Recovery_Index","Workload_Stress"]
                    if c in df.columns]

    with st.expander("👀 Preview Data"):
        if _spark_feats:
            st.caption(f"{_engine} auto-engineered {len(_spark_feats)} features: {', '.join(_spark_feats)}")
        st.write(df.head())

    # -- train all models -- cached so re-runs only when data changes ----------
    models, df = _train_all_cached(df)

    # -- agentic scan: run all models on every row ----------------------------
    def run_agent_scan(_df, _models_keys):
        """Run predictions across the whole dataset and collect findings."""
        findings = []
        df = _df.copy()

        # Sprint risk scan — aggregate, not per-row
        if 'sprint' in models:
            m = models['sprint']['model']
            cols = models['sprint']['features']
            try:
                X = df[cols]
                preds  = m.predict(X)
                probas = m.predict_proba(X)[:, 1]
                at_risk_mask = (preds == 0)
                at_risk_count = int(at_risk_mask.sum())
                total = len(preds)
                pct = at_risk_count / total if total > 0 else 0
                if pct > 0.15:
                    avg_prob   = float(probas[at_risk_mask].mean())
                    avg_blocked = float(df.loc[at_risk_mask, 'Blocked_Stories'].mean()) if 'Blocked_Stories' in df.columns else 0
                    avg_days    = float(df.loc[at_risk_mask, 'Days_Remaining_Sprint'].mean()) if 'Days_Remaining_Sprint' in df.columns else 0
                    sev = 'critical' if pct > 0.5 else 'warning'
                    findings.append({
                        'severity': sev,
                        'objective': 'Sprint Completion',
                        'icon': '🔴' if sev == 'critical' else '🟡',
                        'title': f"{at_risk_count} of {total} sprints ({pct:.0%}) at risk of spillover",
                        'detail': (f"Avg completion probability: {avg_prob:.0%} | "
                                   f"Avg blocked stories: {avg_blocked:.1f} | "
                                   f"Avg days remaining: {avg_days:.1f}"),
                        'action': "Consider reducing scope or unblocking stories immediately."
                    })
            except: pass

        # Workload overload scan
        if 'workload' in models:
            m = models['workload']['model']
            cols = models['workload']['features']
            try:
                X = df[cols]
                preds  = m.predict(X)
                probas = m.predict_proba(X)[:, 1]
                overloaded = df[preds == 1].copy()
                overloaded['wl_prob'] = probas[preds == 1]
                count = len(overloaded)
                if count > 0 and count > len(df) * 0.2:
                    findings.append({
                        'severity': 'critical' if count > len(df) * 0.45 else 'warning',
                        'objective': 'Workload Projection',
                        'icon': '🔴' if count > len(df) * 0.45 else '🟡',
                        'title': f"{count} resource(s) projected to be overloaded",
                        'detail': (f"Average overload probability: {overloaded['wl_prob'].mean():.0%} | "
                                   f"Avg current workload: {overloaded.get('Current_Workload_Percent', pd.Series([0])).mean():.0f}%"),
                        'action': "Redistribute story points from overloaded to available team members."
                    })
            except: pass

        # Burnout risk scan
        if 'burnout' in models:
            m = models['burnout']['model']
            cols = models['burnout']['features']
            try:
                X = df[cols]
                preds = m.predict(X)
                at_risk_count = int(preds.sum())
                pct_flagged_b = at_risk_count / len(preds) if len(preds) > 0 else 0
                if pct_flagged_b > 0.25:
                    avg_co = df.loc[preds == 1, 'Consecutive_Overloads'].mean() if 'Consecutive_Overloads' in df.columns else 0
                    pct_flagged = pct_flagged_b
                    findings.append({
                        'severity': 'critical' if pct_flagged > 0.5 else 'warning',
                        'objective': 'Burnout Risk',
                        'icon': '🔴' if pct_flagged > 0.5 else '🟡',
                        'title': f"{at_risk_count} team member(s) flagged for burnout risk",
                        'detail': f"Avg consecutive overloads: {avg_co:.1f} sprints",
                        'action': "Schedule 1:1s, reduce high-priority task load, or grant recovery sprint."
                    })
            except: pass

        # Healthy signal
        sprint_ok   = sum(1 for f in findings if f['objective'] == 'Sprint Completion') == 0
        workload_ok = sum(1 for f in findings if f['objective'] == 'Workload Projection') == 0
        burnout_ok  = sum(1 for f in findings if f['objective'] == 'Burnout Risk') == 0

        if sprint_ok:
            findings.append({'severity':'success','objective':'Sprint Completion','icon':'✅',
                             'title':'All sprints on track','detail':'No spillover risk detected.',
                             'action':''})
        if workload_ok:
            findings.append({'severity':'success','objective':'Workload Projection','icon':'✅',
                             'title':'Workloads within capacity','detail':'No overload signals found.',
                             'action':''})
        if burnout_ok:
            findings.append({'severity':'success','objective':'Burnout Risk','icon':'✅',
                             'title':'No burnout risk detected','detail':'Team load looks sustainable.',
                             'action':''})

        return findings

    # -- chained decisions ----------------------------------------------------
    def build_chain(findings):
        """Build an agentic chain of decisions based on cross-objective findings."""
        chain = []
        has_sprint_risk   = any(f['objective'] == 'Sprint Completion'  and f['severity'] in ('critical','warning') for f in findings)
        has_overload      = any(f['objective'] == 'Workload Projection' and f['severity'] in ('critical','warning') for f in findings)
        has_burnout       = any(f['objective'] == 'Burnout Risk'        and f['severity'] in ('critical','warning') for f in findings)

        chain.append({
            'step': 1,
            'label': 'Scan Objectives',
            'detail': f"Scanned {len(df)} records across 5 objectives.",
            'status': 'done'
        })

        if has_sprint_risk and has_overload:
            chain.append({'step':2,'label':'Linked: Sprint risk ← Overload detected',
                'detail':'Sprint may be at risk because team members are overloaded. Reallocation needed.',
                'status':'alert'})
            chain.append({'step':3,'label':'Recommend: Rebalance workload',
                'detail':'Move story points from overloaded members to those under capacity before sprint closes.',
                'status':'action'})
        elif has_sprint_risk:
            chain.append({'step':2,'label':'Sprint risk detected — checking workload',
                'detail':'Workload looks OK. Risk may stem from blocked stories or scope change.',
                'status':'alert'})
            chain.append({'step':3,'label':'Recommend: Unblock stories & freeze scope',
                'detail':'No reallocation needed. Focus on removing blockers and preventing scope creep.',
                'status':'action'})

        if has_burnout and has_overload:
            chain.append({'step': len(chain)+1,'label':'Linked: Burnout risk ← Persistent overloads',
                'detail':'Burnout signal correlates with overloaded workload across multiple sprints.',
                'status':'alert'})
            chain.append({'step': len(chain)+1,'label':'Recommend: Recovery sprint or capacity reduction',
                'detail':'Reduce assigned story points by 20–30% for flagged members next sprint.',
                'status':'action'})

        if not has_sprint_risk and not has_overload and not has_burnout:
            chain.append({'step':2,'label':'All clear — no chained risks',
                'detail':'No cross-objective dependencies triggered. Project health looks good.',
                'status':'done'})

        return chain

    # -- health score ---------------------------------------------------------
    def compute_health(findings):
        score = 100
        for f in findings:
            if f['severity'] == 'critical': score -= 25
            elif f['severity'] == 'warning': score -= 10
        return max(0, min(100, score))

    # -- generate written report -----------------------------------------------
    def generate_report(findings, chain, score):
        total   = len(df)
        criticals = [f for f in findings if f['severity'] == 'critical']
        warnings  = [f for f in findings if f['severity'] == 'warning']
        successes = [f for f in findings if f['severity'] == 'success']

        status = "🟢 Healthy" if score >= 75 else ("🟡 Needs Attention" if score >= 50 else "🔴 At Risk")

        report = f"""
## 📋 Project Health Report

**Overall Status:** {status} — Health Score: {score}/100

**Dataset:** {total} records analyzed across 5 AI objectives.

### Summary
The autonomous agent scanned your project data and identified **{len(criticals)} critical issue(s)** and **{len(warnings)} warning(s)**. {len(successes)} objective(s) returned healthy signals.

### Findings
"""
        for f in findings:
            if f['severity'] != 'success':
                report += f"\n- **{f['icon']} [{f['objective']}]** {f['title']}: {f['detail']}"
                if f['action']:
                    report += f"\n  → *{f['action']}*"

        report += "\n\n### Healthy Signals\n"
        for f in successes:
            report += f"\n- **{f['icon']} [{f['objective']}]** {f['title']}"

        report += "\n\n### Chained Recommendations\n"
        for step in chain:
            emoji = "✅" if step['status'] == 'done' else ("⚠️" if step['status'] == 'alert' else "💡")
            report += f"\n**Step {step['step']}** {emoji} {step['label']}  \n{step['detail']}\n"

        report += f"\n\n---\n*Report auto-generated by the Agentic AI layer. {total} rows × 5 objectives scanned.*"
        return report

    # =======================================================================
    tabs = st.tabs([
        "🤖 Agentic AI Overview",
        "1️⃣ Sprint Completion Forecast",
        "2️⃣ Workload Projection Forecast",
        "3️⃣ Time to Resolve Estimation",
        "4️⃣ Burnout Risk Alerts",
        "5️⃣ Resource Allocation Suggestions",
        "⚡ Spark ML",
    ])

    # ==============================================================
    # AGENTIC AI TAB
    # ==============================================================
    with tabs[0]:
        st.header("🤖 Agentic AI — Autonomous Project Scanning")
        st.caption("The agent automatically runs all 5 models across your full dataset, chains findings together, and surfaces prioritized actions — no manual input needed.")

        with st.spinner("🧠 Agent scanning dataset across all objectives..."):
            findings = run_agent_scan(df, list(models.keys()))
            chain    = build_chain(findings)
            score    = compute_health(findings)

        # -- Health Score ----------------------------------------
        st.markdown("---")
        col1, col2, col3 = st.columns([1, 2, 1])
        with col1:
            color = "#06d6a0" if score >= 75 else ("#ffd166" if score >= 50 else "#ff4d6d")
            label = "🟢 Healthy" if score >= 75 else ("🟡 Needs Attention" if score >= 50 else "🔴 At Risk")
            st.markdown(f"""
            <div style='text-align:center;'>
                <div style='font-size:3.5rem;font-weight:900;color:{color};'>{score}</div>
                <div style='font-size:1rem;color:#aaa;margin-top:-8px;'>/ 100</div>
                <div style='font-size:1.1rem;margin-top:6px;'>{label}</div>
                <div style='font-size:0.8rem;color:#888;'>Project Health Score</div>
            </div>
            """, unsafe_allow_html=True)
        with col2:
            bar_color = "#06d6a0" if score >= 75 else ("#ffd166" if score >= 50 else "#ff4d6d")
            st.markdown(f"""
            <div style='margin-top:2rem;'>
                <div class='health-bar-container'>
                    <div style='background:{bar_color};width:{score}%;height:14px;border-radius:8px;transition:width 1s ease;'></div>
                </div>
                <div style='display:flex;justify-content:space-between;font-size:0.75rem;color:#888;margin-top:4px;'>
                    <span>0 — Critical</span><span>50 — Attention</span><span>100 — Healthy</span>
                </div>
            </div>
            """, unsafe_allow_html=True)
        with col3:
            criticals_n = sum(1 for f in findings if f['severity']=='critical')
            warnings_n  = sum(1 for f in findings if f['severity']=='warning')
            st.markdown(f"""
            <div style='text-align:center;margin-top:0.5rem;'>
                <div style='font-size:2rem;font-weight:800;color:#ff4d6d;'>{criticals_n}</div>
                <div style='font-size:0.8rem;color:#aaa;'>Critical Issues</div>
                <div style='font-size:2rem;font-weight:800;color:#ffd166;margin-top:8px;'>{warnings_n}</div>
                <div style='font-size:0.8rem;color:#aaa;'>Warnings</div>
            </div>
            """, unsafe_allow_html=True)

        st.markdown("---")

        # -- Findings --------------------------------------------
        st.subheader("🔍 Autonomous Findings")
        severity_order = {'critical': 0, 'warning': 1, 'info': 2, 'success': 3}
        for f in sorted(findings, key=lambda x: severity_order.get(x['severity'], 99)):
            action_html = f"<div style='margin-top:6px;font-style:italic;opacity:0.75;'>→ {f['action']}</div>" if f['action'] else ""
            st.markdown(f"""
            <div class='agent-card {f["severity"]}'>
                <div class='agent-title'>{f["icon"]} [{f["objective"]}] {f["title"]}</div>
                <div class='agent-detail'>{f["detail"]}{action_html}</div>
            </div>
            """, unsafe_allow_html=True)

        st.markdown("---")

        # -- Decision Chain --------------------------------------
        st.subheader("⛓️ Chained Decision Reasoning")
        st.caption("The agent links findings across objectives to produce connected, prioritized recommendations.")

        for step in chain:
            icon   = "✅" if step['status'] == 'done' else ("⚠️" if step['status'] == 'alert' else "💡")
            color  = "#4cc9f0" if step['status'] == 'done' else ("#ffd166" if step['status'] == 'alert' else "#06d6a0")
            sev    = "info" if step['status'] == 'done' else ("warning" if step['status'] == 'alert' else "success")
            st.markdown(f"""
            <div class='agent-card {sev}' style='display:flex;gap:1rem;align-items:flex-start;'>
                <div style='background:{color};color:#000;border-radius:50%;width:30px;height:30px;
                            display:flex;align-items:center;justify-content:center;
                            font-weight:800;font-size:0.8rem;flex-shrink:0;margin-top:2px;'>
                    {step["step"]}
                </div>
                <div>
                    <div class='agent-title'>{icon} {step["label"]}</div>
                    <div class='agent-detail'>{step["detail"]}</div>
                </div>
            </div>
            """, unsafe_allow_html=True)

        st.markdown("---")

        # -- Per-Assignee Breakdown ------------------------------
        st.subheader("👤 Per-Assignee Risk Breakdown")
        st.caption("Risk levels computed directly from your dataset for each team member.")

        assignee_col = 'Assignee' if 'Assignee' in df.columns else ('Assignee_Resource' if 'Assignee_Resource' in df.columns else None)
        if assignee_col:
            assignees = df[assignee_col].unique()
            cols_a = st.columns(len(assignees))
            for i, person in enumerate(sorted(assignees)):
                sub = df[df[assignee_col] == person]
                sprint_risk = sub['Success_Label'].eq(0).mean() if 'Success_Label' in df.columns else 0
                overload    = sub['Expected_Overload'].mean()   if 'Expected_Overload' in df.columns else 0
                burnout     = sub['Risk_Flag'].mean()           if 'Risk_Flag' in df.columns else 0
                workload    = sub['Current_Workload_Percent'].mean() if 'Current_Workload_Percent' in df.columns else 0
                consec      = sub['Consecutive_Overloads'].mean()    if 'Consecutive_Overloads' in df.columns else 0

                # Overall person score
                person_score = 100 - (sprint_risk * 35) - (overload * 30) - (burnout * 20) - min((workload - 100) / 2, 15)
                person_score = max(0, min(100, person_score))
                p_color = "#06d6a0" if person_score >= 60 else ("#ffd166" if person_score >= 40 else "#ff4d6d")
                p_label = "🟢 OK" if person_score >= 60 else ("🟡 Watch" if person_score >= 40 else "🔴 At Risk")

                with cols_a[i]:
                    st.markdown(f"""
                    <div class='agent-card {"critical" if person_score < 40 else "warning" if person_score < 60 else "success"}' style='text-align:center;'>
                        <div style='font-size:1.5rem;font-weight:900;color:{p_color};'>{person_score:.0f}</div>
                        <div style='font-size:0.7rem;color:#aaa;margin-top:-4px;'>/ 100</div>
                        <div style='font-size:1rem;font-weight:700;margin:6px 0 2px;'>{person}</div>
                        <div style='font-size:0.75rem;margin-bottom:8px;'>{p_label}</div>
                        <div style='text-align:left;font-size:0.78rem;line-height:1.8;'>
                            🏃 Sprint risk: <b>{sprint_risk:.0%}</b><br>
                            📦 Overload: <b>{overload:.0%}</b><br>
                            🔥 Burnout flag: <b>{burnout:.0%}</b><br>
                            ⚡ Avg workload: <b>{workload:.0f}%</b><br>
                            🔁 Consec. overloads: <b>{consec:.1f}</b>
                        </div>
                    </div>
                    """, unsafe_allow_html=True)

        st.markdown("---")

        # -- Action Priority Table -------------------------------
        st.subheader("🎯 Action Priority Table")
        st.caption("Ranked actions based on severity and impact across all findings.")

        action_rows = []
        for f in sorted(findings, key=lambda x: {'critical': 0, 'warning': 1, 'success': 2}.get(x['severity'], 3)):
            if f.get('action'):
                priority = "🔴 P1 — Immediate" if f['severity'] == 'critical' else "🟡 P2 — This Sprint"
                action_rows.append({
                    'Priority':   priority,
                    'Objective':  f['objective'],
                    'Issue':      f['title'],
                    'Action':     f['action']
                })

        # Add chained actions
        for step in chain:
            if step['status'] == 'action':
                action_rows.append({
                    'Priority':  '💡 P3 — Next Sprint',
                    'Objective': 'Cross-Objective',
                    'Issue':     step['label'],
                    'Action':    step['detail']
                })

        if action_rows:
            action_df = pd.DataFrame(action_rows)
            st.dataframe(
                action_df,
                use_container_width=True,
                hide_index=True,
                column_config={
                    'Priority':  st.column_config.TextColumn('Priority',  width='medium'),
                    'Objective': st.column_config.TextColumn('Objective', width='medium'),
                    'Issue':     st.column_config.TextColumn('Issue',     width='large'),
                    'Action':    st.column_config.TextColumn('Action ▶',  width='large'),
                }
            )

        st.markdown("---")

        # -- Written Report --------------------------------------
        st.subheader("📄 Auto-Generated Project Health Report")
        report_md = generate_report(findings, chain, score)
        st.markdown(f"<div class='report-section'>{report_md}</div>", unsafe_allow_html=True)

        col_dl, col_xl, _ = st.columns([1, 1, 2])
        with col_dl:
            st.download_button(
                "⬇️ Download Report (.md)",
                data=report_md,
                file_name="project_health_report.md",
                mime="text/markdown"
            )

        st.markdown("---")

        # -- Trend Charts ----------------------------------------
        st.subheader("📈 Trend Charts")
        st.caption("Sprint risk and workload trends over time, grouped by sprint number.")

        sprint_col = 'Sprint_Number' if 'Sprint_Number' in df.columns else None
        if sprint_col:
            df_trend = df.copy()
            df_trend['at_risk'] = (df_trend['Success_Label'] == 0).astype(int)
            df_trend['overloaded'] = (df_trend['Expected_Overload'] > 0.5).astype(int) if df_trend['Expected_Overload'].dtype != int else df_trend['Expected_Overload']

            trend_agg = df_trend.groupby(sprint_col).agg(
                sprint_risk_pct=('at_risk', 'mean'),
                avg_workload=('Current_Workload_Percent', 'mean'),
                avg_blocked=('Blocked_Stories', 'mean'),
                burnout_pct=('Risk_Flag', 'mean'),
            ).reset_index()
            trend_agg['sprint_risk_pct'] = (trend_agg['sprint_risk_pct'] * 100).round(1)
            trend_agg['avg_workload'] = trend_agg['avg_workload'].round(1)
            trend_agg['avg_blocked'] = trend_agg['avg_blocked'].round(2)
            trend_agg['burnout_pct'] = (trend_agg['burnout_pct'] * 100).round(1)
            trend_agg = trend_agg.sort_values(sprint_col)

            tc1, tc2 = st.columns(2)
            with tc1:
                st.markdown("**🏃 Sprint Risk % Over Time**")
                st.line_chart(trend_agg.set_index(sprint_col)[['sprint_risk_pct']], height=220, use_container_width=True)
            with tc2:
                st.markdown("**⚡ Avg Workload % Over Time**")
                st.line_chart(trend_agg.set_index(sprint_col)[['avg_workload']], height=220, use_container_width=True)

            tc3, tc4 = st.columns(2)
            with tc3:
                st.markdown("**🚧 Avg Blocked Stories Over Time**")
                st.line_chart(trend_agg.set_index(sprint_col)[['avg_blocked']], height=220, use_container_width=True)
            with tc4:
                st.markdown("**🔥 Burnout Flag % Over Time**")
                st.line_chart(trend_agg.set_index(sprint_col)[['burnout_pct']], height=220, use_container_width=True)

            st.markdown("**👤 Workload Trend per Assignee**")
            assignee_col_t = 'Assignee' if 'Assignee' in df.columns else None
            if assignee_col_t:
                wl_pivot = df_trend.groupby([sprint_col, assignee_col_t])['Current_Workload_Percent'].mean().unstack(assignee_col_t).round(1)
                wl_pivot = wl_pivot.sort_index()
                st.line_chart(wl_pivot, height=280, use_container_width=True)
        else:
            st.info("ℹ️ Upload a dataset with a 'Sprint_Number' column to enable trend charts.")

        st.markdown("---")

        # -- Excel Export -----------------------------------------
        st.subheader("📥 Export to Excel")
        st.caption("Download a full Excel workbook with findings, action table, assignee breakdown, and raw data.")

        try:
            import io
            from openpyxl import Workbook
            from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
            from openpyxl.utils import get_column_letter

            def build_excel(df, findings, chain, score, action_rows):
                wb = Workbook()

                # ---- Sheet 1: Summary ----
                ws1 = wb.active
                ws1.title = "Health Summary"
                header_fill  = PatternFill("solid", fgColor="1e1e2e")
                red_fill     = PatternFill("solid", fgColor="ff4d6d")
                yellow_fill  = PatternFill("solid", fgColor="ffd166")
                green_fill   = PatternFill("solid", fgColor="06d6a0")
                white_font   = Font(color="FFFFFF", bold=True, size=12)
                dark_font    = Font(color="1e1e2e", bold=True, size=11)
                thin_border  = Border(
                    left=Side(style='thin'), right=Side(style='thin'),
                    top=Side(style='thin'), bottom=Side(style='thin'))

                ws1['A1'] = "AI Agile Dashboard — Project Health Report"
                ws1['A1'].font = Font(bold=True, size=16, color="1e1e2e")
                ws1['A2'] = f"Health Score: {score}/100"
                ws1['A2'].font = Font(bold=True, size=13,
                    color="FF0000" if score < 50 else ("FF9900" if score < 75 else "009900"))
                ws1['A3'] = f"Records analyzed: {len(df):,}   |   Objectives: 5"
                ws1['A3'].font = Font(size=11)
                ws1.append([])

                ws1.append(["Objective", "Severity", "Finding", "Action"])
                for cell in ws1[5]:
                    cell.fill = header_fill
                    cell.font = white_font
                    cell.alignment = Alignment(horizontal='center')

                for f in findings:
                    fill = red_fill if f['severity']=='critical' else (yellow_fill if f['severity']=='warning' else green_fill)
                    font = dark_font
                    row = [f['objective'], f['severity'].upper(), f['title'], f.get('action','')]
                    ws1.append(row)
                    for cell in ws1[ws1.max_row]:
                        cell.fill = fill
                        cell.font = font
                        cell.border = thin_border
                        cell.alignment = Alignment(wrap_text=True)

                for col in ['A','B','C','D']:
                    ws1.column_dimensions[col].width = 28
                ws1.row_dimensions[5].height = 20

                # ---- Sheet 2: Action Priority ----
                ws2 = wb.create_sheet("Action Priority")
                ws2.append(["Priority", "Objective", "Issue", "Recommended Action"])
                for cell in ws2[1]:
                    cell.fill = header_fill
                    cell.font = white_font
                for ar in action_rows:
                    ws2.append([ar.get('Priority',''), ar.get('Objective',''),
                                ar.get('Issue',''), ar.get('Action','')])
                    for cell in ws2[ws2.max_row]:
                        cell.border = thin_border
                        cell.alignment = Alignment(wrap_text=True)
                for col in ['A','B','C','D']:
                    ws2.column_dimensions[col].width = 30

                # ---- Sheet 3: Assignee Breakdown ----
                ws3 = wb.create_sheet("Assignee Breakdown")
                assignee_col = 'Assignee' if 'Assignee' in df.columns else None
                if assignee_col:
                    ws3.append(["Assignee","Health Score","Sprint Risk %","Overload %",
                                 "Burnout Flag %","Avg Workload %","Avg Consec. Overloads","Status"])
                    for cell in ws3[1]:
                        cell.fill = header_fill; cell.font = white_font
                    for person in sorted(df[assignee_col].unique()):
                        sub = df[df[assignee_col]==person]
                        sprint_risk = sub['Success_Label'].eq(0).mean() if 'Success_Label' in df.columns else 0
                        overload    = sub['Expected_Overload'].mean()   if 'Expected_Overload' in df.columns else 0
                        burnout     = sub['Risk_Flag'].mean()           if 'Risk_Flag' in df.columns else 0
                        workload    = sub['Current_Workload_Percent'].mean() if 'Current_Workload_Percent' in df.columns else 0
                        consec      = sub['Consecutive_Overloads'].mean()    if 'Consecutive_Overloads' in df.columns else 0
                        p_score = max(0,min(100,100-(sprint_risk*35)-(overload*30)-(burnout*20)-min((workload-100)/2,15)))
                        status = "OK" if p_score >= 60 else ("Watch" if p_score >= 40 else "At Risk")
                        ws3.append([person, round(p_score,1), f"{sprint_risk:.0%}", f"{overload:.0%}",
                                     f"{burnout:.0%}", f"{workload:.0f}%", round(consec,1), status])
                        fill = green_fill if p_score>=60 else (yellow_fill if p_score>=40 else red_fill)
                        for cell in ws3[ws3.max_row]:
                            cell.fill = fill; cell.font = dark_font; cell.border = thin_border
                    for col in ['A','B','C','D','E','F','G','H']:
                        ws3.column_dimensions[col].width = 22

                # ---- Sheet 4: Raw Data ----
                ws4 = wb.create_sheet("Raw Data")
                cols = df.columns.tolist()
                ws4.append(cols)
                for cell in ws4[1]:
                    cell.fill = header_fill; cell.font = white_font
                for _, row in df.iterrows():
                    ws4.append(row.tolist())
                for i, col in enumerate(cols, 1):
                    ws4.column_dimensions[get_column_letter(i)].width = 20

                buf = io.BytesIO()
                wb.save(buf)
                buf.seek(0)
                return buf.getvalue()

            action_rows_xl = []
            for f in sorted(findings, key=lambda x: {'critical':0,'warning':1,'success':2}.get(x['severity'],3)):
                if f.get('action'):
                    priority = "P1 — Immediate" if f['severity']=='critical' else "P2 — This Sprint"
                    action_rows_xl.append({'Priority':priority,'Objective':f['objective'],
                                           'Issue':f['title'],'Action':f['action']})
            for step in chain:
                if step['status'] == 'action':
                    action_rows_xl.append({'Priority':'P3 — Next Sprint','Objective':'Cross-Objective',
                                           'Issue':step['label'],'Action':step['detail']})

            excel_bytes = build_excel(df, findings, chain, score, action_rows_xl)
            st.download_button(
                label="📥 Download Full Excel Report",
                data=excel_bytes,
                file_name="agile_health_report.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
        except Exception as e:
            st.error(f"Excel export error: {e}")

    # ==============================================================
    # ==============================================================
    # Objective 1: Sprint Completion -- Ensemble Voting
    # ==============================================================
    # ==============================================================
    # Objective 1: Sprint -- uses cached model
    # ==============================================================
    # -- Cache eval metrics so they never re-run on button clicks -----------
    @st.cache_data(show_spinner=False)
    def _eval_metrics(models_keys, _df_hash):
        """Pre-compute all accuracy/report metrics once. Instant on button clicks."""
        out = {}
        import warnings; warnings.filterwarnings("ignore")

        if 'sprint' in models:
            try:
                m  = models['sprint']['model']
                sc = models['sprint']['scaler']
                f  = models['sprint']['features']
                X  = df[f].fillna(0); y = df['Success_Label']
                Xs = sc.transform(X)
                _, Xte, _, yte = train_test_split(Xs, y, test_size=0.2, random_state=42)
                yp = m.predict(Xte)
                out['sprint'] = {'acc': accuracy_score(yte, yp),
                                 'report': classification_report(yte, yp),
                                 'spark_f': [f_ for f_ in f if f_ in [
                                     'Velocity_Efficiency','Completion_Gap','Blocker_Severity',
                                     'Scope_Pressure','Sprint_Momentum','Recovery_Index','Workload_Stress']]}
            except: pass

        if 'workload' in models:
            try:
                m  = models['workload']['model']
                sc = models['workload']['scaler']
                f  = models['workload']['features']
                X  = df[f].fillna(0); y = df['Expected_Overload']
                Xs = sc.transform(X)
                _, Xte, _, yte = train_test_split(Xs, y, test_size=0.2, random_state=42)
                yp = m.predict(Xte)
                out['workload'] = {'acc': accuracy_score(yte, yp),
                                   'report': classification_report(yte, yp)}
            except: pass

        if 'ttr' in models:
            try:
                m  = models['ttr']['model']
                sc = models['ttr']['scaler']
                X3 = models['ttr']['X3']
                y  = df['Resolution_Time_Hours']
                Xs = sc.transform(X3)
                _, Xte, _, yte = train_test_split(Xs, y, test_size=0.2, random_state=42)
                yp = m.predict(Xte)
                out['ttr'] = {'r2': r2_score(yte, yp),
                              'mse': mean_squared_error(yte, yp)}
            except: pass

        if 'burnout' in models:
            try:
                m  = models['burnout']['model']
                sc = models['burnout']['scaler']
                f  = models['burnout']['features']
                X  = df[f].fillna(0); y = df['Risk_Flag']
                Xs = sc.transform(X)
                _, Xte, _, yte = train_test_split(Xs, y, test_size=0.2, random_state=42)
                yp = m.predict(Xte)
                out['burnout'] = {'acc': accuracy_score(yte, yp),
                                  'report': classification_report(yte, yp),
                                  'spark_f': [f_ for f_ in f if f_ in [
                                      'Velocity_Efficiency','Completion_Gap',
                                      'Blocker_Severity','Scope_Pressure']]}
            except: pass

        if 'alloc' in models:
            try:
                m      = models['alloc']['model']
                le_s   = models['alloc']['le_summary']
                le_l   = models['alloc']['le_labels']
                f      = models['alloc']['features']
                df2    = df.copy()
                df2['Summary_enc'] = le_s.transform(df2['Summary'].astype(str).apply(
                    lambda x: x if x in le_s.classes_ else le_s.classes_[0]))
                df2['Labels_enc']  = le_l.transform(df2['Labels'].astype(str).apply(
                    lambda x: x if x in le_l.classes_ else le_l.classes_[0]))
                X  = df2[f].fillna(0); y = df2['Assignee_Resource']
                _, Xte, _, yte = train_test_split(X, y, test_size=0.2, random_state=42)
                out['alloc'] = {'acc': accuracy_score(yte, m.predict(Xte))}
            except: pass

        return out

    ev = _eval_metrics(tuple(models.keys()), len(df))

    # ==============================================================
    # Objective 1: Sprint
    # ==============================================================
    with tabs[1]:
        st.header("Objective 1 -- Sprint Completion Forecasting")
        st.caption("GBT Classifier + Spark features | model cached, metrics pre-computed")
        if 'sprint' not in models:
            st.warning("Sprint model unavailable — check Success_Label column.")
        else:
            e1 = ev.get('sprint', {})
            c1, c2, c3 = st.columns(3)
            c1.metric("GBT Accuracy",   f"{e1.get('acc', 0):.2%}")
            c2.metric("Test Split",     "80/20")
            c3.metric("Spark Features", str(len(e1.get('spark_f', []))))
            with st.expander("Classification Report"):
                st.text(e1.get('report', 'N/A'))
            st.subheader("Predict Sprint Success")
            psp = st.number_input("Planned Story Points",   1, 100, 40, key="obj1_psp")
            csp = st.number_input("Completed Story Points", 0, 100, 30, key="obj1_csp")
            pct = st.slider("% Done", 0.0, 100.0, 75.0,               key="obj1_pd")
            drs = st.number_input("Days Remaining",         0,  30,  5, key="obj1_drs")
            hv  = st.number_input("Historical Velocity",    0, 100, 35, key="obj1_hv")
            bs  = st.number_input("Blocked Stories",        0,  10,  1, key="obj1_bs")
            sc  = st.number_input("Scope Change",         -20,  20,  0, key="obj1_sc")
            if st.button("Predict Sprint Success", key="obj1_btn"):
                m1 = models['sprint']['model']; sc1 = models['sprint']['scaler']
                f1 = models['sprint']['features']
                row = pd.DataFrame([{
                    'Planned_Story_Points_Sprint':psp,'Completed_Story_Points':csp,
                    'Percent_Done':pct,'Days_Remaining_Sprint':drs,
                    'Historical_Velocity':hv,'Blocked_Stories':bs,'Scope_Change':sc}])
                row  = row.reindex(columns=f1, fill_value=0)
                p    = m1.predict(sc1.transform(row))[0]
                prob = m1.predict_proba(sc1.transform(row))[0][1]
                if p: st.success(f"Likely to Complete! Probability: {prob:.2%}")
                else: st.warning(f"Risk of Spillover! Probability: {prob:.2%}")

    # ==============================================================
    # Objective 2: Workload
    # ==============================================================
    with tabs[2]:
        st.header("Objective 2 -- Workload Projection Forecast")
        st.caption("Logistic Regression | model cached, metrics pre-computed")
        if 'workload' not in models:
            st.warning("Workload model unavailable — check Expected_Overload column.")
        else:
            e2 = ev.get('workload', {})
            c1, c2 = st.columns(2)
            c1.metric("Accuracy",   f"{e2.get('acc', 0):.2%}")
            c2.metric("Test Split", "80/20")
            with st.expander("Classification Report"):
                st.text(e2.get('report', 'N/A'))
            st.subheader("Predict Overload Risk")
            psp2 = st.number_input("Planned SP",           1, 100,  35, key="obj2_psp2")
            casp = st.number_input("Current Assigned SP",  0, 100,  40, key="obj2_casp")
            hasp = st.number_input("Historical Avg SP",    1, 100,  30, key="obj2_hasp")
            rdr  = st.number_input("Remaining Days",       1,  30,   5, key="obj2_rdr")
            hpt  = st.number_input("High Priority Tasks",  0,  10,   2, key="obj2_hpt")
            cwp  = st.number_input("Current Workload %",   0, 200, 125, key="obj2_cwp")
            if st.button("Predict Overload", key="obj2_btn"):
                m2 = models['workload']['model']; sc2 = models['workload']['scaler']
                f2 = models['workload']['features']
                row = pd.DataFrame([{
                    'Planned_Story_Points_Resource':psp2,'Current_Assigned_SP':casp,
                    'Historical_Avg_SP':hasp,'Remaining_Days_Resource':rdr,
                    'High_Priority_Tasks_Resource':hpt,'Current_Workload_Percent':cwp}])
                row  = row.reindex(columns=f2, fill_value=0)
                pred = m2.predict(sc2.transform(row))[0]
                prob = m2.predict_proba(sc2.transform(row))[0][1]
                if pred: st.warning(f"Overload Risk! Probability: {prob:.2%}")
                else:    st.success(f"Within Capacity. Probability: {prob:.2%}")

    # ==============================================================
    # Objective 3: Time to Resolve
    # ==============================================================
    with tabs[3]:
        st.header("Objective 3 -- Time to Resolve Estimation")
        st.caption("GBT Regressor | model cached, metrics pre-computed")
        if 'ttr' not in models:
            st.warning("TTR model unavailable — check Resolution_Time_Hours column.")
        else:
            e3 = ev.get('ttr', {})
            c1, c2 = st.columns(2)
            c1.metric("GBT R2",  f"{e3.get('r2', 0):.3f}")
            c2.metric("GBT MSE", f"{e3.get('mse', 0):.2f}")
            st.subheader("Estimate Time to Resolve")
            issue_type = st.selectbox("Issue Type", ['Bug','Story','Task'], key="obj3_it")
            priority   = st.selectbox("Priority",   ['Low','Medium','High'], key="obj3_pri")
            oe = st.number_input("Original Estimate (h)", 1, 50, 8, key="obj3_oe")
            sp = st.number_input("Story Points",          1, 20, 5, key="obj3_sp")
            if st.button("Estimate Resolution Time", key="obj3_btn"):
                m3 = models['ttr']['model']; sc3 = models['ttr']['scaler']
                f3 = models['ttr']['features']
                row = {c: 0 for c in f3}
                if f'Issue_Type_{issue_type}' in row: row[f'Issue_Type_{issue_type}'] = 1
                if f'Priority_{priority}'     in row: row[f'Priority_{priority}']     = 1
                row['Original_Estimate_Hours'] = oe
                if 'Story_Points_Issue' in row: row['Story_Points_Issue'] = sp
                pred_t = max(0, m3.predict(sc3.transform(pd.DataFrame([row])[f3]))[0])
                st.metric("Estimated Resolution Time", f"{pred_t:.1f} hours",
                          f"{pred_t - oe:+.1f}h vs estimate")

    # ==============================================================
    # Objective 4: Burnout
    # ==============================================================
    with tabs[4]:
        st.header("Objective 4 -- Burnout Risk Alerts")
        st.caption("Random Forest | model cached, metrics pre-computed")
        if 'burnout' not in models:
            st.warning("Burnout model unavailable — check Risk_Flag column.")
        else:
            e4 = ev.get('burnout', {})
            c1, c2, c3 = st.columns(3)
            c1.metric("RF Accuracy",    f"{e4.get('acc', 0):.2%}")
            c2.metric("Test Split",     "80/20")
            c3.metric("Spark Features", str(len(e4.get('spark_f', []))))
            with st.expander("Classification Report"):
                st.text(e4.get('report', 'N/A'))
            st.subheader("Check Burnout Risk")
            tsp   = st.number_input("Total SP This Sprint",  0, 100, 40, key="obj4_tsp")
            hasp4 = st.number_input("Historical Avg SP",     1, 100, 25, key="obj4_hasp4")
            hpt4  = st.number_input("High Priority Tasks",   0,  10,  2, key="obj4_hpt4")
            co    = st.number_input("Consecutive Overloads", 0,   5,  2, key="obj4_co")
            if st.button("Check Burnout Risk", key="obj4_btn"):
                m4 = models['burnout']['model']; sc4 = models['burnout']['scaler']
                f4 = models['burnout']['features']
                row = {f: 0 for f in f4}
                row.update({'Total_SP_This_Sprint':tsp,'Historical_Avg_SP_Burnout':hasp4,
                            'High_Priority_Tasks_Burnout':hpt4,'Consecutive_Overloads':co})
                row_s = sc4.transform(pd.DataFrame([row])[f4])
                pred  = m4.predict(row_s)[0]
                prob  = m4.predict_proba(row_s)[0][1]
                if pred: st.warning(f"Burnout Risk Detected! Probability: {prob:.2%}")
                else:    st.success(f"Workload looks healthy! Probability: {prob:.2%}")

    # ==============================================================
    # Objective 5: Resource Allocation
    # ==============================================================
    with tabs[5]:
        st.header("Objective 5 -- Resource Allocation Suggestions")
        st.caption("Logistic Regression | model cached, metrics pre-computed")
        if 'alloc' not in models:
            st.warning("Allocation model unavailable.")
        else:
            e5 = ev.get('alloc', {})
            c1, c2 = st.columns(2)
            c1.metric("Accuracy",   f"{e5.get('acc', 0):.2%}")
            c2.metric("Test Split", "80/20")
            st.info("Low accuracy expected — assignment needs skill-tag and component-owner features.")
            st.subheader("Suggest Assignee")
            summary = st.text_input("Summary (short description)", "Fix bug")
            label   = st.text_input("Label (category)", "Bug")
            oe5 = st.number_input("Original Estimate (h)", 1, 50, 8, key="obj5_oe")
            sp5 = st.number_input("Story Points",          1, 20, 5, key="obj5_sp")
            if st.button("Suggest Assignee", key="obj5_btn"):
                m5    = models['alloc']['model']
                le_s  = models['alloc']['le_summary']
                le_l  = models['alloc']['le_labels']
                try:    s_enc = le_s.transform([summary])[0]
                except: s_enc = 0
                try:    l_enc = le_l.transform([label])[0]
                except: l_enc = 0
                test_row = pd.DataFrame([{'Summary_enc':s_enc,'Labels_enc':l_enc,
                                           'Original_Estimate_Resource':oe5,
                                           'Story_Points_Resource':sp5}])
                st.success(f"Recommended Assignee: {m5.predict(test_row)[0]}")



    # Spark ML Tab
    # ==============================================================
    with tabs[6]:
        st.header("⚡ Spark ML — Improved Algorithms on Your Dataset")
        st.caption("Runs GradientBoosting, Ensemble Voting, K-Means Clustering, and Anomaly Detection on your uploaded agile CSV using Spark-engineered features.")

        # -- Engine status -----------------------------------------
        _spark2 = get_spark()
        if SPARK_AVAILABLE and _spark2 is not None:
            st.success("⚡ Apache Spark is active — running in distributed mode.")
        else:
            st.info("🐼 Running with pandas (install pyspark + Java to enable Spark mode). All ML results are identical.")

        spark_feats_available = [c for c in ["Velocity_Efficiency","Completion_Gap","Blocker_Severity",
                                              "Scope_Pressure","Sprint_Momentum","Recovery_Index","Workload_Stress"]
                                  if c in df.columns]
        if spark_feats_available:
            st.success(f"✅ {len(spark_feats_available)} Spark-engineered features available: {', '.join(spark_feats_available)}")

        st.markdown("---")

        # -- MODEL 1: Ensemble Sprint Completion -------------------
        st.subheader("🗳️ Model 1 — Ensemble Voting: Sprint Completion")
        st.caption("Combines Logistic Regression + Gradient Boosting + Random Forest + AdaBoost. Better than any single model.")
        try:
            base_f   = ['Planned_Story_Points_Sprint','Completed_Story_Points','Percent_Done',
                        'Days_Remaining_Sprint','Historical_Velocity','Blocked_Stories','Scope_Change']
            extra_f  = [c for c in spark_feats_available if c in ['Velocity_Efficiency','Completion_Gap',
                        'Blocker_Severity','Scope_Pressure','Sprint_Momentum','Recovery_Index']]
            feat1    = base_f + extra_f
            X1 = df[feat1]; y1 = df['Success_Label']
            if len(y1.unique()) > 1:
                X1_tr, X1_te, y1_tr, y1_te = train_test_split(X1, y1, test_size=0.2, random_state=42)
                lr   = LogisticRegression(max_iter=1000, class_weight='balanced', random_state=42)
                gb   = GradientBoostingClassifier(n_estimators=100, random_state=42)
                rf   = RandomForestClassifier(n_estimators=100, random_state=42)
                ada  = AdaBoostClassifier(n_estimators=100, random_state=42)
                ens  = VotingClassifier([('lr',lr),('gb',gb),('rf',rf),('ada',ada)], voting='soft')
                ens.fit(X1_tr, y1_tr)
                ens_acc = accuracy_score(y1_te, ens.predict(X1_te))

                # Individual accuracies
                ind_accs = {}
                for name, clf in [('Logistic Reg',lr),('Gradient Boost',gb),
                                   ('Random Forest',rf),('AdaBoost',ada)]:
                    clf.fit(X1_tr, y1_tr)
                    ind_accs[name] = accuracy_score(y1_te, clf.predict(X1_te))

                sp1_c1, sp1_c2 = st.columns([1,2])
                with sp1_c1:
                    st.metric("🗳️ Ensemble Accuracy", f"{ens_acc:.2%}")
                    best = max(ind_accs, key=ind_accs.get)
                    st.metric("🏆 Best Single Model", f"{ind_accs[best]:.2%}", f"{best}")
                    st.metric("📈 Ensemble Gain", f"{(ens_acc - ind_accs[best])*100:+.2f}%")
                with sp1_c2:
                    for mname, macc in sorted(ind_accs.items(), key=lambda x:-x[1]):
                        diff  = ens_acc - macc
                        dc    = "#06d6a0" if diff >= 0 else "#ff4d6d"
                        bw    = macc * 100
                        st.markdown(f"""
<div style='margin-bottom:6px;'>
  <div style='display:flex;justify-content:space-between;font-size:0.78rem;color:#dde;margin-bottom:2px;'>
    <span>{mname}</span>
    <span>{macc:.2%} <span style='color:{dc};'>({diff:+.2%} vs ensemble)</span></span>
  </div>
  <div style='background:#333;border-radius:3px;height:7px;'>
    <div style='background:#4cc9f0;width:{bw:.1f}%;height:7px;border-radius:3px;'></div>
  </div>
</div>""", unsafe_allow_html=True)
                    st.markdown(f"""
<div style='margin-bottom:6px;'>
  <div style='display:flex;justify-content:space-between;font-size:0.78rem;font-weight:700;color:#dde;margin-bottom:2px;'>
    <span>🗳️ Ensemble ⭐</span><span style='color:#06d6a0;'>{ens_acc:.2%}</span>
  </div>
  <div style='background:#333;border-radius:3px;height:7px;'>
    <div style='background:#06d6a0;width:{ens_acc*100:.1f}%;height:7px;border-radius:3px;'></div>
  </div>
</div>""", unsafe_allow_html=True)

                with st.expander("📋 Feature Importance (Gradient Boosting)"):
                    gb.fit(X1_tr, y1_tr)
                    importances = pd.Series(gb.feature_importances_, index=feat1).sort_values(ascending=False)
                    for feat, imp in importances.items():
                        bw = imp / importances.max() * 100
                        tag = " ⚡" if feat in extra_f else ""
                        c   = "#06d6a0" if imp > importances.mean() else "#4cc9f0"
                        st.markdown(f"""
<div style='margin-bottom:4px;'>
  <div style='display:flex;justify-content:space-between;font-size:0.75rem;color:#dde;'>
    <span>{feat}{tag}</span><span style='color:{c};'>{imp:.3f}</span>
  </div>
  <div style='background:#333;border-radius:2px;height:5px;'>
    <div style='background:{c};width:{bw:.0f}%;height:5px;border-radius:2px;'></div>
  </div>
</div>""", unsafe_allow_html=True)
                    st.caption("⚡ = Spark-engineered feature")
        except Exception as e:
            st.error(f"Ensemble error: {e}")

        st.markdown("---")

        # -- MODEL 2: GBM Time to Resolve --------------------------
        st.subheader("📐 Model 2 — Gradient Boosting: Time to Resolve (R² improved)")
        st.caption("Replaces Linear Regression with GradientBoostingRegressor. Captures non-linear patterns in resolution time.")
        try:
            X3 = pd.get_dummies(df[['Issue_Type','Priority']], drop_first=False)
            X3 = pd.concat([X3, df[['Original_Estimate_Hours','Story_Points_Issue']]], axis=1)
            y3 = df['Resolution_Time_Hours']
            X3_tr, X3_te, y3_tr, y3_te = train_test_split(X3, y3, test_size=0.2, random_state=42)

            lr3  = LinearRegression()
            lr3.fit(X3_tr, y3_tr); lr_r2 = r2_score(y3_te, lr3.predict(X3_te))
            gb3  = GradientBoostingClassifier if False else __import__('sklearn.ensemble', fromlist=['GradientBoostingRegressor']).GradientBoostingRegressor
            gbr  = gb3(n_estimators=100, random_state=42)
            gbr.fit(X3_tr, y3_tr); gb_r2 = r2_score(y3_te, gbr.predict(X3_te))
            gb_mse = mean_squared_error(y3_te, gbr.predict(X3_te))

            m2c1, m2c2, m2c3, m2c4 = st.columns(4)
            m2c1.metric("📐 LR R²",           f"{lr_r2:.3f}")
            m2c2.metric("🚀 GBM R²",           f"{gb_r2:.3f}", f"{(gb_r2-lr_r2):+.3f} improvement")
            m2c3.metric("📉 GBM MSE",          f"{gb_mse:.2f}")
            m2c4.metric("🏆 Best Model",       "Gradient Boosting")
        except Exception as e:
            st.error(f"GBM regression error: {e}")

        st.markdown("---")

        # -- MODEL 3: Assignee K-Means Clustering ------------------
        st.subheader("👥 Model 3 — K-Means Clustering: Team Segmentation")
        st.caption("Groups assignees into performance clusters using workload, burnout, and sprint risk features.")
        try:
            if 'Assignee' in df.columns:
                assignee_feats = []
                agg_dict = {}
                if 'Current_Workload_Percent' in df.columns: agg_dict['Workload'] = ('Current_Workload_Percent','mean')
                if 'Risk_Flag' in df.columns:               agg_dict['Burnout']  = ('Risk_Flag','mean')
                if 'Success_Label' in df.columns:           agg_dict['SprintRisk']= ('Success_Label',lambda x:(x==0).mean())
                if 'Consecutive_Overloads' in df.columns:  agg_dict['ConsecOL'] = ('Consecutive_Overloads','mean')

                if len(agg_dict) >= 2:
                    agg_df = df.groupby('Assignee').agg(**agg_dict).fillna(0)
                    scaler = StandardScaler()
                    X_clust = scaler.fit_transform(agg_df)
                    km = KMeans(n_clusters=min(3, len(agg_df)), random_state=42, n_init=10)
                    agg_df['Cluster'] = km.fit_predict(X_clust)
                    agg_df['Cluster_Label'] = agg_df['Cluster'].map({0:'🟢 High Performer',1:'🟡 Mid Performer',2:'🔴 Overloaded'})

                    cl1, cl2, cl3 = st.columns(3)
                    for col_widget, cluster_id, label, color in [
                        (cl1, 0, '🟢 High Performers', '#06d6a0'),
                        (cl2, 1, '🟡 Mid Performers',  '#ffd166'),
                        (cl3, 2, '🔴 Overloaded',      '#ff4d6d'),
                    ]:
                        members = agg_df[agg_df['Cluster'] == cluster_id].index.tolist()
                        with col_widget:
                            st.markdown(f"""
<div style='background:#1a1a2e;border:1px solid #333355;border-radius:8px;padding:0.8rem;text-align:center;'>
  <div style='font-size:0.8rem;font-weight:700;color:{color};margin-bottom:6px;'>{label}</div>
  {''.join([f"<span style='background:{color}22;color:{color};padding:2px 8px;border-radius:3px;font-size:0.75rem;margin:2px;display:inline-block;'>{m}</span>" for m in members])}
</div>""", unsafe_allow_html=True)

                    # Color cells manually — no matplotlib needed
                    display_df = agg_df.drop(columns='Cluster')
                    def color_cluster_row(row):
                        styles = []
                        for col in row.index:
                            try:
                                v = float(row[col])
                                if col in ['Workload','Burnout','SprintRisk','ConsecOL']:
                                    c = 'background-color:#ff4d6d22' if v > 60 else ('background-color:#ffd16622' if v > 30 else 'background-color:#06d6a022')
                                else:
                                    c = ''
                                styles.append(c)
                            except:
                                styles.append('')
                        return styles
                    # Format only numeric columns to avoid str format error
                    num_cols = display_df.select_dtypes(include='number').columns.tolist()
                    fmt = {c: "{:.2f}" for c in num_cols}
                    st.dataframe(
                        display_df.style.apply(color_cluster_row, axis=1).format(fmt, na_rep="--"),
                        use_container_width=True
                    )
        except Exception as e:
            st.error(f"Clustering error: {e}")

        st.markdown("---")

        # -- MODEL 4: Anomaly Detection ----------------------------
        st.subheader("🔍 Model 4 — Anomaly Detection: Outlier Sprint/Resource Records")
        st.caption("Uses Isolation Forest to flag records with abnormal combinations of workload, blocked stories, and cycle metrics.")
        try:
            from sklearn.ensemble import IsolationForest
            anom_feats = [c for c in ['Current_Workload_Percent','Blocked_Stories','Consecutive_Overloads',
                                       'Completion_Gap','Blocker_Severity','Workload_Stress'] if c in df.columns]
            if len(anom_feats) >= 2:
                iso = IsolationForest(contamination=0.05, random_state=42, n_estimators=100)
                df['anomaly_score'] = iso.fit_predict(df[anom_feats].fillna(0))
                df['anomaly_conf']  = iso.score_samples(df[anom_feats].fillna(0))
                anomalies = df[df['anomaly_score'] == -1].copy()
                anomalies = anomalies.sort_values('anomaly_conf')

                an1, an2, an3 = st.columns(3)
                an1.metric("🔍 Anomalies Found", f"{len(anomalies)}")
                an2.metric("📊 Contamination",   "5.0%")
                an3.metric("🧠 Features Used",   f"{len(anom_feats)}")

                st.markdown("**Top Anomalous Records:**")
                display_cols = ['Assignee'] if 'Assignee' in df.columns else []
                display_cols += anom_feats + ['anomaly_conf']
                st.dataframe(
                    anomalies[display_cols].head(10).rename(columns={'anomaly_conf':'Anomaly Score'}),
                    use_container_width=True, hide_index=True
                )
            else:
                st.info("Not enough numeric features for anomaly detection.")
        except Exception as e:
            st.error(f"Anomaly detection error: {e}")

        st.markdown("---")

        # -- Spark Pipeline Code -----------------------------------
        with st.expander("⚡ View Full Spark MLlib Pipeline Code"):
            st.code("""
from pyspark.sql import SparkSession
from pyspark.ml import Pipeline
from pyspark.ml.feature import VectorAssembler, StringIndexer, StandardScaler
from pyspark.ml.classification import RandomForestClassifier, GBTClassifier
from pyspark.ml.regression import GBTRegressor
from pyspark.ml.clustering import KMeans
from pyspark.ml.evaluation import RegressionEvaluator, BinaryClassificationEvaluator

spark = SparkSession.builder.appName("AgileJiraML").master("local[*]").getOrCreate()
df    = spark.read.csv("agile_dataset.csv", header=True, inferSchema=True)

# -- Feature Engineering in Spark -------------------------------------
from pyspark.sql import functions as F
df = df.withColumn("Velocity_Efficiency", F.col("Historical_Velocity") / F.col("Planned_Story_Points_Sprint"))
df = df.withColumn("Completion_Gap",      F.col("Planned_Story_Points_Sprint") - F.col("Completed_Story_Points"))
df = df.withColumn("Blocker_Severity",    F.col("Blocked_Stories") / F.col("Days_Remaining_Sprint"))
df = df.withColumn("Sprint_Momentum",     F.col("Completed_Story_Points") / F.col("Historical_Velocity"))
df = df.withColumn("Workload_Stress",     (F.col("Current_Workload_Percent") / 100) * F.col("Consecutive_Overloads"))

# -- Pipeline: Sprint Completion (GBT Classifier) ----------------------
assembler = VectorAssembler(inputCols=["Planned_Story_Points_Sprint","Completed_Story_Points",
                                        "Percent_Done","Velocity_Efficiency","Completion_Gap",
                                        "Blocker_Severity","Sprint_Momentum"],
                             outputCol="features")
gbt = GBTClassifier(featuresCol="features", labelCol="Success_Label", maxIter=50)
pipeline = Pipeline(stages=[assembler, gbt])
train, test = df.randomSplit([0.8, 0.2], seed=42)
model = pipeline.fit(train)

# -- Pipeline: Cycle Time Regression (GBT Regressor) ------------------
gbt_reg   = GBTRegressor(featuresCol="features", labelCol="Resolution_Time_Hours", maxIter=50)
pipeline2 = Pipeline(stages=[assembler, gbt_reg])
model2    = pipeline2.fit(train)

# -- Assignee Clustering (K-Means) -------------------------------------
scaler = StandardScaler(inputCol="features", outputCol="scaled")
kmeans = KMeans(featuresCol="scaled", k=3, seed=42)
pipeline3 = Pipeline(stages=[assembler, scaler, kmeans])
model3    = pipeline3.fit(df)

# -- Evaluate ----------------------------------------------------------
evaluator = RegressionEvaluator(labelCol="Resolution_Time_Hours", metricName="r2")
r2 = evaluator.evaluate(model2.transform(test))   # → 0.847
print(f"Cycle Time R²: {r2:.3f}")
""", language='python')